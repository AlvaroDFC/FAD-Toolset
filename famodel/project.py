"""Project class for FAModel, containing information and key methods for
the site information and design information that make up a project."""

import os
import numpy as np
import matplotlib.pyplot as plt

import moorpy as mp
from moorpy.helpers import set_axes_equal, loadLineProps
from moorpy import helpers
import yaml
from copy import deepcopy
from string import ascii_lowercase as ascii_l
from matplotlib.colors import BoundaryNorm
from matplotlib.cm import ScalarMappable
import matplotlib.patches as mpatches
try: 
    import raft as RAFT
except:
    pass

#from shapely.geometry import Point, Polygon, LineString
import famodel.seabed_tools as sbt
from famodel.mooring.mooring import Mooring
from famodel.platform.platform import Platform
from famodel.anchors.anchor import Anchor
from famodel.mooring.connector import Connector, Section
from famodel.substation.substation import Substation
from famodel.cables.cable import Cable
from famodel.cables.dynamic_cable import DynamicCable
from famodel.cables.static_cable import StaticCable
from famodel.cables.cable_properties import getCableProps, getBuoyProps, loadCableProps,loadBuoyProps
from famodel.cables.components import Joint, Jtube
from famodel.platform.fairlead import Fairlead
from famodel.turbine.turbine import Turbine
from famodel.famodel_base import Node, Edge, rotationMatrix
from ._project_helper import _build_array_cable_instances, _build_bathy_grid_and_depth, _build_boundary, _build_cable_instances_from_cables_section, _build_marine_growth, _build_platforms_instance, _build_soil, _parse_array, _parse_array, _parse_all_cable_info, _parse_all_mooring_info, \
                            _parse_anchor_data,_parse_platform_data, _parse_topsides, _build_fairleads_list, _build_jtubes_list, \
                            _build_turbine_and_substation_instances, _build_mooring_anchor_instance_nonshared, _build_raft_dict,\
                            _nearest_soil_names_on_bathymetry_grid, _build_soil_facecolors

# Import select required helper functions
from famodel.helpers import (check_headings, head_adjust, getCableDD, getDynamicCables, 
                            getMoorings, getAnchors, getFromDict, cleanDataTypes, 
                            getStaticCables, getCableDesign, m2nm, loadYAML, 
                            configureAdjuster, route_around_anchors, attachFairleads,
                            calc_heading, calc_midpoint, compareDicts)


class Project():
    '''
    The overall object that defines a floating array for analysis and design 
    purposes. Its main model component is the RAFT model but it also includes
    more general information such as seabed info, metocean info, and...
    
    Ideally, this class can function even if RAFT and MoorPy are not loaded,
    at least for some high-level site processing functions.
    
    '''
    
    def __init__(self, lon=0, lat=0, file=None, depth=202,raft=1):
        '''Initialize a Project. If input data is not provided, it will
        be empty and can be filled in later.
        
        Parameters
        ----------
        file : string or dict, optional
            Name of YAML file, or a python dictionary, containing input
            information describing the project, following the ontology.
        '''
        
        # ----- design information -----
        
        # higher-level design data structures
        self.nPtfm  = 0                                     # number of floating platforms
        self.nAnch  = 0                                     # number of anchors        
        self.coords = np.zeros([self.nPtfm+self.nAnch, 2])  # x-y coordinate table of platforms and anchors
        
        # more detailed design data structures for submodels
        self.array  = None                                  # RAFT model for dynamics analysis
        self.flow   = None                                  # FLORIS interface instance for wake analysis
        
        # Dictionaries describing the array, divided by structure type
        self.turbineList    = {}
        self.platformList   = {}
        self.mooringList    = {}                            # A dictionary of Mooring objects
        self.anchorList     = {}
        self.cableList      = {}                            # CableSystem
        self.substationList = {}
        self.midConnList    = {}                            # multi-line connectors
        
        # Dictionaries of component/product properties used in the array
        self.turbineTypes   = None                          # list of turbine designs (RAFT input file style)
        self.platformTypes  = None                          # list of platform designs (RAFT input file style, with rFair and zFair added)
        self.lineTypes      = None
        self.anchorTypes    = None
        self.cableTypes     = None
        
        
        # ----- site information -----
        self.lat0                   = lat                   # lattitude of site reference point [deg]
        self.lon0                   = lon                   # longitude of site reference point [deg]
        self.g                      = 9.81
        self.rho_water              = 1025                  # density of water (default to saltwater) [kg/m^3]
        self.rho_air                = 1.225                 # density of air [kg/m^3]
        self.mu_air                 = 1.81e-5               # dynamic viscosity of air [Pa*s]
        self.marine_growth          = None
        self.marine_growth_buoys    = None
        self.lineProps              = loadLineProps('default') # moorprops file dictionary

        # Project boundary (vertical stack of x,y coordinate pairs [m])
        self.boundary = np.zeros([0,2])
        
        # Exclusion zones
        self.exclusion = []
        
        # Seabed grid
        self.grid_x         = np.array([0])                 # coordinates of x grid lines [m]
        self.grid_y         = np.array([0])                 # coordinates of y grid lines [m]
        self.depth          = depth
        self.grid_depth     = np.array([[self.depth]])      # depth at each grid point [iy, ix]
        
        # soil parameters at each grid point
        self.soilProps  = {}
        self.soil_names = []
        self.soil_mode  = 0                # soil/anchor model level to use (0: none; 1: simple categories; 2: quantitative)
        self.soil_class = [["none"]]       # soil classification name ('clay', 'sand', or 'rock' with optional modifiers)
        self.soil_gamma = np.zeros((1,1))  # soil effective unit weight [kPa] (all soils)
        self.soil_Su0   = np.zeros((1,1))  # undrained shear strength at mudline [kPa] (clay soils)
        self.soil_K     = np.zeros((1,1))  # undrained shear strength gradient [kPa/m] (clay soils)
        self.soil_alpha = np.zeros((1,1))  # soil skin friction coefficient [-] (clay soils)
        self.soil_phi   = np.zeros((1,1))  # angle of internal friction [deg] (sand soils)
        self.soil_x     = None
        self.soil_y     = None
        
        # MoorPy system associated with the project
        self.ms  = None
        
        # RAFTDict associated with the project
        self.RAFTDict = None
        
        # ----- if an input file has been passed, load it -----
        if file:
            self.load(file,raft=raft)
    
    def load(self, info, raft=True):
        '''
        Load a full set of project information from a dictionary or 
        YAML file. This calls other methods for each part of it.
        
        Parameters
        ----------
        info : dict or filename
            Dictionary or YAML filename containing project info.
        '''
        # standard function to load dict if input is yaml
        if isinstance(info,str):
            # This is not a regular yaml loading. Through a constructor it allows passing other yaml files
            # by using !include filename.yaml syntax
            # TODO: project should be renamed to inputs or similar
            project = loadYAML(info)
            
            # save directory of main yaml for use when reading linked files
            # TODO: this should probably be a class property
            dir = os.path.dirname(os.path.abspath(info))  
            
        else:
            project = info
            dir = ''
        
        # look for site section
        # call load site method
        self.loadSite(project, dir=dir)
        
        # look for design section
        # call load design method
        self.loadDesign(project,raft=raft)

    # ----- Design loading/processing methods -----
    
    def loadDesign(self, d, raft=True):
        '''Load design information from a dictionary or YAML file
        (specified by input). This should be the design portion of
        the floating wind array ontology.'''
        
        print('Loading design')
        # standard function to load dict if input is yaml
        if not isinstance(d,dict):#if the input is not a dictionary, it is a yaml
            self.load(d)#load yaml into dictionary
        #d = 
        
        # ===== load FAM-specific model parts =====
        
        ################### ARRAY ###################
        array_data          = d.get('array', None)
        array_data_uniform  = d.get('uniform_array', None)

        arrayInfo = _parse_array(array_data, array_data_uniform)

        ################### CABLES ###################
        # Data that needs parsing
        # Convert the following to read them as dictionaries using d.get
        
        arrayCableInfo_data     = d.get('array_cables', None)
        dyn_cable_configs_data  = d.get('dynamic_cable_configs', None)
        cable_types_data        = d.get('cable_types', None)
        cable_appendages_data   = d.get('cable_appendages', None)

        # No parsing is required
        cableInfo = d.get('cables', {})

        arrayCableInfo, dyn_cable_configs, cable_types, cable_appendages = \
            _parse_all_cable_info(arrayCableInfo_data, 
                                  dyn_cable_configs_data, 
                                  cable_types_data, 
                                  cable_appendages_data)
        
        ################### MOORING ###################
        # ----- array mooring -----
        array_mooring_data              = d.get('array_mooring', None)
        mooring_systems_data            = d.get('mooring_systems', None)
        mooring_line_types_data         = d.get('mooring_line_types', None)
        mooring_connector_types_data    = d.get('mooring_connector_types', None)
        mooring_line_configs_data       = d.get('mooring_line_configs', None)

        arrayMooring, mSystems, self.lineTypes, self.lineProps, connectorTypes, lineConfigs = \
            _parse_all_mooring_info(self.lineProps,
                                    array_mooring_data, 
                                    mooring_systems_data, 
                                    mooring_line_types_data, 
                                    mooring_connector_types_data, 
                                    mooring_line_configs_data, 
                                    arrayInfo)

        ################### ANCHORS ###################
        anchor_data         = d.get('array_mooring',{}).get('anchor_data',None)
        anchor_keys         = d.get('array_mooring',{}).get('anchor_keys',None)
        anchor_types_data   = d.get('anchor_types')

        arrayAnchor, self.anchorTypes = \
            _parse_anchor_data(anchor_data,
                               anchor_keys,
                               anchor_types_data)
        
        
        # ----- platforms -----
        platform_data   = d.get('platform', None)
        platforms_data  = d.get('platforms', None)

        self.platformTypes, RAFTDict = \
            _parse_platform_data(platform_data, 
                                 platforms_data)
        
        ################### TOPSIDES ###################
        topsides = d.get('topsides',[])

        self.turbineTypes, substations = _parse_topsides(topsides)
        
        # ----- set up dictionary for each individual mooring line, create anchor, mooring, and platform classes ----

        ########################### Build design ########################### 
        # NOTE: declaring variables outside cos they need to be used later on
        alph                    = list(ascii_l)
        jtube_by_platform       = {}
        fairlead_by_platform    = {} # dict of platform ids as keys and fairlead objects list as values         
        # TODO: move this as a post_init(arrayInfo, arrayMooring, arrayAnchor, arrayCableInfo, CableInfo, raft)
        # check that all necessary sections of design dictionary exist
        if arrayInfo:
            mct                     = 0 # counter for number of mooring lines. NOTE: not used
            # NOTE: create a platform instance for each row in the array table
            # TODO: consider moving all of variable[id] = var_aux inside of building methods to avoid empty declarations and if statements here. The building methods can check if the variable exists and create it if not, then add to it. This would also be more modular and allow for more flexible ordering of the building steps.
            for i in range(0, len(arrayInfo)):
                arrayInfo_row = arrayInfo[i]

                # This method will modify self.platformTypes, self.platformList in place as mutable objects
                platform    = _build_platforms_instance(arrayInfo_row, self.platformTypes, self.platformList)
            
                # # get index of platform from array table
                pfID        = int(arrayInfo_row['platformID']-1)   

                # Return a list of Fairlead instances for the platform
                pf_fairs    = _build_fairleads_list(self.platformTypes, pfID, platform)

                # Save the pf_fairs list in the fairlead_by_platform dictionary with the platform ID as the key     
                fairlead_by_platform[platform.id]   = pf_fairs # fairlead_by_platform seems unused
                
                # Return a list of J-tubes dictionaries for the platform 
                # NOTE: nor the original lines nor these ones fill anythng! --> Returns list of empty dicts
                pf_jtubes   = _build_jtubes_list(self.platformTypes, pfID, platform)

                # Save the pf_jtubes list in the jtube_by_platform dictionary with the platform ID as the key
                jtube_by_platform[platform.id]      = pf_jtubes

                # Create the turbine or substation instance and add to project platform list
                # self.turbineTypes, self.turbineList and self.substationList are also modified in place as mutable objects in this method
                turbine_or_substation_instance = \
                    _build_turbine_and_substation_instances(arrayInfo_row, 
                                                            i, 
                                                            platform, 
                                                            topsides, 
                                                            self.turbineList, 
                                                            self.turbineTypes, 
                                                            self.substationList)
                if turbine_or_substation_instance is not None:
                    platform.attach(turbine_or_substation_instance)

                # The goal of this portion is to create the list of mooring instances and anchor instances,
                # NOTE: the attachement of mooring and anchor to platforms is done inside the method below. Might change.
                mooringList, anchorList = \
                    _build_mooring_anchor_instance_nonshared(arrayInfo_row,
                                                            mSystems,
                                                            lineConfigs,
                                                            alph,
                                                            connectorTypes,
                                                            self.lineProps,
                                                            self.lineTypes,
                                                            platform,
                                                            self.platformTypes[pfID],
                                                            pf_fairs,
                                                            self.anchorTypes,
                                                            # the following are kwargs
                                                            rho_water=self.rho_water,
                                                            g=self.g,
                                                            depth=self.depth)
                self.mooringList.update(mooringList) 
                self.anchorList.update(anchorList)
                   
                # update position of platform, moorings, anchors
                platform.setPosition(r=platform.r, project=self)
                        
        
        # ----- set up dictionary for each shared mooring line or shared anchor, create mooring and anchor classes ----
        
        # create any shared mooring lines / lines connected to shared anchors / lines called out in arrayMooring table
        mooringList, anchorList = self._build_mooring_anchor_instance_shared(arrayMooring, arrayInfo,arrayAnchor, lineConfigs, connectorTypes, alph)

        # NOTE: in the original code, the anchor and mooring instances are not added to mooringList and anchorList?
        self.mooringList.update(mooringList) 
        self.anchorList.update(anchorList)
                
        # update all anchors
        self.updateAnchor()
                
        # ===== load Cables ======
        cableList, cable_reposition_args = _build_array_cable_instances(
            arrayCableInfo,
            dyn_cable_configs,
            cable_types,
            cable_appendages,
            self.platformList,
            jtube_by_platform,
            depth=self.depth,
            rho_water=self.rho_water,
            g=self.g,
            start_index=len(self.cableList),
        )

        # Update the cable list dictionary that belongs to the project class
        self.cableList.update(cableList)

        # reposition the cables now that all are created and attached
        for cable_id, cable in cableList.items():
            cable.reposition(
                project=self,
                **cable_reposition_args[cable_id]) 
                        
        # create any cables from cables section (this is a more descriptive cable format that may have routing etc)   
        cableList, cable_reposition_args = \
            _build_cable_instances_from_cables_section(cableInfo,
                                                        dyn_cable_configs,
                                                        cable_types,
                                                        cable_appendages,
                                                        self.platformList,
                                                        jtube_by_platform,
                                                        depth=self.depth,
                                                        rho_water=self.rho_water,
                                                        g=self.g,
                                                        start_index=len(self.cableList))

        self.cableList.update(cableList)

        # NOTE: we could merge both for-loops into one. Keeping them separate for now.
        for cable_id, cable in cableList.items():
            cable.reposition(
                project=self,
                **cable_reposition_args[cable_id])     
        
        # NOTE: not sure what this is doing, review in future version
        for pf in self.platformList.values():
            pf.setPosition(pf.r, project=self)

        # ===== load RAFT model parts =====                
        raft_dict = _build_raft_dict(raft,
                                RAFTDict,
                                self.turbineTypes,
                                arrayInfo,
                                d["site"],
                                depth=self.depth,
                                rho_water=self.rho_water,
                                rho_air=self.rho_air,
                                mu_air=self.mu_air)

        if raft and ("platforms" in raft_dict or "platform" in raft_dict):
            self.getRAFT(raft_dict) # NOTE: I don't think this is doing anything --> wrong it creates the self.array RAFT model

        self.RAFTDict = deepcopy(raft_dict)
            
    # ----- Site conditions processing functions -----

    def loadSite(self, yaml_input, dir=''):
        '''Load site information from a dictionary or YAML file
        (specified by input). This should be the site portion of
        the floating wind array ontology.

        Each of the keys in the yaml file (general, bathymetry, boundary, soil, marine_growth) is optional, 
        and if not provided, the project will keep the default values for the corresponding variables.
        This is done through the defaults_dict variable that is passed to the helper functions that 
        build each portion of the site information.
        
        yaml_input : passing the full yaml dictionary here
        dir : optional directory of main yaml file'''
        # standard function to load dict if input is yaml

        ###########
        site_dict = yaml_input.get('site', None)
        # site_dict = site # NOTE: to use before updating the load method
        if site_dict is None:
            print('[Info] No site information provided')
            return
        
        general         = site_dict.get('general', {})
        self.depth      = getFromDict(general, 'water_depth', default=self.depth)
        self.rho_water  = getFromDict(general, 'rho_water', default=1025.0)
        self.rho_air    = getFromDict(general, 'rho_air', default=1.225)
        self.mu_air     = getFromDict(general, 'mu_air', default=1.81e-5)

        # load bathymetry information, if provided
        defaults_dict = {'grid_x': self.grid_x,
                         'grid_y': self.grid_y,
                         'grid_depth': self.grid_depth}
        
        self.grid_x, self.grid_y, self.grid_depth = _build_bathy_grid_and_depth(site_dict, depth=self.depth, dir=dir, defaults_dict=defaults_dict)
  
        # Load project boundary, if provided, else return the default boundary (np.zeros([0,2]) 
        self.boundary = _build_boundary(site_dict, lat0=self.lat0, lon0=self.lon0, default_boundary=self.boundary)
 
        # Create a defaults dict in case these variables are not defined and we don't change the data type
        defaults_dict = {'soilProps': self.soilProps,
                         'soil_x': self.soil_x,
                         'soil_y': self.soil_y,
                         'soil_mode': self.soil_mode,
                         'soil_names': self.soil_names}
        
        self.soilProps, self.soil_x, self.soil_y, self.soil_names, self.soil_mode = _build_soil(site_dict, dir, defaults_dict)
 
        # load marine growth info NOTE: a defualts dictionary is not passed because these variables are initialized as None
        self.marine_growth, self.marine_growth_buoys = _build_marine_growth(site_dict)    

    # Helper functions

    def getDepthAtLocation(self, x, y, return_n=False):
        '''Compute the depth at a specified x-y location based on the
        bathymetry grid stored in the project. Setting return_n to 
        True will return depth and the seabed normal vector.
        '''
        
        z, n = sbt.getDepthFromBathymetry(x, y, self.grid_x, self.grid_y, 
                                          self.grid_depth)
        
        if return_n:
            return z, n
        else:
            return z


    def seabedIntersect(self, r, u):
        '''
        Compute location at which a ray (defined by coordinate i and direction
        vector u) crosses the seabed.
        
        Paramaters
        ----------
        r : float, array
            Absolute xyz coordinate of a point along the ray [m].
        u : float, array
            3D direction vector of the ray.

        Returns
        -------
        r_i: float, array
            xyz coordinate of the seabed intersection [m].
        '''
        
        # Initialize 
        r_i = np.array(r)  # calculated seabed intersect starts at the r provided
        ix_last = -1  # index of intersected grid panel (from previous iteration)
        iy_last = -1
        
        for i in range(10):  # iterate up to 10 times (only needed for fine grids)
            
            # Get depth of seabed beneath r_i, and index of grid panel
            depth, nvec, ix, iy = sbt.getDepthFromBathymetry(r_i[0], r_i[1], self.grid_x, 
                                       self.grid_y, self.grid_depth, index=True)
            
            #print(f" {i}  {depth:6.0f}   {ix}, {iy}   {r_i[0]:5.0f},{r_i[1]:5.0f},{r_i[2]:5.0f}")
            
            # Convergence check (if we're on the same grid panel, we're good)
            if ix==ix_last and iy==iy_last:
                break

            # Save some values for convergence checking
            ix_last = ix
            iy_last = iy
            r_last  = np.array(r_i)
            
            # vectors from r_i to a point on the bathymetry panel beneath it
            # w = np.array([r_anch[0], r_anch[1], -depth]) - r_i
            w = np.array([0, 0, -depth - r_i[2]]) # same!! as above
            
            # fraction along u where it crosses the seabed (can be greater than 1)
            fac = np.dot(nvec, w) / np.dot(nvec, u)
            
            r_i = r_i + u*fac  # updated seabed crossing estimate 

        
        return r_i

    # TODO: move to seabed tools
    def projectAlongSeabed(self, x, y):
        '''Project a set of x-y coordinates along a seabed surface (grid),
        returning the corresponding z coordinates.'''
        
        if len(x) == len(y):
            n = len(x)        
            z = np.zeros(n)   # z coordinate of each point [m]
            a = np.zeros(n)   # could also do slope (dz/dh)
            for i in range(n):
                z[i] = self.getDepthAtLocation(x[i], y[i])
        
        else:
            z = np.zeros([len(y), len(x)])
            for i in range(len(y)):
                for j in range(len(x)):
                    z[i,j] = self.getDepthAtLocation(x[j], y[i])
            
        return z

    # METHODS TO USE WITH ANCHOR TOOLS
    def loadSoil(self, filename=None, yaml=None, soil_mode='uniform', profile_source=None):
        '''
        Load geotechnical information from input file or YAML.
        Supports two soil modes: 'uniform' and 'layered'.

        Parameters
        ----------
        filename : str, optional
            Path to .txt/.dat file with soil labels/profile IDs and coordinates
        yaml : dict, optional
            Dictionary containing soil data and properties (used when filename is None)
        soil_mode : str
            Either 'uniform' or 'layered'
        profile_source : str, optional
            Path to YAML file with layered profile definitions (only used if soil_mode='layered')
        '''
        xs = None
        ys = None
        soil_names = None
        soilProps = None

        # Case 1: File input (grid + properties)
        if filename is not None:
            if filename.endswith('.shp'):
                raise ValueError("Shapefiles not supported in Project class")

            elif filename.endswith('.txt') or filename.endswith('.dat'):
                # Load label/profile_id grid
                xs, ys, soil_names = sbt.readBathymetryFile(filename, dtype=str)

                # Load soil properties
                soilProps = sbt.getSoilTypes(filename, soil_mode=soil_mode, profile_source=profile_source)

            if yaml:
                soilProps = yaml.get('soil_types', soilProps)  # allow overwriting via YAML

        # Case 2: YAML only (no filename)
        elif filename is None:
            if yaml:
                xs = yaml['x']
                ys = yaml['y']
                soil_names = yaml['type_array']
                raw_soil_types = yaml['soil_types']
        
                # Ensure all soil types have a 'layers' field
                soilProps = {}
                for key, entry in raw_soil_types.items():
                    if 'layers' in entry:
                        soilProps[key] = entry 
                    else:
                        # Wrap old flat format into single-layer profile (optional fallback)
                        layer = dict(entry)
                        layer.setdefault('top', 0)
                        layer.setdefault('bottom', 50)
                        layer.setdefault('soil_type', key)
                        soilProps[key] = {'layers': [layer]}
            else:
                print('[Warning] No soil input provided — using default values')
                xs = [0]
                ys = [0]
                soil_names = [['mud']]  # note: should be 2D to match grid structure
                soilProps = {
                    'mud': {'layers': [{
                        'soil_type': 'clay',
                        'top': 0, 'bottom': 50,
                        'gamma_top': 10, 'gamma_bot': 10,
                        'Su_top': 2.39, 'Su_bot': 59.39
                    }]},
                    'rock': {'layers': [{
                        'soil_type': 'rock',
                        'top': 0, 'bottom': 50,
                        'UCS_top': 5, 'UCS_bot': 5,
                        'Em_top': 7, 'Em_bot': 7
                    }]}
                }


        else:
            raise ValueError("Invalid combination of filename/yaml inputs")

        # --- Set defaults only for uniform mode (when values are missing) ---
        if soil_mode == 'uniform':
            for key, props in soilProps.items():
                props['Su0']   = getFromDict(props, 'Su0',   shape=-1, dtype=list, default=[2.39])
                props['k']     = getFromDict(props, 'k',     shape=-1, dtype=list, default=[1.41])
                props['alpha'] = getFromDict(props, 'alpha', shape=-1, dtype=list, default=[0.7])
                props['gamma'] = getFromDict(props, 'gamma', shape=-1, dtype=list, default=[8.7])
                props['phi']   = getFromDict(props, 'phi',   shape=-1, dtype=list, default=[0.0])
                props['UCS']   = getFromDict(props, 'UCS',   shape=-1, dtype=list, default=[7.0])
                props['Em']    = getFromDict(props, 'Em',    shape=-1, dtype=list, default=[50.0])

                # ensure no array-like leftovers
                for k, prop in props.items():
                    if hasattr(prop, '__array__'):
                        props[k] = np.array(prop)

        # --- Store to project ---
        self.soilProps = soilProps

        if xs is not None:
            self.soil_x = np.array(xs)
            self.soil_y = np.array(ys)
            self.soil_names = np.array(soil_names)

        self.soil_mode = soil_mode
        print(f"Loaded soilProps keys: {list(soilProps.keys())}")

        # --- Update anchor objects if available ---
        if self.anchorList:
            for anchor in self.anchorList.values():
                name, props = self.getSoilAtLocation(anchor.r[0], anchor.r[1])
                anchor.soilProps = {name: props}

    def getSoilAtLocation(self, x, y):
        '''
        Retrieve the soil information at a specific location, supporting both uniform and layered modes.

        Returns
        -------
        (str, dict or list): soil name or profile ID, and associated soil properties or layered profile
        '''
        self.profile_map = []
        
        if self.soil_x is not None:
            ix = np.argmin([abs(x - sx) for sx in self.soil_x])
            iy = np.argmin([abs(y - sy) for sy in self.soil_y])
            soil_id = self.soil_names[iy, ix]  # could be label or profile_id

            if self.soil_mode == 'uniform':
                soil_info = self.soilProps[soil_id]               
                if not self.profile_map:
                    self.convertUniformToLayered(default_layer=50.0)
                    
                # Replace with a single entry corresponding to this soil_id
                self.profile_map = [
                    next(e for e in self.profile_map if e['name'] == str(soil_id))]
                self.profile_map[0]['layers']
                
                return soil_id, soil_info
                
            elif self.soil_mode == 'layered':
                layers = self.soilProps[soil_id]  # list of layer dicts
                profile_entry = {'name': str(soil_id), 'layers': layers}
                self.profile_map.append(profile_entry)
                
                return soil_id, layers
                
            else:
                raise ValueError(f"Unknown soil_mode: {self.soil_mode}")

            print(f"[DEBUG] soil_id at location ({x}, {y}) is: {soil_id}")
            print(f"[DEBUG] Available soilProps keys: {list(self.soilProps.keys())}")
        
        else:
            raise ValueError("No soil grid defined")
            
    def convertUniformToLayered(self, default_layer=50.0):
        '''
        Converts self.soilProps (uniform format) into profile_map (layered format)
        using a default thickness and assuming uniform clay profile.
        Matches the structure of layered CPT-based soil profiles.
        '''
        self.profile_map = []

        for name, props in self.soilProps.items():
            name = str(name)

            gamma = float(props['gamma'][0])
            Su0   = float(props['Su0'][0])
            k     = float(props['k'][0])

            layer = {
                'soil_type': 'clay',
                'top': 0.0,
                'bottom': default_layer,
                'gamma_top': gamma,
                'gamma_bot': gamma,
                'Su_top': Su0,
                'Su_bot': Su0 + k*default_layer}

            profile_entry = {'name': name, 'layers': [layer]}
            self.profile_map.append(profile_entry)
            
    def convertLayeredToUniform(self):
        '''
        Converts self.profile_map (layered format) into soilProps (uniform format)
        assuming a single clay layer with linear Su(z) = Su0 + k*z.
        Matches the structure expected by uniform soil models.
        '''
        self.soilProps = {}

        for name, layers in self.profile_map.items():
            if not layers or len(layers) != 1:
                raise ValueError('convertLayeredToUniform only supports a single-layer profile')

            layer = layers[0]
            if str(layer.get('soil_type', '')).lower() != 'clay':
                raise ValueError('convertLayeredToUniform only supports clay')

            top = float(layer['top'])
            bot = float(layer['bottom'])
            Su_top = float(layer['Su_top'])
            Su_bot = float(layer['Su_bot'])
            gamma = float(layer['gamma_top'])  # gamma_top == gamma_bot in your format

            if bot <= top:
                raise ValueError('Invalid layer thickness (bottom <= top)')

            thickness = bot - top
            k = (Su_bot - Su_top)/thickness
            Su0 = Su_top - k*top

            self.soilProps[name] = {
                'gamma': [gamma],
                'Su0': [Su0],
                'k': [k]}

    # # ----- Anchor 
    def updateAnchor(self,anch='all',update_loc=True):
        #breakpoint()
        if anch == 'all':
            anchList = [anch for anch in self.anchorList.values()]
        elif isinstance(anch, Anchor):
            anchList = [anch]
            
        for anchor in anchList:
            if update_loc:
                att = next(iter(anchor.attachments.values()), None)
                if att['end'] in ['a', 'A', 0]:
                    anchor.r = att['obj'].rA
                else:
                    anchor.r = att['obj'].rB

            x,y = anchor.r[:2]
                
            anchor.r[2] = -self.getDepthAtLocation(x,y) # update depth
            for att in anchor.attachments.values():
                if att['end'] in ['a', 'A', 0]:
                    att['obj'].rA[2] = anchor.r[2]
                else:
                    att['obj'].rB[2] = anchor.r[2]
                
            if hasattr(self,'soil_x') and self.soil_x is not None:
                name, props = self.getSoilAtLocation(x,y) # update soil
                anchor.soilProps = {name:props}
            
    def setSoilAtLocation(self, anchor):
        name, props = self.getSoilAtLocation(anchor.r[0], anchor.r[1])
    
        # Add required metadata
        layer = dict(props)  # shallow copy of props
        layer['soil_type'] = name  # or force to 'clay'/'rock' if needed
        layer['top'] = props.get('top', 0)
        layer['bottom'] = props.get('bottom', 50)  
    
        # Wrap in expected profile_map format
        profile_map = [{'name': name, 'layers': [layer]}]
        anchor.setSoilProfile(profile_map)

   
    def setCableLayout(self):

        # 2-D

        # 3-D
        pass
    
    # ----- general design-related calculations -----
    
    def makeDistanceMatrix(self):
        '''Compute the distance matrix for an array of turbines. This matrix
        is filled with the horizontal distance between every turbine's 
        undisplaced position.
        '''
        
        dists = np.zeros([self.nPtfm, self.nPtfm])  # distance matrix
        
        for i in range(self.nPtfm):
            for j in range(self.nPtfm):
                delta = self.coords[i] - self.coords[j]
                dists[i,j] = np.diag(delta)
                dists[j,i] = dists[i,j]
                
        return dmat
    
    
    # ----- cable calculation methods -----
    
    def calcCableLength(self, cable):
        '''Calculates a cable's length based on its routing.
        '''

        # select cable
        
        # figure out cable length considering end coordinates and path
        
        return length
    
    
    def checkCableExclusions(self, cable):
        '''Checks whether a cable crosses over any exclusions
        or other out of bounds areas.
        '''

        # select cable
        
        # check its path against any exclusion areas or boundaries
        
        # make a list of any exclusion/nodes that it is too close to
        
        return score, list_of_violations
    


    def loadBoundary(self, filename):
        '''
        Load a lease area boundary for the project from an input file.
        
        Parameters
        ----------
        filename : path
            path/name of file containing bathymetry data (format TBD)
        '''
        
        # load data from file
        Xs, Ys = sbt.processBoundary(filename, self.lat0, self.lon0)
        
        self.setBoundary(Xs, Ys)


    def setBoundary(self, Xs, Ys):
        '''Set the boundaries of the project based on x-y polygon vertices.'''
        
        # check compatibility with project grid size
        
        # save as project boundaries
        self.boundary = np.vstack([[Xs[i],Ys[i]] for i in range(len(Xs))])
        # self.boundary = np.vstack([Xs, Ys])
        
        # if the boundary doesn't repeat the first vertex at the end, add it
        if not all(self.boundary[0,:] == self.boundary[-1,:]):
            self.boundary = np.vstack([self.boundary, self.boundary[0,:]])
        
        # figure out masking to exclude grid data outside the project boundary
        
    def setExclusionZone(self, Xs, Ys):
        '''
        Set exclusion zones of the project based on x-y verts
        '''
        self.exclusion.append(np.vstack([[Xs[i],Ys[i]] for i in range(len(Xs))]))
        
        # if the exclusion doesn't repeat the first vertex at the end, add it
        if not all(self.exclusion[-1][0,:] == self.exclusion[-1][-1,:]):
            self.exclusion[-1] = np.vstack([self.exclusion[-1], self.exclusion[-1][0,:]])
    
    def trimGrids(self, buffer=100):
        '''Trims bathymetry and soil grid information that is outside the
        project boundaries, for faster execution and plotting.'''

        # boundary extents
        xmin = np.min(self.boundary[:,0]) - buffer
        xmax = np.max(self.boundary[:,0]) + buffer
        ymin = np.min(self.boundary[:,1]) - buffer
        ymax = np.max(self.boundary[:,1]) + buffer
        
        # figure out indices to trim at if needed
        i_x1 = np.max((np.argmax(self.grid_x > xmin) - 1, 0))  # start x index
        i_y1 = np.max((np.argmax(self.grid_y > ymin) - 1, 0))  # start y index
        if xmax < np.max(self.grid_x):           
            i_x2 = np.max((np.argmin(self.grid_x < xmax) + 1, 0))  # end x index+1
        else:
            i_x2 = len(self.grid_x)
        if ymax < np.max(self.grid_y):           
            i_y2 = np.max((np.argmin(self.grid_y < ymax) + 1, 0))  # end y index+1
        else:
            i_y2 = len(self.grid_y)
        
        # trim things
        self.grid_x     = self.grid_x    [i_x1:i_x2]
        self.grid_y     = self.grid_y    [i_y1:i_y2]
        self.grid_depth = self.grid_depth[i_y1:i_y2, i_x1:i_x2]           
        
    def addCablesConnections(self,connDict,cableType_def='dynamic_cable_66',oss=False,
                             substation_r=[None],ss_id=200,id_method='location',
                             keep_old_cables=False, connect_ss=True, 
                             cableConfig=None, configType=0,heading_buffer=30,
                             route_anchors=False, adj_dir=1, 
                             consider_alternate_side=False):

        '''Adds cables and connects them to existing platforms/substations based on info in connDict
        Designed to work with cable optimization output designed by Michael Biglu

        Parameters
        ----------
        connDict : dict
            Connection dictionary that describes the cables to create and their connections
        cableType_def : str, optional
            Cable family name corresponding to a name in the CableProps_default file. Default is 'dynamic_cable_66'
        oss : bool, optional
            Controls whether to create an offshore substation object. Default is False.
        substation_r : list, optional
            x-y location of substation object to be created. Default is None
        id_method : str, optional
            Method to identify which platforms/substations to connect to each cable. Options are 'location' or 'id'.
            'location' option connects cable to platforms/substation that has the same location as the start and end coordinates listed 
            provided for that cable in the connDict. 'id' option connects cable to platforms/substation that has the same id 
            as the global ids listed for that cable in the connDict. Default is 'location'.
        keep_old_cables : bool, optional
            Controls whether to keep any existing cables in the project object or disconnect and remove all of them.
            Default is False.
        connect_ss : bool, optional
            Controls whether to add the cable connections to the substation
        cableConfig : dict, optional
            Dictionary listing details on cable configurations to apply to the cable objects. 
        configType : int, optional
            0 = default to dynamic-static-dynamic cables, 1 = default to suspended cable systems
        heading_buffer : float, optional
            Minimum buffer between moorings and cables (degrees). Default is 30
        route_anchors: bool, optional
            True=automatically route cables around anchors
        adj_dir: int, optional
            Control initial direction to adjust cable headings to avoid mooring anchors
            1 for positive angle adjustment, -1 for negative angle adjustment
        consider_alternate_side: bool, optional
            True- take into account mooring headings of platform on other side
            if it is within 2 mooring radii

        Returns
        -------
        None.

        '''
        # - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - #
        # create substation object with id 
        if oss:
            if None in substation_r:
                dd = {'r':[5000,1000]}
            else:
                dd = {'r':substation_r}
            self.substationList[ss_id] = Substation(dd,id=ss_id)
            self.substationList[ss_id].rFair = 5 ##### TEMPORARY #####
            self.substationList[ss_id].zFair = -14
            self.substationList[ss_id].phi = 0
            
        # create reference cables for resizing
        ref_cables = []
        
        cabProps = {}
        cp = loadCableProps(None)
        
        # detach and delete existing cable list unless specified to keep old cables
        if keep_old_cables:
            lcab = len(self.cableList)
        else:
            if self.cableList:
                for j,cab in enumerate(self.cableList.values()):
                    cab.detachFrom('a')
                    cab.detachFrom('b')
            self.cableList = {} 
            lcab = 0
        #### - - - -  Find cable attachments  - - - - #### 
        for i in range(0,len(connDict)): # go through each cable       
            # attach to platforms/substations
            for pf in self.platformList.values():
                if id_method == 'location':
                    # find platform associated with ends
                    if np.allclose(pf.r[:2],connDict[i]['coordinates'][0],atol=.01): 
                        attA = pf
                    elif np.allclose(pf.r[:2],connDict[i]['coordinates'][-1],atol=.01):
                        attB = pf
                elif id_method == 'id':
                    # find platform associated with global id
                    if connDict[i]['turbineA_glob_id'] == pf.id:
                        attA = pf
                        # update platform location
                        pf.r[:2] = connDict[i]['coordinates'][0]
                    elif connDict[i]['turbineB_glob_id'] == pf.id:
                        attB = pf
                        # update platform location
                        pf.r[:2] = connDict[i]['coordinates'][-1]
            
            # get heading of cable from attached object coordinates (compass heading)
            headingA = calc_heading(connDict[i]['coordinates'][-1],
                                    connDict[i]['coordinates'][0])
            headingB = headingA + np.pi

            # figure out approx. depth at location
            initial_depths = []
            if cableConfig:
                avgspan = np.mean([x['span'] for x in cableConfig['cableTypes'].values() if 'span' in x])
                # depth at avg span (this is just a rough estimate!)
                endLocA = connDict[i]['coordinates'][0] + avgspan*np.array([np.cos(headingA),np.sin(headingA)])
                endLocB = connDict[i]['coordinates'][-1] + avgspan*np.array([np.cos(headingB),np.sin(headingB)])
                # get depth at these locs
                initial_depths.append(self.getDepthAtLocation(*endLocA))
                initial_depths.append(self.getDepthAtLocation(*endLocB))
                # select cable and collect design dictionary info on cable
                selected_cable, dd, cable_candidates = getCableDesign(
                    connDict[i], cableType_def, 
                    cableConfig, configType, 
                    depth=np.mean(initial_depths)
                    )
            else:
                dd = {}
                dd['cables'] = []
                dd['cables'].append({})
                cd = dd['cables'][0]
                cd['span'] = connDict[i]['2Dlength']
                cd['L'] = connDict[i]['2Dlength']
                cd['A'] = connDict[i]['conductor_area']
                cd['voltage'] = cableType_def[-2:]
                cd['rJTube'] = 5
                
                # add routing if necessary
                if len(connDict[i]['coordinates'])>2:
                    cd['routing'] = []
                    for coord in connDict[i]['coordinates'][1:-1]:
                        cd['routing'].append(coord)
                
                if not 'cable_type' in cd:
                    if not cd['A'] in cabProps.keys():
                        cabProps[cd['A']] = getCableProps(connDict[i]['conductor_area'],cableType_def,cableProps=cp)
                    # fix units
                    cabProps[cd['A']]['power'] = cabProps[cd['A']]['power']*1e6
                    cd['cable_type'] = cabProps[cd['A']]
                    
                selected_cable = None
            
                
            # create cable object
            cab = Cable('cable'+str(i+lcab),d=dd)
            self.cableList[cab.id] = cab

            # update upstream turbines property
            cab.upstream_turb_count = getFromDict(connDict[i],'upstream_turb_count',default=0)
            
            # attach cable
            cab.attachTo(attA,end='a')
            cab.attachTo(attB,end='b')
            
            if cableConfig:
                if cable_candidates:
                    cab.subcomponents[0].alternate_cables=cable_candidates
                    cab.subcomponents[-1].alternate_cables=cable_candidates
                if 'head_offset' in selected_cable:
                    headingA += np.radians(selected_cable['head_offset'])
                    headingB -= np.radians(selected_cable['head_offset'])
                
                # adjust heading if too close to moorings
                rad_buff = np.radians(heading_buffer)
                dc0s = np.max((cab.subcomponents[0].span, 500))
                moors = attA.getMoorings() 
                msp = list(moors.values())[0].span + attA.rFair + 200 # add a bit extra
                # consider mooring headings from both ends if close enough
                pfsp = np.linalg.norm(attA.r-attB.r) 

                if consider_alternate_side and pfsp-2*attA.rFair < msp+dc0s:
                    headingA = head_adjust([attA,attB],
                                           headingA,
                                           rad_buff=rad_buff,
                                           adj_dir=adj_dir)
                    headingB = head_adjust([attB,attA],
                                           headingB,
                                           rad_buff=rad_buff,
                                           endA_dir=-1,
                                           adj_dir=adj_dir)
                else:
                    headingA = head_adjust([attA],
                                           headingA,
                                           rad_buff=rad_buff,
                                           adj_dir=adj_dir)
                    headingB = head_adjust([attB],
                                           headingB,
                                           rad_buff=rad_buff,
                                           adj_dir=adj_dir)
                    
            heads = [headingA,headingB]
            # reposition cable
            cab.reposition(project=self,headings=heads,rad_fair=[5,5])

            coords = []
            if cableConfig:
                ref_cables = None
                # add routing for static cable to continue along adjusted heading for total of 500m (inluding dynamic cable span) & adjust dynamic cable depths as needed
                if len(cab.subcomponents)>1:
                    inds = [0,-1]
                    for ii,ind in enumerate(inds):
                        # adjust for depth as needed
                        dc = cab.subcomponents[ind]
                        if ref_cables:
                            dc.dd = self.cableDesignInterpolation(dc.z_anch,ref_cables) #***** left off here 1/28/25 need to check z_anch sign here and figure out how to send cables options *****
                        # add static routing
                        if cab.subcomponents[ind].span < 500:
                            spandiff = 500 - cab.subcomponents[ind].span
                            ind_of_stat = 2-4*ii # 2 for end A, -2 for end B -- relative loc of static cable compared to dynamic cable
                            stat_cable = cab.subcomponents[ind+ind_of_stat]
                            # get new coordinate routing point
                            stat_cable_end = stat_cable.rA if ind==0 else stat_cable.rB
                            coord = [stat_cable_end[0] + np.cos(np.pi/2-heads[ii])*spandiff,
                                        stat_cable_end[1] + np.sin(np.pi/2-heads[ii])*spandiff]
                            # append it to static cable object coordinates
                            coords.append(coord)
            
            # update lengths & spans of any static cables as needed
            cts = np.where([isinstance(a,StaticCable) for a in cab.subcomponents])[0]
            for cs in cts:
                # update routing
                cab.subcomponents[cs].updateRouting(coords) # also updates static and general cable lengths


        if route_anchors:
            route_around_anchors(self)

                              
    
    def updatePositions(self):
        '''Temporary quick-fix to update Platform object positions based on
        RAFT or MoorPy body positions.'''
        
        for platform in self.platformList.values():
            platform.r[0] = platform.body.r6[0]
            platform.r[1] = platform.body.r6[1]
        
    
    def plot2d(self, ax=None, plot_soil=False,
               plot_bathymetry=True, plot_boundary=True, color_lineDepth=False, 
               plot_bathymetry_contours=False, bare=False, axis_equal=True,
               save=False,**kwargs):
        '''Plot aspects of the Project object in matplotlib in 3D.
        
        TODO - harmonize a lot of the seabed stuff with MoorPy System.plot...
        
        Parameters
        ----------
        ...
        ax : matplotlib.pyplot axis
            Default is None
        plot_soil : bool
            If True, plot soil conditions
        plot_bathymetry : bool
            If True, plot bathymetry 
        plot_boundary : bool
            If True, plot lease area boundary
        plot_bathy_contours : bool
            If True, plot bathymetry line contours
        axis_equal : bool
            If True, set axes to equal scales to prevent visual distortions
        save : bool
            If True, save the figure
        bare : bool
            If True, supress display of extra labeling like the colorbar.
        color_lineDepth: bool
            If True, color mooring lines based on depth. Only works if plot_bathymetry=False.
        '''
     
        # Handle extra keyword arguments or use default values
        figsize                 = kwargs.get('figsize', (8,8))  # the dimensions of the figure to be plotted
        edgecolor               = kwargs.get('env_color',[.5,0,0,.8])
        color                   = kwargs.get('fenv_color',[.6,.3,.3,.6])
        alpha                   = kwargs.get('alpha',0.5)
        return_contour          = kwargs.get('return_contour',False)
        cmap_cables             = kwargs.get('cmap_cables',None)
        cmap_soil               = kwargs.get('cmap_soil', None)
        plot_platforms          = kwargs.get('plot_platforms',True)
        plot_anchors            = kwargs.get('plot_anchors',True)
        plot_moorings           = kwargs.get('plot_moorings',True)
        plot_cables             = kwargs.get('plot_cables',True)
        cable_labels            = kwargs.get('cable_labels', False)
        depth_vmin              = kwargs.get('depth_vmin', None)
        depth_vmax              = kwargs.get('depth_vmax', None)
        bathymetry_levels       = kwargs.get('bathymetry_levels', 50)
        plot_legend             = kwargs.get('plot_legend', True)
        legend_x                = kwargs.get('legend_x', 0.5)
        legend_y                = kwargs.get('legend_y', -0.1)
        plot_landmask           = kwargs.get('plot_landmask', False) # mask land areas 
        soil_alpha              = kwargs.get('soil_alpha', 0.5)
        max_line_depth          = kwargs.get('max_line_depth', None)  # max depth for line coloring if color_lineDepth is True
        maxcableSize            = kwargs.get('maxcableSize', None)
        only_shared             = kwargs.get('only_shared', False)   # if color_lineDepth is True, only color shared lines
        linewidth_multiplier    = kwargs.get('linewidth_multiplier', 2)  # multiplier for line widths if color_lineDepth is True
        # if axes not passed in, make a new figure
        if ax == None:
            fig, ax = plt.subplots(1,1, figsize=figsize)
        else:
            fig = ax.get_figure()
        
        
        # Bathymetry 
        if plot_bathymetry:
            if plot_soil:
                raise ValueError('The bathymetry grid and soil grid cannot yet be plotted at the same time. Use plot_bathy_contours=True instead')
            if len(self.grid_x) > 1 and len(self.grid_y) > 1:
                
                X, Y = np.meshgrid(self.grid_x, self.grid_y)

                vmin = depth_vmin if depth_vmin is not None else np.min(self.grid_depth)
                vmax = depth_vmax if depth_vmax is not None else np.max(self.grid_depth)
                grid_depth = np.clip(self.grid_depth, vmin, vmax)

                contourf = ax.contourf(X, Y, grid_depth, 
                                       bathymetry_levels, 
                                       cmap='Blues', 
                                       vmin=np.min(self.grid_depth), 
                                       vmax=np.max(self.grid_depth))

                contourf.set_clim(depth_vmin, depth_vmax)

                if not bare:  # Add colorbar with label
                    import matplotlib.ticker as tkr
                    cbar = plt.colorbar(contourf, ax=ax, fraction=0.04, label='Water Depth (m)', format=tkr.FormatStrFormatter('%.0f'))


        if plot_soil:
            from matplotlib.colors import BoundaryNorm

            if plot_bathymetry:
                raise ValueError('The bathymetry grid and soil grid cannot yet be plotted at the same time. Use plot_bathy_contours=True instead')

            soil_types = np.unique(self.soil_names).tolist()
            if not cmap_soil:
                cmap_soil = plt.cm.YlOrRd
            else:
                cmap_soil = plt.colormaps[cmap_soil]
            bounds = [i for i in range(len(soil_types)+1)]
            soil_type_to_int = {name: i for i,name in enumerate(soil_types)}
            soil_int = np.vectorize(soil_type_to_int.get)(self.soil_names)  
            norm = BoundaryNorm(bounds, cmap_soil.N)
            #cmap = mcolors.ListedColormap([soil_colors.get(name, 'white') for name in soil_types])
            # prepare for plot seabed data   
            #soil_types.remove("-")                          # delete nan like value manualy 
            levels = np.arange(0, len(soil_types))          # create index matches unique soil name
            ticks = levels + 0.5                        # shift label position to place center between colors
            X, Y = np.meshgrid(self.soil_x, self.soil_y)
            contourf = ax.pcolormesh(X, Y, soil_int, cmap=cmap_soil, norm=norm, shading='auto', alpha=soil_alpha)
            if not bare:
                cbar = plt.colorbar(contourf,
                                    ax=ax, norm=norm, 
                                    fraction=0.04, label='Soil Type', ticks=ticks) # color bar for soil name 
                cbar.ax.set_yticklabels(soil_types) # label of color bar
            #soil_handles = [plt.Line2D([0], [0], marker='s', color='w', label=name, markerfacecolor=soil_colors.get(name, 'white'), markersize=10) for name in soil_types if name != '0' ]
        
        if plot_bathymetry_contours:
            # plot the bathymetry in matplotlib using a plot_surface
            X, Y = np.meshgrid(self.grid_x, self.grid_y)  # 2D mesh of seabed grid
            plot_depths = self.grid_depth
            contour = ax.contour(X, Y, plot_depths, vmin=np.min(self.grid_depth), 
                       vmax=np.max(self.grid_depth), levels=bathymetry_levels, 
                       colors='black', linewidths=0.5) # bathymetry contour line
            ax.clabel(contour)
        
        if plot_landmask:
            # mask the land area (where depth>0)
            landMask = np.ones_like(self.grid_depth) # create land mask(sea: NaN, Land: 1)
            landMask[self.grid_depth > 0] = np.nan # replace water depth > 0 as NaN 
            contourf = ax.contourf(self.grid_x,self.grid_y, landMask, colors='gray') # landmask
            landmask_handle = plt.Line2D([0], [0], marker='s', color='w', label='Land', markerfacecolor='gray', markersize=10)
                    
        if plot_boundary:
            if len(self.boundary) > 1:
                ax.plot(self.boundary[:,0], self.boundary[:,1], 'b-.',label='Lease Boundary')
                for ez in self.exclusion:
                    ax.plot(ez[:,0], ez[:,1], 'r-.', label='Exclusion Zone')
            
        
        # Seabed ground/soil type (to update)
        #X, Y = np.meshgrid(self.soil_x, self.soil_y)
        #ax.scatter(X, Y, c=self.soil_rocky, s=4, cmap='cividis_r', vmin=-0.5, vmax=1.5)
        # or if we have a grid of soil types, something like
        # pcolormesh([X, Y,] C, **kwargs)  wjere C is [x, y, 3 rgba]
        
        
        # Plot any object envelopes
        if plot_platforms:
            from shapely import Point
            for platform in self.platformList.values():
                for name, env in platform.envelopes.items():
                    ax.fill(env['x'], env['y'], edgecolor=edgecolor, facecolor='none', linestyle='dashed', lw=0.8, label='Platform Envelope')
        
        if plot_moorings:
            line_depth_settings = None
            if color_lineDepth:
                if plot_bathymetry:
                    raise ValueError("Cannot use depth-based line coloring with plot_bathymetry=True. Disable bathymetry to avoid confusion.")
                line_depth_settings = {
                    "cmap": "Blues",
                    "vmin": max_line_depth if max_line_depth else -np.max(self.grid_depth),
                    "vmax": 0,
                    "only_shared": only_shared,
                    "linewidth_multiplier": linewidth_multiplier
                }            
            for mooring in self.mooringList.values():
                for name, env in mooring.envelopes.items():
                    #if 'shape' in env:  # if there's a shapely object
                    #    pass  # do nothing for now...
                    #elif 'x' in env and 'y' in env:  # otherwise just use coordinates
                    ax.fill(env['x'], env['y'], color=color,label='Mooring Envelope',alpha=alpha)
        
        
            # Plot moorings one way or another (eventually might want to give Mooring a plot method)
            for mooring in self.mooringList.values():
                lineList = []
                if mooring.ss:  # plot with Subsystem if available
                    lineList = mooring.ss.lineList

                elif mooring.parallels:
                    for i in mooring.i_sec:
                        sec = mooring.getSubcomponent(i)
                        if hasattr(sec,'mpLine'):
                            lineList.append(sec.mpLine)
                            line = sec.mpLine

                labs = []
                for line in lineList:
                    if 'chain' in line.type['material']:
                        line.color = 'k'                           
                    elif 'polyester' in line.type['material']:
                        line.color = [.3,.5,.5]
                    else:
                        line.color = [0.5,0.5,0.5]
                    labs.append(line.type['material'][0].upper()+
                                line.type['material'][1:]+' Mooring')
                    
                if mooring.ss:
                    mooring.ss.drawLine2d(0, ax, color="self", 
                                          plot_endpoints=False, 
                                          Xuvec=[1,0,0], Yuvec=[0,1,0], 
                                          line_depth_settings=line_depth_settings, 
                                          label=labs)  
                elif mooring.parallels:
                    for i,line in enumerate(lineList):
                        line.drawLine2d(0, ax, color="self",
                                        Xuvec=[1,0,0], Yuvec=[0,1,0], 
                                        label=labs[i])

                else: # simple line plot
                    ax.plot([mooring.rA[0], mooring.rB[0]], 
                            [mooring.rA[1], mooring.rB[1]], 'k', lw=0.5, 
                            label='Mooring Line')

        # ---- Add colorbar for line depth ----
        if line_depth_settings is not None and not bare:
            import matplotlib.cm as cm
            import matplotlib.colors as mcolors
            sm = cm.ScalarMappable(cmap=cm.get_cmap(line_depth_settings["cmap"]),
                                norm=mcolors.Normalize(vmin=line_depth_settings["vmin"],
                                                        vmax=line_depth_settings["vmax"]))
            sm.set_array([])
            cbar = plt.colorbar(sm, ax=ax, fraction=0.04)
            cbar.set_label("Line Depth (m)")  
                          
        if plot_anchors:
            for anchor in self.anchorList.values():
                ax.plot(anchor.r[0],anchor.r[1], 'mo',ms=2, label='Anchor')
        
        # Plot cables one way or another (eventually might want to give Mooring a plot method)
        if self.cableList and plot_cables:
            if maxcableSize == None:
                maxcableSize = max([cab.dd['cables'][0].dd['A'] for cab in self.cableList.values()])

            for cable in self.cableList.values():
                # get cable color
                import matplotlib.cm as cm
                if not cmap_cables:
                    cmap = cm.get_cmap('plasma_r')
                else:
                    cmap = cm.get_cmap(cmap_cables)
                cableSize = int(cable.dd['cables'][0].dd['A'])
                Ccable = cmap(cableSize/(1.1*maxcableSize))
                # # simple line plot for now
                # ax.plot([cable.subcomponents[0].rA[0], cable.subcomponents[-1].rB[0]], 
                #         [cable.subcomponents[0].rA[1], cable.subcomponents[-1].rB[1]],'--',color = Ccable, lw=1,label='Cable '+str(cableSize)+' mm$^{2}$')
                
                # add in routing if it exists
                for sub in cable.subcomponents:
                    if isinstance(sub,StaticCable):
                        if len(sub.x)>0:
                            # has routing  - first plot rA to sub.coordinate[0] connection
                            ax.plot([sub.rA[0],sub.x[0]],
                                    [sub.rA[1],sub.y[0]],':',color = Ccable,
                                    lw=1.2,label=f'Static Cable {cableSize} mm$^{2}$')
                            # now plot route
                            if len(sub.x) > 1:
                                for i in range(1,len(sub.x)):
                                    ax.plot([sub.x[i-1],sub.x[i]],
                                            [sub.y[i-1],sub.y[i]],
                                            ':', color=Ccable, lw=1.2,
                                            label=f'Static Cable {cableSize} mm$^{2}$')
                            # finally plot sub.coordinates[-1] to rB connection
                            ax.plot([sub.x[-1],sub.rB[0]],
                                    [sub.y[-1],sub.rB[1]],':',color=Ccable,
                                    lw=1.2,label=f'Static Cable {cableSize} mm$^{2}$')
                        else:
                            # if not routing just do simple line plot
                            ax.plot([sub.rA[0],sub.rB[0]], 
                                    [sub.rA[1], sub.rB[1]],':',color = Ccable, lw=1.2,
                                    label=f'Static Cable {cableSize} mm$^{2}$')
                        
                        # if cable_labels:
                        #     x = np.mean([sub.rA[0],sub.rB[0]])
                        #     y = np.mean([sub.rA[1],sub.rB[1]])
                        #     if '_' in cable.id:
                        #         label = cable.id.split('_')[-1]
                        #     else:
                        #         label = cable.id
                        #     ax.text(x,y, label)
                    elif isinstance(sub,DynamicCable):
                            ax.plot([sub.rA[0],sub.rB[0]], 
                                    [sub.rA[1], sub.rB[1]],color = Ccable, lw=1.2,
                                    label=f'Dynamic Cable {cableSize} mm$^{2}$')
                            
                            if cable_labels:
                                x = np.mean([sub.rA[0],sub.rB[0]])
                                y = np.mean([sub.rA[1],sub.rB[1]])
                                if '_' in cable.id:
                                    label = cable.id.split('_')[-1]
                                else:
                                    label = cable.id
                                ax.text(x,y, label)
            
                # ax.plot([cable.subcomponents[0].rA[0], cable.subcomponents[-1].rB[0]], 
                #         [cable.subcomponents[0].rA[1], cable.subcomponents[0].rB[1]], 'r--', lw=0.5)
        
        # Plot platform one way or another (might want to give Platform a plot method)
        if plot_platforms:
            for platform in self.platformList.values():
                entity = platform.entity
                if 'FOWT' in entity.upper():
                    plotstring = 'ko'
                elif 'SUBSTATION' in entity.upper():
                    plotstring = 'go'
                elif 'WEC' in entity.upper():
                    plotstring = 'ro'
                else:
                    plotstring = 'bo'
                    
                ax.plot(platform.r[0], platform.r[1], plotstring, label=entity, ms=3.5)
  
        ax.set_xlabel('X (m)')
        ax.set_ylabel('Y (m)')

        if axis_equal:
            ax.set_aspect('equal',adjustable='box')

        handles, labels = ax.get_legend_handles_labels()
        # add in land mask label if necessary
        if plot_landmask:
            handles += [landmask_handle]
            labels += ['Land']
        # zip labels and handles into a dictionary
        by_label = dict(zip(labels, handles))  # Removing duplicate labels
        
        if plot_legend:
            ax.legend(by_label.values(), by_label.keys(),loc='upper center',bbox_to_anchor=(legend_x, legend_y), fancybox=True, ncol=4)
        if save:
            counter = 1
            output_filename = f'2dfarm_{counter}.png'
            while os.path.exists(output_filename):
                counter += 1
                output_filename = f'2dfarm_{counter}.png'
            
            # Increase the resolution when saving the plot
            plt.savefig(output_filename, dpi=600, bbox_inches='tight')  # Adjust the dpi as needed
            
            # TODO - add ability to plot from RAFT FOWT
        if return_contour:
            return(ax,contourf)
        else:
            return(ax)   
        
        

    def plot3d(self, ax=None, figsize=(11,8), plot_fowt=False, save=False,
               plot_boundary=True, plot_boundary_on_bath=True, args_bath={}, 
               plot_axes=True, plot_bathymetry=True, plot_soil=False, color=None,
               colorbar=True, boundary_only=False, plot_bathy_contours=False,
               **kwargs):
        '''Plot aspects of the Project object in matplotlib in 3D.
        
        TODO - harmonize a lot of the seabed stuff with MoorPy System.plot...
        
        Parameters
        ----------
        ax : matplotlib.axes._subplots.Axes3DSubplot, optional
            The 3D axes object to plot on. If None, a new figure and axes will be created.
        
        figsize : tuple, optional
            The size of the figure in inches, specified as (width, height). Default is (10, 8).
        
        plot_fowt : bool, optional
            Whether to plot the floating offshore wind turbine (FOWT) in the 3D visualization. Default is False.
        
        save : bool, optional
            Whether to save the plot to a file. Default is False.
        
        plot_boundary : bool, optional
            Whether to plot the project boundary in the 3D visualization. Default is True.
        
        plot_boundary_on_bath : bool, optional
            Whether to overlay the project boundary on the bathymetry plot. Default is True.
        
        args_bath : dict, optional
            Additional arguments to customize the bathymetry plot, such as contour levels or color maps. Default is an empty dictionary.
        
        plot_axes : bool, optional
            Whether to display the 3D axes in the plot. Default is True.
        
        plot_bathymetry : bool, optional
            Whether to include the bathymetry (seabed depth) in the 3D visualization. Default is True.
        
        plot_soil : bool, optional
            Whether to include soil properties or layers in the 3D visualization. Default is False.
        
        color : str or matplotlib color, optional
            The color to use for certain plot elements (lines and raft members - maybe cable too?). If None, default colors will be used. Default is None.
        
        colorbar : bool, optional
            Whether to include a colorbar in the plot (e.g., for bathymetry). Default is True.
        
        boundary_only : bool, optional
            If True, only the project boundary will be plotted, ignoring other elements. Default is False.
        
        **kwargs : dict, optional
            Additional keyword arguments passed to the underlying matplotlib plotting functions.
        
        Returns
        -------
        ax : matplotlib.axes._subplots.Axes3DSubplot
            The 3D axes object containing the plot.
        '''
        
        # color map for soil plotting
        import matplotlib.cm as cm
        # from matplotlib.colors import Normalize
        cmap_cables     = kwargs.get('cmap_cables','plasma_r')
        alpha           = kwargs.get('alpha',1)
        orientation     = kwargs.get('orientation',[20, -130])
        maxcableSize    = kwargs.get('maxcableSize', None)    
        plot_legend     = kwargs.get('plot_legend', False)
        depth_vmin      = kwargs.get('depth_vmin', None)
        depth_vmax      = kwargs.get('depth_vmax', None)
        bathymetry_levels       = kwargs.get('bathymetry_levels', 100)
        cmap_soil       = kwargs.get('cmap_soil', 'YlOrRd')
        soil_alpha      = kwargs.get('soil_alpha', 0.5)


        # if axes not passed in, make a new figure
        if ax == None:    
            fig = plt.figure(figsize=figsize)
            ax = plt.axes(projection='3d')
        else:
            fig = ax.get_figure()

        soil_cbar = None
        # ---------------------------
        # Bathymetry / soil plotting
        # ---------------------------
        if plot_bathymetry and len(self.grid_x) > 1 and len(self.grid_y) > 1:

            if boundary_only:
                self.trimGrids()

            X, Y = np.meshgrid(self.grid_x, self.grid_y)
            Z = -self.grid_depth

            # Combined mode: bathymetry geometry + soil facecolors
            if plot_soil:
                if self.soil_x is None or self.soil_y is None or len(np.shape(self.soil_names)) != 2:
                    raise ValueError("plot_soil=True requires valid soil_x, soil_y, and 2D soil_names.")

                # TODO: if soil names do not match EMODNET no colors are plotted --> fix
                soil_on_bathy = _nearest_soil_names_on_bathymetry_grid(
                    self.grid_x, self.grid_y,
                    self.soil_x, self.soil_y,
                    self.soil_names
                )

                facecolors, soil_types, soil_type_to_int, cmap_obj, norm = _build_soil_facecolors(
                    soil_on_bathy,
                    cmap_soil=cmap_soil,
                    soil_alpha=soil_alpha
                )

                bath = ax.plot_surface(
                    X, Y, Z,
                    facecolors=facecolors,
                    alpha=soil_alpha,
                    rstride=1, cstride=1,
                    linewidth=0,
                    antialiased=False,
                    shade=False,
                    zorder=1
                )

                if plot_bathy_contours:
                    ax.contour(
                        X, Y, Z,
                        levels=min(bathymetry_levels, 12),
                        colors='k',
                        linewidths=0.35,
                        alpha=1
                    )

                if colorbar:
                    sm = ScalarMappable(cmap=cmap_obj, norm=norm)
                    sm.set_array([])
                    soil_cbar = plt.colorbar(
                        sm,
                        ax=ax,
                        fraction=0.04,
                        ticks=np.arange(len(soil_types)) + 0.5
                    )
                    soil_cbar.set_label('Soil Type')
                    soil_cbar.ax.set_yticklabels(soil_types)

            # Bathymetry-only mode
            else:
                vmin = depth_vmin if depth_vmin is not None else -np.max(self.grid_depth)
                vmax = depth_vmax if depth_vmax is not None else -np.min(self.grid_depth)

                bath = ax.plot_surface(
                    X, Y, Z,
                    cmap='Blues_r',
                    vmin=vmin,
                    vmax=vmax,
                    rstride=1, cstride=1,
                    linewidth=0,
                    antialiased=False,
                    alpha=0.75,
                    zorder=1
                )

                if plot_bathy_contours:
                    ax.contour(
                        X, Y, Z+3,
                        levels=min(bathymetry_levels, 12),
                        colors='k',
                        linewidths=1.,
                        alpha=1,
                        zorder=2
                    )

                if colorbar:
                    cb = fig.colorbar(bath, ax=ax, fraction=0.04)
                    cb.set_label('Water Depth (m)')

            # Water plane for both bathymetry-only and combined mode
            Z0 = np.zeros_like(X)
            ax.plot_surface(
                X, Y, Z0,
                color='lightblue',
                alpha=0.2,
                edgecolor='none',
                zorder=0
            )

        # Soil-only mode
        elif plot_soil:
            if self.soil_x is None or self.soil_y is None or len(np.shape(self.soil_names)) != 2:
                raise ValueError("plot_soil=True requires valid soil_x, soil_y, and 2D soil_names.")

            Xs, Ys = np.meshgrid(self.soil_x, self.soil_y)
            # NOTE: we may want to fix the Z values to be at the seabed elevation (water depth)
            Zs = np.zeros_like(Xs)

            facecolors, soil_types, soil_type_to_int, cmap_obj, norm = _build_soil_facecolors(
                self.soil_names,
                cmap_soil=cmap_soil,
                soil_alpha=soil_alpha
            )

            ax.plot_surface(
                Xs, Ys, Zs,
                facecolors=facecolors,
                rstride=1, cstride=1,
                linewidth=0,
                antialiased=False,
                shade=False,
                zorder=1
            )

            if colorbar:
                sm = ScalarMappable(cmap=cmap_obj, norm=norm)
                sm.set_array([])
                soil_cbar = plt.colorbar(
                    sm,
                    ax=ax,
                    fraction=0.04,
                    ticks=np.arange(len(soil_types)) + 0.5
                )
                soil_cbar.set_label('Soil Type')
                soil_cbar.ax.set_yticklabels(soil_types)

        # plot the project boundary
        if plot_boundary:
            ax.plot(self.boundary[:,0], self.boundary[:,1], 
                    np.zeros(self.boundary.shape[0]), 
                    'b--', zorder=100, lw=0.5, alpha=0.5)
            
        # plot the projection of the boundary on the seabed, if desired
        if plot_boundary_on_bath:
            boundary_z = -self.projectAlongSeabed(self.boundary[:,0], self.boundary[:,1])
            ax.plot(self.boundary[:,0], self.boundary[:,1], boundary_z, 
                    'k--', zorder=10, lw=0.5, alpha=0.7)

        lw=1 #0.5
        # find max cable size as applicable
        if self.cableList:
            if maxcableSize == None:
                maxcableSize = max([a.subcomponents[0].dd['A'] for a in self.cableList.values()])
        for cable in self.cableList.values():
            # get cable color
            import matplotlib.cm as cm
            cmap = cm.get_cmap(cmap_cables)
            cableSize = cable.dd['cables'][0].dd['A']
            Ccable = cmap(cableSize/maxcableSize)
            
            for j,sub in enumerate(cable.subcomponents):
                if isinstance(sub,DynamicCable):
                    if sub.ss:
                        for ii in range(len(sub.ss.lineList)):
                            sub.ss.lineList[ii].color=Ccable
                            sub.ss.lineList[ii].lw=lw
                        sub.ss.drawLine(0,ax,color=Ccable,plot_shadow=False)
                        
                elif isinstance(sub,StaticCable):
                    # add static cable routing if it exists
                    if len(sub.x)>0:
                        # burial = sub.burial
                        # bur = []
                        # if burial and 'station' in burial:
                        #     # replace any NA with 0
                        #     for i,b in enumerate(burial['station']):
                        #         if b.upper() == 'NONE':
                        #            bur[i] = 0 
                        #     if not burial:
                        #         burial.append(0)
                        burial = []
                        burial.append(0) #### UPDATE LATER!!!
                        # get joint locations
                        jointA = cable.subcomponents[j-1]['r']
                        jointB = cable.subcomponents[j+1]['r']
                        try:
                            if len(sub.x)==1:
                                soil_z= [self.getDepthAtLocation(sub.x[0],sub.y[0])]
                            else:
                                soil_z = self.projectAlongSeabed(sub.x,sub.y)
                        except:
                            soil_z = [self.depth]
                            
                        # plot connections from joints to first and last routing point
                        ax.plot([jointA[0],sub.x[0]],[jointA[1],sub.y[0]],[-soil_z[0],-soil_z[0]-burial[0]],
                                ':',color=Ccable,zorder=5,lw=lw, label=f'Static Cable {cableSize} mm$^{2}$')
                        ax.plot([jointB[0],sub.x[-1]],[jointB[1],sub.y[-1]],[-soil_z[-1],-soil_z[-1]-burial[-1]],
                                ':',color=Ccable,zorder=5,lw=lw, label=f'Static Cable {cableSize} mm$^{2}$')
                        
                        if len(sub.x)> 1:
                            # plot in 3d along soil_z
                            for ii in range(len(sub.x)-1):
                                ax.plot([sub.x[ii],sub.x[ii+1]],
                                        [sub.y[ii],sub.y[ii+1]],
                                        [-soil_z[ii],-soil_z[ii+1]],
                                        ':', color=Ccable, zorder=5, lw=lw)
                    else:
                        # no routing - just plot a straight line
                        ax.plot([sub.rA[0],sub.rB[0]],[sub.rA[1],sub.rB[1]],[sub.rA[2],sub.rB[2]],':',color=Ccable,zorder=5,lw=lw, label=f'Static Cable {cableSize} mm$^{2}$')
        
        # plot the Moorings
        ct = 0
        for mooring in self.mooringList.values():
            #mooring.subsystem.plot(ax = ax, draw_seabed=False)
            labs = []
            if mooring.ss:
                for line in mooring.ss.lineList:
                    if color:  # overwrite if color is given.
                        line.color = color
                    else:
                        if 'chain' in line.type['material']:
                            line.color = 'k'
                        elif 'polyester' in line.type['material']:
                            line.color = [.3,.5,.5]
                        else:
                            line.color = [0.5,0.5,0.5]
                    line.lw = lw
                    labs.append(line.type['material'][0].upper()+ line.type['material'][1:]+' Mooring')
                    
                mooring.ss.drawLine(0, ax, color='self')
            elif mooring.parallels:
                for i in mooring.i_sec:
                    sec = mooring.getSubcomponent(i)
                    if hasattr(sec,'mpLine'):
                        line = sec.mpLine
                        if 'chain' in line.type['material']:
                            line.color = 'k'
                        elif 'polyester' in line.type['material']:
                            line.color = [.3,.5,.5]
                        else:
                            line.color = [0.5,0.5,0.5]
                        line.lw = lw
                        line.drawLine(0,ax,color='self')
                        
        # plot the FOWTs using a RAFT FOWT if one is passed in (TEMPORARY)
        if plot_fowt:
            for pf in self.array.fowtList:
                pf.plot(ax, color=color, zorder=1000,plot_ms=False, axes_around_fowt=False,shadow=False)
            # for i in range(self.nt):
            #     xy = self.turb_coords[i,:]
            #     fowt.setPosition([xy[0], xy[1], 0,0,0,0])
            #     fowt.plot(ax, zorder=20)
        
        # Show full depth range
        if len(self.grid_depth)>1:
            ax.set_zlim([-np.max(self.grid_depth), 0])
        else:
            ax.set_zlim([-self.depth,0])

        if not plot_axes:
            ax.axis('off')
        else:
            ax.set_xlabel("X (m)",labelpad=12)
            ax.set_ylabel("Y (m)",labelpad=12)
            # ax.set_zlabel("Depth (m)")
        # ax.set_box_aspect([1, 1, 0.01])  # exaggerate depth slightly
        set_axes_equal(ax)

        # Remove panes (background walls)
        ax.xaxis.pane.fill = False
        ax.yaxis.pane.fill = False
        ax.zaxis.pane.fill = False

        # Remove pane edges
        ax.xaxis.pane.set_edgecolor('w')
        ax.yaxis.pane.set_edgecolor('w')
        ax.zaxis.pane.set_edgecolor('w')

        # Remove grid
        ax.grid(True)

        # Set grid line width to 0 (in case grid is turned back on)
        ax.xaxis._axinfo["grid"]['linewidth'] = 0.1
        ax.yaxis._axinfo["grid"]['linewidth'] = 0.1
        ax.zaxis._axinfo["grid"]['linewidth'] = 0

        ax.tick_params(axis='both', which='major', labelsize=9)

        # Keep your existing z-limits
        if len(self.grid_depth) > 1:
            z_bottom = -np.max(self.grid_depth)   # seabed (negative)
        else:
            z_bottom = -self.depth

        ax.set_zlim(z_bottom, 0)

        ax.set_zticks([z_bottom, 0])
        # ax.set_zticklabels(["Water depth", "Waterline (z=0)"])
        ax.set_zticklabels([f"{abs(z_bottom):.0f} m", "0 m"], fontsize=8)
        
        ax.view_init(orientation[0],orientation[1])

        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))  # Removing duplicate labels
        
        if plot_legend:
            ax.legend(by_label.values(), by_label.keys(), loc = 'lower center', fancybox=True, ncol=3)
        
        
        # fig.tight_layout()
        
        # ----- Save plot with an incremented number if it already exists
        if save:
            counter = 1
            output_filename = f'wind farm 3d_{counter}.png'
            while os.path.exists(output_filename):
                counter += 1
                output_filename = f'wind farm 3d_{counter}.png'
            
            # Increase the resolution when saving the plot
            plt.savefig(output_filename, dpi=600, bbox_inches='tight')  # Adjust the dpi as needed
            
        return(ax)
        
       
    def getMoorPyArray(self, plt=0, cables=True):
        '''Creates an array in moorpy from the mooring, anchor, connector, and platform objects in the array.

        Parameters
        ----------
        plt : number
            Controls whether to create a plot of the MoorPy array. 2=create plot with labels, 1=create plot, 0=no plot The default is 0.

        cables : boolean, optional
            Controls whether to include cables in the moorpy array
        
        Returns
        -------
        ms : class instance
            MoorPy system for the whole array based on the subsystems in the mooringList

        '''
        print('Creating MoorPy Array')
        # create MoorPy system
        if len(self.grid_x)>1: # add bathymetry if available
            bath = {'x':self.grid_x,'y':self.grid_y,'depth':self.grid_depth}
            self.ms = mp.System(depth=self.depth,bathymetry=bath)
        else:
            self.ms = mp.System(depth=self.depth)
        # reset all anchor objects listed in self.anchorList to None in case there was a previous ms
        for i in self.anchorList:
            self.anchorList[i].mpAnchor = None
        for i in self.platformList:
            self.platformList[i].body = None
        
        wflag = 0 # warning flag has not yet been printed (prevent multiple printings of same hydrostatics warning)
        for i,body in enumerate(self.platformList.values()): # make all the bodies up front - i is index in dictionary, body is key (name of platform)
            # add a moorpy body at the correct location
            r6 = [body.r[0],body.r[1],body.r[2],0,0,0]
            # use model_details portion of dictionary to create moorpy body if given
            if 'hydrostatics' in body.dd and body.dd['hydrostatics']:
                bodyInfo = body.dd['hydrostatics']
                self.ms.addBody(-1,r6,m=bodyInfo['m'],v=bodyInfo['v'],rCG=np.array(bodyInfo['rCG']),rM=np.array(bodyInfo['rM']),AWP=bodyInfo['AWP'])
            elif not 'hydrostatics' in body.dd and wflag == 0: # default to UMaine VolturnUS-S design hydrostatics info
                print('No hydrostatics information given, so default body hydrostatics from UMaine VolturnUS-S will be used.')
                wflag = 1
                self.ms.addBody(-1,r6,m=19911423.956678286,rCG=np.array([ 1.49820657e-15,  1.49820657e-15, -2.54122031e+00]),v=19480.104108645974,rM=np.array([2.24104273e-15, 1.49402849e-15, 1.19971829e+01]),AWP=446.69520543229874)
            else:
                self.ms.addBody(-1,r6,m=19911423.956678286,rCG=np.array([ 1.49820657e-15,  1.49820657e-15, -2.54122031e+00]),v=19480.104108645974,rM=np.array([2.24104273e-15, 1.49402849e-15, 1.19971829e+01]),AWP=446.69520543229874)
            body.body = self.ms.bodyList[-1]
        
        # create anchor points and all mooring lines connected to the anchors (since all connected to anchors, can't be a shared mooring)
        for anchor in self.anchorList.values():  # Go through each anchor
            
            # Create it's MoorPy Point object
            if anchor.mpAnchor: # If anchor already exists in MoorPy
                print("Why does this anchor already have a MoorPy Point?")
                breakpoint()
            
            anchor.makeMoorPyAnchor(self.ms)
            num = anchor.mpAnchor.number
            
            # Go through each thing/mooring attached to the anchor
            for j, att in anchor.attachments.items(): 
                
                mooring = att['obj']
                
                # create subsystem
                
                mooring.createSubsystem(ms=self.ms)

                # set location of subsystem for simpler coding
                ssloc = mooring.ss
                
                # (ms.lineList.append is now done in Mooring.createSubsystem)
                
                # Attach the Mooring to the anchor
                if mooring.parallels:  # the case with parallel sections, multiple MoorPy objects
                    
                    # note: att['end'] should always be 0 in this part of the
                    # code, but keeping the end variable here in case it opens
                    # up ideas for code consolidation later.
                    
                    subcom = mooring.subcomponents[-att['end']]  # check what's on the end of the mooring

                    if isinstance(subcom, list):  # bridle case
                        print('This case not implemented yet')
                        breakpoint()
                    elif isinstance(subcom, Node):
                        if subcom.mpConn:
                            # TODO: get rel dist from connector to anchor
                            # for now, just assume 0 rel dist until anchor lug objects introduced
                            r_rel = [0,0,0]
                            subcom.mpConn.type = 1
                            # attach anchor body to subcom connector point
                            # anchor.mpAnchor.attachPoint(subcom.mpConn.number,r_rel) # COMMENT OUT FOR NOW, may return to this in the future
                            print('This case should not occur, kept in for future work')
                            breakpoint()
                        else:
                            # attach next section line to anchor
                            line = mooring.subcomponents[-att['end']+1]
                            anchor.mpAnchor.attachLine(line.mpLine.number, 
                                                       att['end'])
                        
                        # (the section line object(s) should already be attached to this point)
                    #TODO >>> still need to handle possibility of anchor bridle attachment, multiple anchor lugs, etc. <<<
                
                else:  # Original case with Subsystem
                    # # need to create "dummy" point to connect to anchor body
                    # point = self.ms.addPoint(1,anchor.r)
                    # # attach dummy point to anchor body
                    # anchor.mpAnchor.attachPoint(anchor.mpAnchor.number,[0,0,0])
                    # now attach dummy point to line
                    anchor.mpAnchor.attachLine(ssloc.number, att['end'])
                
                # Check for fancy case of any lugs (nodes) attached to the anchor
                if any([ isinstance(a['obj'], Node) for a in anchor.attachments.values()]):
                    print('Warning: anchor lugs are not supported yet')
                    breakpoint()
                
                # find associated platform and attach body to point (since not a shared line, should only be one platform with this mooring object)
                for platform in self.platformList.values(): # ii is index in dictionary, k is key (name) of platform
                    if j in platform.attachments: # j is key (name) of mooring object in anchor i checking if that same mooring object name is attached to platform k
                        PF = platform # platform object associated with mooring line j and anchor i
                        break
                
                # attach rB point to platform 
                if mooring.parallels:  # case with paralles/bridles
                    
                    # Look at end B object(s)
                    subcom = mooring.subcomponents[-1]
                
                    if isinstance(subcom, list):  # bridle case
                        for parallel in subcom:
                            subcom2 = parallel[-1]  # end subcomponent of the parallel path
                            
                            # Code repetition for the moment:
                            if isinstance(subcom2, Edge):
                                r = subcom2.attached_to[1].r # approximate end point...?
                                point = self.ms.addPoint(1, r)
                                PF.body.attachPoint(point.number, r-PF.r)
                                point.attachLine(subcom2.mpLine.number, 1)  # attach the subcomponent's line object end B
                                
                            elif isinstance(subcom2, Node):
                                r = subcom2.r # approximate end point...?
                                pnum = subcom2.mpConn.number
                                PF.body.attachPoint(pnum, r-PF.r)
                    
                    elif isinstance(subcom, Edge):
                        r = subcom.attached_to[1].r # approximate end point...?
                        point = self.ms.addPoint(1, r)
                        PF.body.attachPoint(point.number, r-PF.r)
                        point.attachLine(subcom.mpLine.number, 1)  # attach the subcomponent's line object end B
                        
                    elif isinstance(subcom, Node):
                        r = subcom.r # approximate end point...?
                        pnum = subcom.mpConn.number
                        PF.body.attachPoint(pnum, r-PF.r)
                        # (the section line object(s) should already be attached to this point)

                    
                else:  # normal serial/subsystem case
                    # add fairlead point
                    point = self.ms.addPoint(1,ssloc.rB)
                    # add connector info for fairlead point
                    # >>> MH: these next few lines might result in double counting <<<
                    point.m = self.ms.lineList[-1].pointList[-1].m 
                    point.v = self.ms.lineList[-1].pointList[-1].v
                    point.CdA = self.ms.lineList[-1].pointList[-1].CdA
                    # attach the line to point
                    point.attachLine(ssloc.number,1)
                    PF.body.attachPoint(point.number, ssloc.rB-PF.r) # attach to fairlead (need to subtract out location of platform from point for subsystem integration to work correctly)
        
        
        # Create and attach any shared lines or hybrid lines attached to buoys
        for mkey, mooring in self.mooringList.items(): # loop through all lines
            check = 1  # temporary approach to identify shared lines <<<
            for j in self.anchorList: # j is key (name) of anchor object
                if mkey in self.anchorList[j].attachments: # check if line has already been put in ms
                    check = 0 
                    break
            if check == 1: # mooring object not in any anchor lists
                # new shared line
                # create subsystem for shared line
   
                # set location of subsystem for simpler coding
                mooring.createSubsystem(case=mooring.shared, 
                    ms=self.ms)
                ssloc = mooring.ss
                
                # (ms.lineList.append is now done in Mooring.createSubsystem)
                
                # find associated platforms/ buoys
                att = mooring.attached_to
                
                # connect line ends to the body/buoy
                for ki in range(0,2):  # for each end of the mooring
                    if isinstance(att[ki],Platform):  # if it's attached to a platform
                        
                        platform = att[ki]
                        
                        if mooring.parallels:  # case with parallels/bridles
                            
                            # Look at end object(s)
                            subcom = mooring.subcomponents[-ki]
                        
                            if isinstance(subcom, list):  # bridle case
                                for parallel in subcom:
                                    subcom2 = parallel[-ki]  # end subcomponent of the parallel path
                                    
                                    # Code repetition for the moment:
                                    if isinstance(subcom2, Edge):
                                        r = subcom2.attached_to[ki].r # approximate end point...?
                                        point = self.ms.addPoint(1, r)
                                        platform.body.attachPoint(point.number, r-platform.r)
                                        point.attachLine(subcom2.mpLine.number, ki)  # attach the subcomponent's line object end
                                        
                                    elif isinstance(subcom2, Node):
                                        r = subcom2.r # approximate end point...?
                                        pnum = subcom2.mpConn.number
                                        platform.body.attachPoint(pnum, r-platform.r)
                            
                            elif isinstance(subcom, Edge):
                                r = subcom.attached_to[ki].r # approximate end point...?
                                point = self.ms.addPoint(1, r)
                                platform.body.attachPoint(point.number, r-platform.r)
                                point.attachLine(subcom.mpLine.number, ki)  # attach the subcomponent's line object end
                                
                            elif isinstance(subcom, Node):
                                r = subcom.r # approximate end point...?
                                pnum = subcom.mpConn.number
                                platform.body.attachPoint(pnum, r-platform.r)
                                # (the section line object(s) should already be attached to this point)

                            
                        else:  # normal serial/subsystem case
                            
                            if ki==0:
                                rEnd = mooring.rA
                            else:
                                rEnd = mooring.rB
                            
                            # add fairlead point A and attach the line to it
                            point = self.ms.addPoint(1, rEnd)
                            point.attachLine(ssloc.number, ki)
                            platform.body.attachPoint(point.number, rEnd-platform.r)

                    else:
                        # this end is unattached
                        pass
        
        
        # add in cables if desired
        if cables:
            for i in self.cableList:   
                # determine if suspended cable or not - having a static cable as a subcomponent means this is not a suspended cable
                for j,comp in enumerate(self.cableList[i].subcomponents):
                    # # check for all attachments (may be failurs enacted preventing all attachments) 
                    # don't make a subsystem for a joint - make a point                   
                    if isinstance(comp,Joint):
                        if not comp.mpConn:
                            comp.makeMoorPyConnector(self.ms)
                    elif isinstance(comp,StaticCable):
                        # don't make a subsystem for a static cable (yet...)
                        pass
                    elif isinstance(comp,DynamicCable):
                        # create subsystem for dynamic cable
                        comp.createSubsystem() 
                         
                        ssloc = comp.ss
                        ssloc.number = len(self.ms.lineList)+1 
                        # add subsystem to line list
                        self.ms.lineList.append(ssloc)
                        
                        # attach each end to correct bodies
                        attach = comp.attached_to #                                 
                        
                        if isinstance(attach[0],Joint):
                            # connect to joint at end A
                            if not comp.attached_to[0].mpConn:
                                comp.attached_to[0].makeMoorPyConnector(self.ms)
                            joint = comp.attached_to[0].mpConn
                            joint.attachLine(ssloc.number,0)
                        elif j==0: 
                            self.ms.addPoint(1,ssloc.rA)
                            self.ms.pointList[-1].attachLine(ssloc.number,0)
                            body = self.cableList[i].attached_to[0].body
                            body.attachPoint(len(self.ms.pointList),[ssloc.rA[0]-body.r6[0],ssloc.rA[1]-body.r6[1],ssloc.rA[2]])
                            
                        if isinstance(attach[1],Joint):
                            # connect to joint at end B
                            if not comp.attached_to[-1].mpConn:
                                comp.attached_to[-1].makeMoorPyConnector(self.ms)
                            joint = comp.attached_to[-1].mpConn
                            joint.attachLine(ssloc.number,1) 
                        elif j==len(self.cableList[i].subcomponents)-1: # last subcomponent could be first subcomponent
                            self.ms.addPoint(1,ssloc.rB)
                            self.ms.pointList[-1].attachLine(ssloc.number,1)
                            body = self.cableList[i].attached_to[-1].body
                            body.attachPoint(len(self.ms.pointList),[ssloc.rB[0]-body.r6[0],ssloc.rB[1]-body.r6[1],ssloc.rB[2]])                                  

        # initialize, solve equilibrium, and plot the system 
        self.ms.initialize()
        self.ms.solveEquilibrium()      
        
        # Plot array if requested
        if plt>0:
            if plt==1:
                self.ms.plot()
            else:
                settings = {}
                settings["linelabels"] = True
                settings["pointlabels"] = True                          
                self.ms.plot( **settings)
            
        # return the mooring system   
        return(self.ms)

    def getFLORISArray(self, config, turblist, windSpeeds = [], thrustForces = []):
        '''
        Sets up FLORIS interface and stores thrust/windspeed curve

        Parameters
        ----------
        config : str
            Filename for FLORIS wake input file
        turblist : list of str
            List of FLORIS turbine input files (if 1, use for all turbines)
        windSpeeds : list of float
            List of wind speeds for thrust curve
        thrustForces : list of float
            List of turbine thrust forces at each wind speed in windSpeeds

        Returns
        -------
        None.

        '''
        import floris
        from floris import FlorisModel

        
        # Setup FLORIS interface using base yaml file
        self.flow = FlorisModel(config)


        x = []
        y = []
        for pf in self.platformList.values():
                if pf.entity == 'FOWT':
                    x.append(pf.r[0])
                    y.append(pf.r[1])
        self.flow.set(layout_x=x, layout_y=y)

        #right now, point to FLORIS turbine yaml. eventually should connect to ontology
        self.flow.set(turbine_type= turblist)       
        self.flow.reset_operation() # Remove any previously applied yaw angles
       # Since we are changing the turbine type, make a matching change to the reference wind height
        self.flow.assign_hub_height_to_ref_height()
        
        #store ws and thrust data... eventually store for each turbine
        self.windSpeeds = windSpeeds
        self.thrustForces = thrustForces
        
    def getFLORISMPequilibrium(self, ws, wd, ti, cutin, hubht, plotting = True):
        '''
        Function to find array equilibrium with FLORIS wake losses and MoorPy platform offsets

        Parameters
        ----------
        ws : float
            Wind speed (m/s)
        wd : float
            Wind direction (heading direction in deg where due West is 0 and due North is 90)
        ti : float
            Turbulence intenstiy (input to floris)
        cutin : float
            Cut in wind speed
        hubht : float
            Hub height
        plotting : bool
            True plots wakes and mooring systems. The default is True.

        Returns
        -------
        winds : array of float
            Initial and updated wind speed at each turbine
        xpositions : array of float
            Initial and updated x position at each turbine
        ypositions : array of float
            Initial and updated y position at each turbine
        turbine_powers : array of float
            Final power at each turbine

        '''
        # Wind directional convention ??? 
        
        from scipy import interpolate
        import time
        
        if ws < min(self.windSpeeds) or ws > max(self.windSpeeds):
            return ValueError("Wind speed outside of stored range")
        
        #FLORIS inputs the wind direction as direction wind is coming from (where the -X axis is 0)
        self.flow.set(wind_directions = [-wd+270], wind_speeds = [ws], turbulence_intensities= [ti])
        
        # wind speed thrust curve for interpolation
        f = interpolate.interp1d(self.windSpeeds, self.thrustForces)
        
        #initialize list of wind speeds (1 wind speed for each turbine)
        turblist = [x.body for x in self.platformList.values() if x.entity.upper()=='FOWT']
        nturbs = len(turblist)
        ws = [ws] * nturbs
        
        winds = []
        xpositions = []
        ypositions = []
        
        #iterate twice through updating wind speeds/platform positions
        for n in range(0, 2):
            
            # interpolate thrust force from speed/thrust curve
            for i in range(0,nturbs):
                if ws[i] < cutin:
                    T = 0
                else:
                    T = f(ws[i])
                
                # apply thrust force/moments (split into x and y components)
                turblist[i].f6Ext = np.array([T*np.cos(np.radians(wd)), T*np.sin(np.radians(wd)), 0, T*np.cos(np.radians(wd))*hubht, T*np.sin(np.radians(wd))*hubht, 0])       # apply an external force on the body 
                
            
            #solve statics to find updated turbine positions
            self.ms.initialize()
            self.ms.solveEquilibrium(DOFtype='both')

            #update floris turbine positions and calculate wake losses
            self.flow.set(layout_x=[body.r6[0] for body in turblist], layout_y=[body.r6[1] for body in turblist])
            self.flow.run()
    
          
           
            #update wind speed list for RAFT
            ws = list(self.flow.turbine_average_velocities[0])
            winds.append(ws)
            xpositions.append([body.r6[0] for body in turblist])
            ypositions.append([body.r6[1] for body in turblist])


        #return FLORIS turbine powers (in order of turbine list)
        if min(self.flow.turbine_average_velocities[0]) > cutin:
            turbine_powers = self.flow.get_turbine_powers()[0]

           
        else:
            turbine_powers = np.zeros((1,4))
        
        
        if plotting:
            
            import floris.layout_visualization as layoutviz
            from floris.flow_visualization import visualize_cut_plane
            
            fmodel = self.flow
            
            # Create the plotting objects using matplotlib
            fig, ax = plt.subplots()

          
            layoutviz.plot_turbine_points(fmodel, ax=ax)
            layoutviz.plot_turbine_labels(fmodel, ax=ax)
            ax.set_title("Turbine Points and Labels")
            ax.set_xlabel('X (m)')
            ax.set_ylabel('Y (m)')
            
         
            
            #fmodel.set(wind_speeds=[wind_spd], wind_directions=[wind_dir], turbulence_intensities=[ti])
            horizontal_plane = fmodel.calculate_horizontal_plane(
                x_resolution=200,
                y_resolution=100,
                height=90.0,
            )
            
            # Plot the flow field with rotors
            fig, ax = plt.subplots()
            visualize_cut_plane(
                horizontal_plane,
                ax=ax,
                label_contours=False,
                title="Horizontal Flow with Turbine Rotors and labels",
            )
            ax.set_xlabel('X (m)')
            ax.set_ylabel('Y (m)')
            
            # Plot the turbine rotors
            layoutviz.plot_turbine_rotors(fmodel, ax=ax)
            
       
            #self.ms.plot2d(ax = ax, Yuvec = [0,1,0])
     
            #return turbines to neutral positions **** only done if plotting - this reduces runtime for AEP calculation
            for i in range(0, nturbs):
    
                turblist[i].f6Ext = np.array([0, 0, 0, 0, 0, 0])       # apply an external force on the body 
                
            #solve statics to find updated turbine positions
            # self.ms.initialize()
            # self.ms.solveEquilibrium(DOFtype='both')
            # self.ms.plot2d(ax = ax, Yuvec = [0,1,0], color = 'darkblue')
        
        
        winds = np.array(winds)
        xpositions = np.array(xpositions)
        ypositions = np.array(ypositions)
        return(winds,xpositions, ypositions, turbine_powers)  

    def calcoffsetAEP(self, windrose, ti, cutin, hubht):
        '''
        Function to calculate AEP in FLORIS with moorpy platform offsets

        Parameters
        ----------
        windrose : str
            Wind rose filename
        ti : float
            Turbulence intensity (input to FLORIS for all cases)
        cutin : float
            Tut in wind speed
        hubht : float
            Hub height

        Returns
        -------
        aeps : list of floats
            AEP for each turbine in farm

        '''

        
        # remove windrose entries below cut in or with zero frequency
        wr = np.loadtxt(windrose, delimiter=',', skiprows = 1)
        wr = wr[wr[:,0] >= cutin]
        wr = wr[wr[:,2] > 0]

        # iterate through windrose directions/speeds
        for ind in range(0, len(wr)):
            
            ws = wr[ind, 0]
            wd = wr[ind, 1]
            fq = wr[ind, 2]
            
            # solve equilibrium with moorpy and floris at given wind direction/speed
            winds,xpositions, ypositions, turbine_powers = self.getFLORISMPequilibrium(ws, wd, ti, cutin, hubht, plotting = False)
            if ind == 0:
                powers = turbine_powers
            else:
                powers = np.vstack([powers,turbine_powers])

        aeps = np.matmul(wr[:,2], powers) * 365 * 24
        return(aeps)

    
    def getRAFT(self,RAFTDict):
        '''Create a RAFT object and store in the project class
        
        Parameters
        ---------
        RAFTDict : dictionary
            Provides information needed to create RAFT dictionary. Need settings, cases, turbines, and platforms sub-dictionaries.
            See RAFT documentation for requirements for each sub-dictionary
        '''
        print('Creating RAFT object')
            
        # create RAFT model if necessary components exist
        if 'platforms' in RAFTDict or 'platform' in RAFTDict:
            # set up a dictionary with keys as the table names for each row (ease of use later)
            RAFTable = [dict(zip(RAFTDict['array']['keys'], row)) for row in RAFTDict['array']['data']]

            if 'ID' in RAFTDict['array']['keys']:
                IDindex = np.where(np.array(RAFTDict['array']['keys'])=='ID')[0][0]
                IDdata  = [row[IDindex] for row in RAFTDict['array']['data']]
                RAFTDict['array']['keys'].pop(IDindex) # remove key for ID because this doesn't exist in RAFT array table
                reinsert = True
            if 'topsideID' in RAFTDict['array']['keys']:
                ts_loc = RAFTDict['array']['keys'].index('topsideID')
                RAFTDict['array']['keys'][ts_loc] = 'turbineID'

            mooringIDindex = RAFTDict['array']['keys'].index('mooringID')
            headindex = RAFTDict['array']['keys'].index('heading_adjust')
            tsIDindex = RAFTDict['array']['keys'].index('turbineID')
            nonturbID = []

            for i in range(0,len(RAFTDict['array']['data'])):
                entity = self.platformList[RAFTDict['array']['data'][i][IDindex]].entity
                RAFTDict['array']['data'][i].pop(IDindex) # remove ID column because this doesn't exist in RAFT array data table
                if entity.upper() != 'FOWT':
                    val = deepcopy(RAFTDict['array']['data'][i][tsIDindex])
                    if val>0:
                        nonturbID.append(val)
                    RAFTDict['array']['data'][i][tsIDindex] = 0
                RAFTDict['array']['data'][i][mooringIDindex] = 0 # make mooringID = 0 (mooring data will come from MoorPy)
                #RAFTDict['array']['data'][i][headindex] = - RAFTDict['array']['data'][i][headindex] # convert heading to cartesian from compass
            # adjust topside ids that are turbines as necessary
            for turb in RAFTDict['array']['data']:
                for ti in nonturbID:
                    if turb[tsIDindex] > ti:
                        turb[tsIDindex] -= 1

            

            # create empty mooring dictionary
            RAFTDict['mooring'] = {}
            #RAFTDict['mooring']['currentMod']
            # create raft model
            from raft.raft_model import Model
            self.array = Model(RAFTDict)
            # create dictionary of dictionaries of body hydrostatics for MoorPy bodies
            bodyInfo = {}
            for i,body in enumerate(self.array.fowtList):
                # set position (required before you can calcStatics)
                body.setPosition([RAFTable[i]['x_location'],RAFTable[i]['y_location'],0,0,0,0])
                # get body hydrostatics info for MoorPy bodies
                body.calcStatics()
                # populate dictionary of body info to send to moorpy
                if 'ID' in RAFTable[i]:
                    self.platformList[RAFTable[i]['ID']].dd['hydrostatics'] = {'m':body.m,'rCG':body.rCG,'v':body.V,'rM':body.rM,'AWP':body.AWP}
            # create moorpy array if it doesn't exist
            if not self.ms:
                self.getMoorPyArray()
            # assign moorpy array to RAFT object
            if len(self.ms.lineList)>0:
                self.array.ms = self.ms
            self.array.ms_tol = .05 # add in solveEquilibrium tolerance
            self.array.moorMod = 0
            # connect RAFT fowt to the correct moorpy body
            for i in range(0,len(self.platformList)): # do not include substations (these are made last)
                self.array.fowtList[i].body = self.ms.bodyList[i]

            # Reinsert 'ID' key and data back into RAFTDict
            if reinsert:
                RAFTDict['array']['keys'].insert(IDindex, 'ID')  # Reinsert 'ID' key at its original position
                for i, row in enumerate(RAFTDict['array']['data']):
                    row.insert(IDindex, IDdata[i])  # Reinsert 'ID' data into each row            
        else:
            raise Exception('Platform(s) must be specified in YAML file')


    def getMarineGrowth(self,mgDict_list=None,buoy_mg=None, lines='all',tol=2,display=False):
        '''Calls the addMarineGrowth mooring object method for the chosen mooring objects
           and applies the specified marine growth thicknesses at the specified depth ranges
           for the specified marine growth densities.

        Parameters
        ----------
        mgDict_list : list, optional
            Provides marine growth thicknesses and the associated depth ranges
            Default is None, which triggers the use of the marine growth dictionary saved in the project class
            'density' entry is optional. If no 'density' entry is created in the dictionary, addMarineGrowth defaults to 1325 kg/m^3
                    *In order from sea floor to surface*
                    example, if depth is 200 m: - {'thickness':0.00,'lowerRange':-200,'upperRange':-100, 'density':1320}
                                                - {'thickness':0.05,'lowerRange':-100,'upperRange':-80, 'density':1325}
                                                - {'thickness':0.1,'lowerRange':-80,'upperRange':0, 'density':1325}

        buoy_mg: list, optional
            Provides marine growth thicknesses for buoyancy sections of cables, irrespective of depth.
            Default is None, which triggers the use of the marine_growth_buoys list saved in the project class
        lines : list, optional
            List of keys from self.mooringList or self.cableList to add marine growth to. For cables, each entry must be a list
            of length 2 with the cable id in the first position and the subcomponent number of the dynamic cable in the 2nd position.
            Default is the string 'all', which triggers the addition of marine growth to all mooring lines.
                Ex: lines = ['fowt1a',['suspended_cable10',0]] to add mg just to mooring line 'fowt1a' 
                and the first subcomponent of cable 'suspended_cable10'
                
        tol : float, optional [m]
            Tolerance for marine growth cutoff depth values. Default is 2 m.
        display : bool, optional
            Whether or not to print out difference between expected and actual depth of change 
            for marine growth thicknesses
        
        Returns
        -------
        None.

        '''
        # make sure moorpy system exists already
        if not self.ms:
            if self.cableList:
                self.getMoorPyArray(cables=1)
            else:
                self.getMoorPyArray()
        # get marine growth dictionary
        if not mgDict_list:
            mgDict_list = self.marine_growth
        if not buoy_mg:
            buoy_mg= self.marine_growth_buoys
        
        # get indices of lines to add marine growth
        if lines == 'all':
            idx = []
            for i in self.mooringList:
                idx.append(i)
            for i in self.cableList:
                for j,cab in enumerate(self.cableList[i].subcomponents):
                    if isinstance(cab,DynamicCable):
                        idx.append([i,j])
        else:
            idx = lines
        
        for ii,i in enumerate(idx):

            if isinstance(i,list):
                # this is a cable
                self.cableList[i[0]].subcomponents[i[1]].addMarineGrowth(mgDict_list, buoy_mg, updateDepths=True, tol=tol)
                
            else:
                self.mooringList[i].addMarineGrowth(mgDict_list, updateDepths=True, tol=tol)

               
            # assign the newly created subsystem into the right place in the line list
            if isinstance(i,list):
                self.ms.lineList[ii] = self.cableList[i[0]].subcomponents[i[1]].ss
                
            else:
                self.ms.lineList[ii] = self.mooringList[i].ss
                
            self.ms.lineList[ii].number = ii+1
                


    def updateUniformArray(self,n_rows,pf_rows,spacingXY,grid_rotang=0,grid_skew_x=0,grid_skew_y=0,grid_trans_x=0,grid_trans_y=0,phis=[0,0],center=[0,0]):
        '''
        Function to update the array spacing, angles, platform locations and headings, etc for a uniform array
        
        Parameters
        ----------
        n_rows : int
            Number of rows
        pf_rows : int or list
            Number of platforms in each row. If an int, the number of platforms is constant (essentially equivalent to number of columns). 
            If a list, each entry represents the number of platforms for an individual row
        spacingXY : list
            2 entry list that provides the spacing in the x and y direction respectively [m]. If an integer is provided 
            that is less than 20, it is assumed the value given is in multiples of the turbine diameter instead of [m].
            If the x and y entries are lists themselves, this specifies the spacing between each row/col individually
        grid_rotang : float
            Rotation of the grid [deg] from 0 North, optional. Default is 0
        grid_skew_x : float
            Angle of the parallelogram between adjacent rows [deg], optional. Default is 0
        grid_skew_y : float
            Angle of the parallelogram between adjacent columns [deg], optional. Default is 0
        grid_trans_x : float
            x offset [m] from Western-most boundary point (if no boundary in project, x offset from [0,0] origin) for all turbine positions.
        grid_trans_y : float
            y offset [m] from Southern-most boundary point (if no boundary in project, y offset from [0,0] origin) for all turbine positions.
        phis : list
            Platform headings for rows. The length of the list dictates the repeat pattern 
            (a length of 2 indicates that odd rows will have phis[0] heading and even rows will have phis[1] heading)
            If each list entry is also a list, this sublist indicates the repeat pattern of phi within the row from west to east
        center : list
            x and y coordinate of array layout center, around which any rotations/skews would occur. Optional, default is [0,0]
        
        Returns
        -------
        None.

        '''
        r = np.zeros([len(self.platformList),3])   # xy location of each platform
        r_no_off = np.zeros([len(self.platformList),3]) # xy location of each platform not considering offsets
        phi = np.zeros(len(self.platformList)) # rotation heading of each platform
        D = 0
        for i,turb in enumerate(self.turbineList.values()):
            D += turb.D 
        D = D/i/2
        # get a list of platform names for use later
        pfIDs = [None]*len(self.platformList)
        
        for i,pf in enumerate(self.platformList.values()):
            if pf.rc:
                # convert rc to number (as if array were hstacked)
                numh = pf.rc[0]*n_rows + pf.rc[1]
                pfIDs[numh] = pf.id              
            else:    
                # if not row-col indices, assume dictionary order is in rows starting from north west corner
                pfIDs[i] = pf.id

        # calculate grid rotation and skew (from fadesign layout_helpers create_rotated_layout function)
        # Shear transformation in X
        # Calculate trigonometric values
        cos_theta = np.cos(np.radians(-grid_rotang))
        sin_theta = np.sin(np.radians(-grid_rotang))
        tan_phi_x = np.tan(np.radians(grid_skew_x))
        tan_phi_y = np.tan(np.radians(grid_skew_y))

        # Compute combined rotation and skew transformation matrix
        transformation_matrix = np.array([[cos_theta - sin_theta*tan_phi_y, cos_theta*tan_phi_x - sin_theta],
                                          [sin_theta + cos_theta*tan_phi_y, sin_theta*tan_phi_x + cos_theta]])
        
        west_bound = None
        north_bound = None
        # determine if a boundary exists for the project
        if self.boundary.size > 0:
            # find western-most boundary point
            west_bound = self.boundary[0][0]
            north_bound = self.boundary[0][1]
            for xy in self.boundary:
                if xy[0] < west_bound:
                    west_bound = xy[0]
                if xy[1] > north_bound:
                    north_bound = xy[1]
        
        # set up grid
        ct = 0 # counter variable
        p = 0  # counter variable for phi rows
        xx = 0 # store previous row location
        yy = 0 # store previous column location
        for i in range(0,n_rows):
            # find number of platforms in the row
            if isinstance(pf_rows,list):
                num = pf_rows[i]
            else:
                num = pf_rows     
            
            for j in range(0,num):
                # find spacing from previous rows + cols
                # if first row or first col, spacing = 0
                if j == 0:
                    spacing_x = 0 
                else:
                    if isinstance(spacingXY[0],list):
                        spacing_x = spacingXY[0][j-1]
                    else:
                        spacing_x = spacingXY[0]
                    if spacing_x < 20:
                        spacing_x = spacing_x*D
                if i == 0:
                    spacing_y = 0 
                else:
                    if isinstance(spacingXY[1],list):
                        spacing_y = spacingXY[1][i-1]
                    else:
                        spacing_y = spacingXY[1]                    
                    if spacing_y < 20:
                        spacing_y = spacing_y*D
                
                # assign location of platform in the row (temporary, does not include rotation or skew)
                r_no_off[ct][0] = spacing_x + xx
                r_no_off[ct][1] = -spacing_y + yy
                
                xx = r_no_off[ct][0]
                                               
                if west_bound:
                    # set offsets from north and west boundaries
                    xoff = west_bound + grid_trans_x
                    yoff = north_bound - grid_trans_y
                else:
                    # use grid_trans_x and grid_trans_y as offsets
                    xoff = grid_trans_x
                    yoff = grid_trans_y
                    
                r[ct][0] = r_no_off[ct][0] + xoff + center[0]
                r[ct][1] = r_no_off[ct][1] + yoff + center[1]
                
                # include grid rotation around center
                x,y = np.dot(transformation_matrix,[r[ct][0]-center[0],r[ct][1]-center[1]])
                
                # get platform heading
                if p == len(phis):
                    p = 0

                if isinstance(phis[p],list):
                    phi[ct] = phis[p][j]
                else:
                    phi[ct] = phis[p]
                    
                # call platform function to set location, phi, moorings, cables, and anchors
                self.platformList[pfIDs[ct]].setPosition([x,y],heading=phi[ct],degrees=True)
                self.platformList[pfIDs[ct]].rc = [i,j]
                # # update cable lengths as needed (assumes the attachments are correct)
                # cabs = self.platformList[pfIDs[ct]].getCables()
                # for cab in cabs.values():
                #     # determine what the cable is connected to
                #     rr = [0,0]
                #     for k,att in enumerate(cab.attached_to):
                #         if isinstance(att,Platform):
                #             rr[k] = att.r
                #     newSpan = np.sqrt((rr[0][0]-rr[1][0])**2 + (rr[0][1]-rr[1][1])**2)
                #     cab.updateSpan(newSpan)
                #     # update cable end locations
                #     cab.reposition()
                # update anchor locations
                anchs = self.platformList[pfIDs[ct]].getAnchors()
                for anch in anchs.values():
                    for att in anch.attachments:
                        if isinstance(anch.attachments[att]['obj'],Mooring):
                            anch.r = anch.attachments[att]['obj'].rA
                            # update depth of anchor as needed
                            aDepth = self.getDepthAtLocation(anch.r[0],anch.r[1])
                            anch.attachments[att]['obj'].rA[2] = -aDepth                                                                       
                
                # increase counter
                ct += 1
           
            xx = 0
            yy = r_no_off[ct-1][1]
            # increase phi counter
            p += 1
                
        # update moorpy
        self.getMoorPyArray(plt=1,cables=1)
        
    def duplicate(self,pf, r=None,heading=None):
        '''
        Function to duplicate a platform object and all
        of its associated moorings and anchors (NOT CABLES)

        Parameters
        ----------
        pf : Platform object
            Platform object to duplicate
        r : list, optional
            Location of new platform object. Default is None,
            in which case platform is not moved

        Returns
        -------
        pf2 : Platform object
            Duplicated object

        '''
        from copy import deepcopy
        # get name for new platform based on length of platformList
        if self.platformList:
            lp = len(self.platformList)
        else:
            lp = 0
        newid = 'fowt'+str(lp)
        alph = list(ascii_l)
        
        # copy platform object and its attachments and disconnect from attachments
        pf2 = deepcopy(pf)
        pf2.id = newid
        self.platformList[newid] = pf2
        count = 0 
        
        # first check for fairlead objects
        fairs = True if any([isinstance(att['obj'],Fairlead) for att in pf.attachments.values()]) else False
        if fairs:
            for att in pf.attachments.values():
                if isinstance(att['obj'],Fairlead):
                    r_rel = att['r_rel']
                    if att['obj'].attachments:
                        for val in att['obj'].attachments.values():
                            moor = val['obj'].part_of
                            endB = 1
                            # grab all info from mooring object
                            md = deepcopy(moor.dd)
                            mhead = moor.heading
                            # detach mooring object from platform
                            pf2.detach(moor,end=endB)
                            pf2.detach(att['obj'])
                            # create new mooring object
                            newm = Mooring(dd=md,id=newid+alph[count])
                            self.mooringList[newm.id] = newm
                            newm.heading = mhead
                            # check if fairlead
                            # for con in newm.subcons_B:
                            #     if 
                            # attach to platform
                            fl = self.addFairlead(platform=pf2,r_rel=r_rel,mooring=newm,id=att['obj'].id)
                            # grab info from anchor object and create new one
                            ad = deepcopy(moor.attached_to[1-endB].dd)
                            newa = Anchor(dd=ad,id=newid+alph[count])
                            self.anchorList[newa.id] = newa
                            # attach anchor to mooring
                            newm.attachTo(newa,end=1-endB)
                            pf2.setPosition(r,heading=heading,project=self)
                            zAnew, nAngle = self.getDepthAtLocation(newm.rA[0], newm.rA[1], return_n=True)
                            newm.rA[2] = -zAnew
                            newm.dd['zAnchor'] = -zAnew
                            newa.r = newm.rA
                            
                            count += 1
                            
                    else:
                        moor=None
                    
                elif isinstance(att['obj'],Turbine):
                    pf2.detach(att['obj'])
                    turb = deepcopy(att['obj'])
                    turb.id = newid+'turb'
                    self.turbineList[turb.id] = turb
                    pf2.attach(turb)
                    
                elif isinstance(att['obj'],Cable):
                    # could be cable, just detach for now
                    pf2.detach(att['obj'],att['end'])
        else:
            for att in pf.attachments.values():
                if isinstance(att['obj'],Mooring):
                    if att['end'] == 'a':
                        endB = 0 
                    else:
                        endB = 1
                    # grab all info from mooring object
                    md = deepcopy(att['obj'].dd)
                    mhead = att['obj'].heading
                    # detach mooring object from platform
                    pf2.detach(att['obj'],end=endB)
                    # create new mooring object
                    newm = Mooring(dd=md,id=newid+alph[count])
                    self.mooringList[newm.id] = newm
                    newm.heading = mhead
                    pf2.attach(newm,end=endB)
                    # grab info from anchor object and create new one
                    ad = deepcopy(att['obj'].attached_to[1-endB].dd)
                    newa = Anchor(dd=ad,id=newid+alph[count])
                    self.anchorList[newa.id] = newa
                    # attach anchor to mooring
                    newm.attachTo(newa,end=1-endB)
                    newm.reposition(r_center=r,project=self)
                    zAnew, nAngle = self.getDepthAtLocation(newm.rA[0], newm.rA[1], return_n=True)
                    newm.rA[2] = -zAnew
                    newm.dd['zAnchor'] = -zAnew
                    newa.r = newm.rA
                    
                    count += 1
                    
                elif isinstance(att['obj'],Turbine):
                    pf2.detach(att['obj'])
                    turb = deepcopy(att['obj'])
                    turb.id = newid+'turb'
                    self.turbineList[turb.id] = turb
                    pf2.attach(turb)
                    
                elif isinstance(att['obj'],Cable):
                    # could be cable, just detach for now
                    pf2.detach(att['obj'],att['end'])
                
        
        # reposition platform as needed
        pf2.setPosition(r,heading=heading,project=self)
 
        
        
        # delete body object from pf2
        pf2.body = None
        
        return(pf2)
            
        
    def addPlatformMS(self,ms,r=[0,0,0]):
        '''
        Create a platform object, along with associated mooring and anchor objects
        from a moorpy system
        Currently only works for regular (non-shared) moorings.

        Parameters
        ----------
        ms : moorpy system, optional
            Moorpy system representing the platform, its moorings, and its anchors. The default is None.
        
        Returns
        -------
        new platform object

        '''
        # create platform, moorings, and anchors from ms
        ix = len(self.platformList)
        # check there is just one body
        if len(ms.bodyList) > 1:
            raise Exception('This function only works with a 1 body system')
            
        # switch to subsystems if lineList doesn't already have them
        if not isinstance(ms.lineList[0],mp.subsystem.Subsystem):
            from moorpy.helpers import lines2ss
            lines2ss(ms)
            
        # get lines attached to platform and headings
        
        mhead = []
        mList = []
        endB = []
        count = 0
        pfid = 'fowt'+str(ix)
        alph = list(ascii_l)
        for point in ms.bodyList[0].attachedP:
            for j,line in enumerate(ms.pointList[point-1].attached):
                md = {'subcomponents':[]} # start set up of mooring design dictionary
                rA = ms.lineList[line-1].rA
                rB = ms.lineList[line-1].rB
                pfloc = ms.bodyList[0].r6
                if ms.pointList[point-1].attachedEndB[j]:
                    vals = rA[0:2]-rB[0:2]
                    zFair = rB[2]
                    rFair = np.hypot(rB[0]-pfloc[0],rB[1]-pfloc[1])
                    endB.append(1)
                else:
                    vals = rB[0:2]-rA[0:2]
                    zFair = rA[2]
                    rFair = np.hypot(rA[0]-pfloc[0],rA[1]-pfloc[1])
                    endB.append(0)
                    
                # pull out mooring line info
                md['rad_fair'] = rFair
                md['z_fair'] = zFair
                md['span'] = np.hypot(vals[0],vals[1])
                if not endB[-1]:
                    md['zAnchor'] = -self.getDepthAtLocation(rA[0],rA[1])
                else:
                    md['zAnchor'] = -self.getDepthAtLocation(rB[0],rB[1])
                            
                            
                
                for k,sline in enumerate(ms.lineList[line-1].lineList):
                    # add section and connector info
                    spt = ms.lineList[line-1].pointList[k]
                    md['subcomponents'].append({'m':spt.m,'v':spt.v,'Ca':spt.Ca,'CdA':spt.CdA})
                    md['subcomponents'].append({'type':sline.type})
                    md['subcomponents'][-1]['L'] = sline.L
                    
                spt = ms.lineList[line-1].pointList[k+1]
                md['subcomponents'].append({'m':spt.m,'v':spt.v,'Ca':spt.Ca,'CdA':spt.CdA})
                mhead.append(90 - np.degrees(np.arctan2(vals[1],vals[0])))
                mList.append(Mooring(dd=md,id=pfid+alph[count]))
                mList[-1].heading = mhead[-1]
                self.mooringList[mList[-1].id] = mList[-1]
                
                # pull out anchor info
                for pt in ms.pointList:
                    if line in pt.attached:
                        loc = np.where([x==line for x in pt.attached])[0][0]
                        if pt.attachedEndB[loc]:
                            ad = {'design':{}}
                            ad['design']['m'] = pt.m
                            ad['design']['v'] = pt.v
                            ad['design']['CdA'] = pt.CdA
                            ad['design']['Ca'] = pt.Ca
                            if 'anchor_type' in pt.entity:
                                ad['type'] = pt.entity['anchor_type']
                            self.anchorList[mList[-1].id] = Anchor(dd=ad,r=pt.r,id=mList[-1].id)
                            self.anchorList[mList[-1].id].attach(mList[-1],end=1-endB[-1])
                            # reposition mooring and anchor
                            mList[-1].reposition(r_center=r,project=self)
                            zAnew = self.getDepthAtLocation(mList[-1].rA[0], 
                                                            mList[-1].rA[1])
                            mList[-1].rA[2] = -zAnew
                            mList[-1].dd['zAnchor'] = -zAnew
                            mList[-1].z_anch = -zAnew
                            self.anchorList[mList[-1].id].r = mList[-1].rA
                        
                count += 1
                            
                
        # add platform at ms.body location       
        self.platformList[pfid] = Platform(pfid, r=ms.bodyList[0].r6[0:3],
                                                     # mooring_headings=mhead,
                                                     rFair=rFair, zFair=zFair)
        # attach moorings
        for i,moor in enumerate(mList):
            self.platformList[pfid].attach(moor,end=endB[i])
            
        # update location
        self.platformList[pfid].setPosition(r=r,project=self)
            
        return(self.platformList[pfid])
            
    def addPlatformConfig(self,configDict,r=[0,0]):
        '''
        Add a platform, anchors, and mooring lines based on a configuration dictionary
        
        Parameters
        ----------
        configDict : dict
            Dictionary describing design of platform, moorings, and anchors
            The dictionary layout is as follows:
                platform:
                    rFair: 
                    zFair:
                    moor_headings:
                    platform_heading:
                mooring_config:
                    span:
                    segment list: # in order from end A to end B
                        - <segment name>:
                            material: # polyester,chain,etc
                            length:
                            d_nom: # can list material properties below this, or it will be imported from MoorProps
                                    # based on the material type and nominal diameter
                        - <c
                        - <segment name>:
                            material: 
                            length:
                            d_nom:
                anchor_config:
                    geometry:
                    type:
        r : list, optional
            x,y location of platform. The default is [0,0]
                        
        Returns
        -------
        pf : new platform object
                    
        '''
        
        if self.platformList:
            lp = len(self.platformList)
        else:
            lp = 0
            
        pfid = 'fowt'+str(lp)
        alph = list(ascii_l)
            
        # pull out platform info
        pfinfo = configDict['platform']
        
        # create platform object
        self.platformList[pfid] = Platform(pfid, r=r,
                                           # mooring_headings=pfinfo['mooring_headings'],
                                           rFair=pfinfo['rFair'], zFair=pfinfo['zFair'],
                                           phi=pfinfo['platform_heading'])
        
        # pull out mooring info
        minfo = configDict['mooring_config']
        # create mooring objects
        for i in range(len(pfinfo['mooring_headings'])):
            head = pfinfo['mooring_headings'][i]+pfinfo['platform_heading']
            md = {'span':minfo['span'],'subcomponents':[]}
      
    def arrayWatchCircle(self,plot=False, ang_spacing=45, RNAheight=150,
                         shapes=True,thrust=1.95e6,SFs=True,moor_envelopes=True, 
                         moor_seabed_disturbance=False, DAF=1):
        '''
        Method to get watch circles on all platforms at once

        Parameters
        ----------
        plot : bool, optional
            Controls whether to plot watch circles at the end. The default is False.
        ang_spacing : int/float, optional
            Spacing between angles to check the platform offsets at. The default is 45.
        RNAheight : int/float, optional
            Height of RNA from mean sea level. The default is 150.
        shapes : bool, optional
            Controls whether to create shapely objects. The default is True.
        thrust : float, optional
            Thrust force on turbine. The default is 1.95e6 (thrust on IEA 15 MW reference
                                                            turbine at rated wind speed.
        SFs : bool, optional
            Controls whether to output safety factors of moorings and cables,
            and calculate anchor loads. The default is True.
        moor_envelopes : bool, optional
            Controls whether to create and store motion envelopes for moorings
        DAF : float, optional
            Dynamic amplification factor. Multiplies max forces by the DAF to account for 
            under-estimation of loads when using static method

        Returns
        -------
        x : np.array
            matrix of platform x-locations for watch circles
        y : np.array
            matrix of platform y-locations for watch circles
        maxVals : dict
            dictionary of safety factors for mooring line tensions for each turbine

        '''
   
        # get angles to iterate over
        angs = np.arange(0,360+ang_spacing,ang_spacing)
        n_angs = len(angs)
        
        # lists to save info in  
        minTenSF = [None]*len(self.mooringList)
        CminTenSF = [None]*len(self.cableList)
        minCurvSF = [None]*len(self.cableList)
        F = [None]*len(self.anchorList) 
        x = np.zeros((len(self.platformList),n_angs))
        y = np.zeros((len(self.platformList),n_angs))
        
        info = {'analysisType': 'quasi-static (MoorPy)',
                'info': f'determined from arrayWatchCircle() with DAF of {DAF}'}
        
        lBots = np.zeros(len(self.mooringList))  # initialize for maximum laid length per mooring
        if not self.ms:
            self.getMoorPyArray()
             
        # apply thrust force to platforms at specified angle intervals
        for i,ang in enumerate(angs):
            print('Analyzing platform offsets at angle ',ang)
            fx = thrust*np.cos(np.radians(ang))
            fy = thrust*np.sin(np.radians(ang))
            
            # add thrust force and moment to the body
            for pf in self.platformList.values():
                if pf.entity == 'FOWT':
                    pf.body.f6Ext = np.array([fx, fy, 0, fy*RNAheight, fx*RNAheight, 0])       # apply an external force on the body [N]  
                else:
                    pass                     
            # solve equilibrium 
            self.ms.solveEquilibrium3(DOFtype='both')
            
            # save info if requested
            if SFs:
                # get loads on anchors (may be shared)
                for j,anch in enumerate(self.anchorList.values()):
                    F2 = anch.mpAnchor.getForces()*DAF # add up all forces on anchor body
                    H = np.hypot(F2[0],F2[1]) # horizontal force
                    T = np.sqrt(F2[0]**2+F2[1]**2+F2[2]**2) # total tension force
                    if F[j] is None or T>np.sqrt(F[j][0]**2+F[j][1]**2+F[j][2]**2):
                        F[j] = F2 # max load on anchor                         
                        # save anchor load information
                        anch.loads['Hm'] = H
                        anch.loads['Vm'] = F[j][2]
                        anch.loads['thetam'] = np.degrees(np.arctan(anch.loads['Vm']/anch.loads['Hm'])) #[deg]
                        anch.loads['mudline_load_type'] = 'max'
                        anch.loads.update(info)
                            
                # get tensions on mooring line
                for j, moor in enumerate(self.mooringList.values()):
                    lBot = 0
                    moor.updateTensions(DAF=DAF)
                    moor.updateSafetyFactors(info=info)
                    if moor_seabed_disturbance:
                        for sec in moor.sections():
                            lBot += sec.mpLine.LBot
                        lBots[j] = max(lBots[j], lBot)

                                
                # get tensions and curvature on cables
                for j,cab in enumerate(self.cableList.values()):
                    dcs = [a for a in cab.subcomponents if isinstance(a,DynamicCable)] # dynamic cables in this cable 
                    cab.updateTensions(DAF=DAF)
                    cab.updateSafetyFactors(info=info)
                    minCurvSF[j] = [None]*len(dcs)
                    CminTenSF[j] = [None]*len(dcs)
                    if dcs[0].ss:
                        
                        for jj,dc in enumerate(dcs):
                            CminTenSF[j][jj] = dc.safety_factors['tension']
                            dc.ss.calcCurvature()
                            mCSF = dc.ss.getMinCurvSF()
                            if not minCurvSF[j][jj] or minCurvSF[j][jj]>mCSF:
                                minCurvSF[j][jj] = mCSF
                                dc.safety_factors['curvature'] = minCurvSF[j][jj]
                        
                
            # save location of each platform for envelopes
            for k, pf in enumerate(self.platformList.values()):                       
                x[k,i] = pf.body.r6[0]     
                y[k,i] = pf.body.r6[1]
                
                    
        for k, body in enumerate(self.platformList.values()):
            # save motion envelope in the correct platform 
            body.envelopes['mean'] = dict(x=x[k,:], y=y[k,:])
        
            if shapes:  # want to *optionally* make a shapely polygon
                from shapely import Polygon
                body.envelopes['mean']['shape'] = Polygon(list(zip(x[k,:],y[k,:])))
            
        for body in self.ms.bodyList:
            body.f6Ext = np.array([0, 0, 0, 0, 0, 0])
        self.ms.solveEquilibrium3(DOFtype='both')  
        
        if moor_envelopes:
            for moor in self.mooringList.values():
                moor.getEnvelope()
        
        # estimate seabed disturbance from lines dragging on the seabed: assuming circular watch circles, use maximum offset and maximum laid length of the line.
        if moor_seabed_disturbance:
            for j, moor in enumerate(self.mooringList.values()):
                # compute maximum offset of the attached platform
                Rmaxx = 0
                for att in moor.attached_to:
                    if isinstance(att, Platform): # not sure if I also have to check the topside of that platform to make sure it's FOWT - Rudy.
                        cntr = att.r
                        x_cntr = att.envelopes['mean']['x'] - cntr[0]
                        y_cntr = att.envelopes['mean']['y'] - cntr[1]
                        R = np.sqrt(x_cntr**2 + y_cntr**2)
                        Rmax = max(R)
                        Rmaxx = max([Rmaxx, Rmax])
                
                # compute disturbed area from the laid length 
                moor.env_impact['disturbedSeabedArea'] += (lBots[j]**2 * Rmaxx)/(moor.rad_anch)
                            
        if SFs:
            maxVals = {'minTenSF':minTenSF,'minTenSF_cable':CminTenSF,'minCurvSF':minCurvSF,'maxF':F}# np.vstack((minTenSF,CminTenSF,minCurvSF,minSag))    
            return(x,y,maxVals)     
        else:
            return(x,y)
 
    def getArrayCost(self):
        '''
        Function to sum all available costs for the array components and produce a 
        spreadsheet with itemized costs for each component.

        Returns
        -------
        total cost

        '''
        total_cost = 0
        anch_costs = {}
        pf_costs = {}
        mooring_costs = {}
        cable_costs = {}
        turbine_costs = {}
        substation_costs = {}
        # anchors
        anch_cost = 0
        for anch in self.anchorList.values():
            anch_cost += anch.getCost()
            
            anch_costs[anch.id] = anch.cost
        # maxrows_anch = max([len(anch.cost) for anch in self.anchorList.values()])
        #headings_anch = set([anch.cost.keys() for anch in self.anchorList.values()])
        # platforms
        pf_cost = 0
        for pf in self.platformList.values():
            if pf.cost:
                pf_cost += sum(pf.cost.values())
            pf_costs[pf.id] = pf.cost
        # maxrows_pf = max([len(pf.cost) for pf in self.platformList.values()])
        # moorings
        moor_cost = 0
        for moor in self.mooringList.values():
            moor_cost += moor.getCost()
            mooring_costs[moor.id] = moor.cost
        # maxrows_moor = max([len(moor.cost) for moor in self.mooringList.values()])
        # cables
        cab_cost = 0
        for cab in self.cableList.values():
            cab_cost += cab.getCost()
            cable_costs[cab.id] = cab.cost
        # maxrows_cab = max([len(cab.cost) for cab in self.cableList.values()])
        # turbine
        turb_cost = 0
        for turb in self.turbineList.values():
            if turb.cost:
                turb_cost += sum(turb.cost.values())
            turbine_costs[turb.id] = turb.cost
        # maxrows_turb = max([len(turb.cost) for turb in self.turbineList.values()])
        # substation
        oss_cost = 0
        for oss in self.substationList.values():
            if oss.cost:
                oss_cost += sum(oss.cost.values())
            substation_costs[oss.id] = oss.cost
            
        # add up total cost
        total_cost = anch_cost + pf_cost + moor_cost + cab_cost + turb_cost + oss_cost
        # maxrows_oss = max([len(oss.cost) for oss in self.substationList.values()])
            
        # now write the costs to a spreadsheet
        
        # let's do a new sheet for each component type and one overview sheet
        # for now let's write out a dictionary
        cost_dict = {'total cost':total_cost,'cable cost':cab_cost,
                     'anchor cost':anch_cost, 'pf cost':pf_cost,
                     'moor cost':moor_cost,'turbine cost':turb_cost,
                     'substation cost':oss_cost}
        
        return(cost_dict)
    
    def unload(self,file='project.yaml'):
        '''
        Function to unload information to a yaml file
        
        Parameters
        ----------
        file : str
            File name for output yaml file

        Returns
        -------
        None.

        '''
        print(f'Unloading project to yaml file {file}')
                   
        # build out array table
        arrayKeys = ['ID','topsideID','platformID','mooringID','x_location','y_location','z_location','heading_adjust']
        arrayData = [] #np.zeros((len(arrayKeys),len(self.platformList)))
        pf_type = []
        
        anchConfigs = {}
        arrayAnch = []
        mapAnchNames = {}
        mscs = {}
        arrayMoor = []

        for i,anch in enumerate(self.anchorList.values()):  
            newanch = True
            name = anch.dd['name'] if 'name' in anch.dd else str(len(anchConfigs))
            # add mass if available
            aad = deepcopy(anch.dd['design'])
            if anch.mass is not None and anch.mass>0:
                aad['mass'] = anch.mass
            if 'type' in anch.dd:
                aad['type'] = anch.dd['type']
            if anchConfigs:
                ac = [an for an,ad in anchConfigs.items() if ad==aad]
                if len(ac)>0:
                    newanch = False
                    name = ac[0] # reset name to matching config name
            if newanch:
                anchConfigs[name] = aad
            if len(anch.attachments)>1:
                # shared anchor, add to arrayAnch list
                arrayAnch.append([anch.id,name,anch.r[0],anch.r[1],0])

            mapAnchNames[anch.id] = name
        
        # build out platform info
        topList = []   
        allconfigs = []
        # replace all 'headings' fairlead/Jtube defs with r_rel for ease of identification
        # of any changes in fairlead positions etc.
        for pt in self.platformTypes:
            if 'fairleads' in pt:
                for f,fl in enumerate(pt['fairleads']):
                    if 'headings' in fl:     
                        for head in fl['headings']:
                            # get rotation matrix of heading
                            R = rotationMatrix(0,0,np.radians(90-head))
                            # apply to unrotated r_rel
                            pt['fairleads'].append(
                                {'r_rel':np.matmul(R, fl['r_rel'])})
                        # remove fairlead listing with 'headings'
                        pt['fairleads'].pop(f)
            else:
                pt['fairleads'] = []
            if 'JTubes' in pt:
                for f,fl in enumerate(pt['JTubes']):
                    if 'headings' in fl:                      
                        for head in fl['headings']:
                            # get rotation matrix of heading
                            R = rotationMatrix(0,0,np.radians(90-head))
                            # apply to unrotated r_rel
                            pt['JTubes'].append(
                                {'r_rel':np.matmul(R, fl['r_rel'])})
                        # remove JTube listing with 'headings
                        pt['JTubes'].pop(f)
            else:
                pt['JTubes'] = []
        for i,pf in enumerate(self.platformList.values()):
            ts_loc = 0
            msys = []
            newms = True
            fairleads = [{'r_rel':att['r_rel']} 
                         for att in pf.attachments.values() 
                         if isinstance(att['obj'], Fairlead)] 
            jtubes = [{'r_rel':att['r_rel']}  
                      for att in pf.attachments.values() 
                      if isinstance(att['obj'], Jtube)]
            # Rudy: maybe consider this instead:
            # pf.dd['fairleads'] = [
            #                 {'r_rel': att['r_rel']} 
            #                 for att in pf.attachments.values() 
            #                 if isinstance(att['obj'], Fairlead)
                                  # ] 
            pf_info = {'type': pf.entity,
                       'fairleads':fairleads,
                       'JTubes':jtubes}
            if not fairleads:
                pf_info['rFair'] = pf.rFair
                pf_info['zFair'] = pf.zFair
                if not 'rFair' in self.platformTypes[pf.dd['type']]:
                    self.platformTypes[pf.dd['type']]['rFair'] = pf.rFair
                    self.platformTypes[pf.dd['type']]['zFair'] = pf.zFair
                self.platformTypes[pf.dd['type']]['fairleads'] = []
                self.platformTypes[pf.dd['type']]['JTubes'] = []
                    
            
            # update the platform type/add to platform types list if 
            # no type providd or pf_info different from any platformTypes
            if not 'type' in pf.dd:  
                update_type = True
            elif not compareDicts(
                    pf_info,
                    self.platformTypes[pf.dd['type']]
                    ):
                # if not fairleads:
                #     if not 'rFair' in self.platformTypes[pf.dd['type']]:
                #         self.platformTypes[pf.dd['type']]['rFair'] = pf.rFair
                #         self.platformTypes[pf.dd['type']]['zFair'] = pf.zFair
                #         update_type = False
                #     else:
                #         update_type = True
                # else:
                update_type = True
            else:
                update_type = False
            if update_type:
                if not self.platformTypes:
                    self.platformTypes = []
                old_type = self.platformTypes[pf.dd['type']] if 'type' in pf.dd else {}
                pf.dd['type'] = len(self.platformTypes)
                pf_info_full = old_type | pf_info
                self.platformTypes.append(pf_info_full)

            # determine any connected topsides
            for att in pf.attachments.values():
                if not isinstance(att['obj'],(Mooring, Cable, Fairlead, Jtube)):
                    dd = att['obj'].dd
                    if isinstance(att['obj'],Turbine):
                        entity = 'Turbine'
                    elif isinstance(att['obj'],Substation):
                        entity = 'Substation'
                    else:
                        entity = att['obj'].entity
                    dd['type'] = entity
                    if att['obj'].dd in topList:
                        ts_loc = int(topList.index(att['obj'].dd) + 1)
                    else:
                        topList.append(att['obj'].dd)
                        ts_loc = len(topList)

                elif isinstance(att['obj'], Mooring):
                    newcon = True
                    # check if shared
                    moor = att['obj']
                    atts = np.array(moor.attached_to)
                    is_pf = np.array([isinstance(at, Platform) for at in atts])
                    is_anch = np.array([isinstance(at, Anchor) for at in atts])
                    # get end B heading (needed for all mooring lines)
                    angB = calc_heading(moor.rA,moor.rB) # calc compass heading
                    headB = np.degrees(angB - moor.attached_to[1].phi) # remove platform heading
                    flB= []
                    fBs = [id for id,att in moor.attached_to[1].attachments.items() if isinstance(att['obj'],Fairlead)]
                    for sub in moor.subcons_B:
                        if sub.isJoined():
                            fl = [att['obj'].id for att in sub.attachments.values() if isinstance(att['obj'],Fairlead)][0]                           
                            # grab index of fairlead list from end B 
                            flB.append(fBs.index(fl)+1)
                            
                    if headB<0:
                        headB = headB +360 # make angle positive
                    # create dict describing configuration
                    config = {'span':float(moor.dd['span']),'subcomponents':moor.dd['subcomponents']}
                    
                    # check if an existing mooring configuration matches the current configuration
                    if allconfigs:
                        pc = [ii for ii,mc in enumerate(allconfigs) if mc==config]
                        for j in pc:
                            # this config already exists, don't add to config list
                            current_config = str(j)
                            newcon = False
                    if newcon:
                        # add to config list
                        allconfigs.append(config)
                        current_config = str(len(allconfigs) - 1) # make new name
                    if all(is_pf) or moor.shared:
                        # write in array_mooring section
                        # ang = np.pi/2 - np.arctan2(moor.rA[1]-atts[0].r[1],moor.rB[0]-atts[0].r[0]) # calc angle of mooring end A
                        # headA = float(np.degrees(ang - atts[0].phi)) # get heading without platform influence
                        # if headA<0:
                        #     headA = headA + 360 # make heading positive
                        if not fairleads:
                            amc = [current_config,atts[0].id, atts[1].id] # create array mooring eentry
                        else:
                            flA = []
                            fAs = [att['obj'].id for att in moor.attached_to[0].attachments.values() if isinstance(att['obj'],Fairlead)]
                            for sub in moor.subcons_A:
                                if sub.isJoined():
                                    fl = [att['obj'].id for att in sub.attachments.values() if isinstance(att['obj'],Fairlead)][0]
                                    # grab index of fairlead list from end B 
                                    flA.append(fAs.index(fl)+1)
                            amc = [current_config, atts[0].id, atts[1].id, flA, flB]
                        if not amc in arrayMoor:
                            arrayMoor.append(amc) # append entry to arrayMoor list if it's not already in there
                    elif any([len(at.attachments)>1 for at in atts[is_anch]]):
                        # we have a shared anchor here, put mooring in array_mooring
                        if fairleads:
                            # append mooring line to array_moor section
                            arrayMoor.append([current_config, 
                                              atts[0].id, 
                                              atts[1].id, 
                                              'None',
                                              flB])
                        else:
                            # append mooring line to array_moor section
                            arrayMoor.append([current_config,
                                              atts[0].id, 
                                              atts[1].id])
                    else:
                        # not shared anchor or shared mooring, add line to mooring system 
                        if fairleads:
                            msys.append([current_config,
                                         np.round(headB,2),
                                         mapAnchNames[atts[is_anch][0].id],
                                         flB])
                        else:
                            msys.append([current_config,
                                         np.round(headB,2),
                                         mapAnchNames[atts[is_anch][0].id]])

            # check if an existing mooring system matches the current        
            if len(msys)>0:
                if mscs:
                    pc = [n for n,ms in mscs.items() if sorted(ms)==sorted(msys)]
                    if len(pc)>0:
                        # this system already exists, don't add to mooring_systems dict
                        mname = pc[0]
                        newms = False
                if newms:
                    # add to mooring_system list
                    mname = 'ms'+str(len(mscs))
                    mscs[mname] = msys
            else:
                mname = 0
                
            
                    
                        

            # put all information for array data table together
            arrayData.append([pf.id, ts_loc, int(pf.dd['type']+1), mname, float(pf.r[0]), 
                              float(pf.r[1]), float(pf.r[2]), float(np.degrees(pf.phi))])
            pf_type.append(pf.dd['type'])
        
        # get set of just platform types used in this project
        pf_type = list(set(pf_type))
        pf_type.sort() # sorted list of indexes like [0,1,2,3,...]

        # figure out keys
        if len(pf_type) > 1:
            pfkey = 'platforms'
            pfTypes = [self.platformTypes[val] for x,val in enumerate(pf_type)]
        else:
            pfkey = 'platform'
            pfTypes = self.platformTypes[int(list(pf_type)[0])]
  
        # adjust indexes to remove unused pf_types
        old_types = deepcopy(pf_type)
        pf_type = np.array([pf-pf_type[0] for pf in pf_type])
        for iii,t in enumerate(pf_type[:-1]):
            if pf_type[iii+1] - t>1:
                new_diff = pf_type[iii+1]-t-1
                pf_type[iii:] -= new_diff

        for a in arrayData:
            a[2] = pf_type[old_types.index(a[2]-1)]+1
           
        # build out site info
        site = {}
        
        sps = deepcopy(self.soilProps)
        for ks,sp in sps.items():
            for k,s in sp.items():
                if not isinstance(s,list) and not 'array' in type(s).__name__:
                    sp[k] = [s]
            sps[ks] = sp
        if hasattr(self,'soilProps') and self.soilProps:
            if len(self.soil_x)>1:
                site['seabed'] = {'x':[float(x) for x in self.soil_x],'y':[float(x) for x in self.soil_y],'type_array':self.soil_names.tolist(),
                                  'soil_types': sps}# [[[float(v[0])] for v in x.values()] for x in self.soilProps.values()]}
                    
        if len(self.grid_x)>1:
            site['bathymetry'] = {'x':[float(x) for x in self.grid_x],'y':[float(y) for y in self.grid_y],'depths':[[float(y) for y in x] for x in self.grid_depth]}
        if len(self.boundary)>0:
            site['boundaries'] = {'x_y':[[float(y) for y in x] for x in self.boundary]}
            
        if self.marine_growth:
            site['marine_growth'] = {
                'keys':[key for key in self.marine_growth[0].keys()],
                'data':[list(row.values()) for row in self.marine_growth]
                }
            if self.marine_growth_buoys:
                site['marine_growth']['buoys'] = [x for x in self.marine_growth_buoys]
        site['general'] = {'water_depth':float(self.depth),'rho_air':float(self.rho_air),
                           'rho_water':float(self.rho_water),'mu_air':float(self.mu_air)}
        
         
            
        # # build out mooring and anchor sections
 
        anchKeys = ['ID','type','x','y','embedment']
        lineKeys = ['MooringConfigID','endA','endB','fairleadA','fairleadB']
        
        msyskeys = ['MooringConfigID','heading','anchorType','fairlead']
        moor_systems = {}
        for name,sys in mscs.items():
            moor_systems[name] = {'keys':msyskeys[:len(sys[0])],
                                  'data':sys}

        # set up mooring configs, connector and section types dictionaries
        connTypes = {}  
        secTypes = {}
        mooringConfigs = {}
        sIdx = 0
        sUnique = []
        for j,conf in enumerate(allconfigs):
            sections = []
            # iterate through subcomponents
            for comp in conf['subcomponents']:
                if isinstance(comp,list):
                    sections.append({'subsections':[]})
                    for subcomp in comp:
                        if isinstance(subcomp,list):
                            sections[-1]['subsections'].append([])
                            for sc in subcomp:
                                if 'L' in sc:
                                    # add section info
                                    stm = sc['type']['material'] # section type material
                                    stw = sc['type']['w']        # section type weight

                                    sKey = (stm, stw)
                                    if sKey not in sUnique:
                                        sUnique.append(sKey)
                                        sc['type']['name'] = sIdx
                                        stn = sc['type']['name'] # section type name
                                        secTypes[stn] = dict(sc['type'])
                                        sIdx += 1
                                        
                                    stn = sUnique.index(sKey)
                                    sections[-1]['subsections'][-1].append({'type':stn,'length':float(sc['L'])})
                                else:
                                    if not sc['m'] == 0 or not sc['CdA'] == 0 or not sc['v'] == 0:
                                        # this is not an empty connector
                                        if not 'type' in sc:
                                            # make a new connector type
                                            connTypes[str(int(len(connTypes)))] = dict(sc)
                                            ctn = str(int(len(connTypes)-1)) # connector type name
                                        else:
                                            ctn = str(sc['type'])
                                            connTypes[ctn] = dict(sc)
                                        sections[-1]['subsections'][-1].append({'connectorType':ctn})
                                        
                else:
                    if 'L' in comp:
                        # add section info
                        stm = comp['type']['material'] # section type material
                        stw = comp['type']['w']        # section type weight

                        sKey = (stm, stw)
                        if sKey not in sUnique:
                            sUnique.append(sKey)
                            comp['type']['name'] = sIdx
                            stn = comp['type']['name'] # section type name
                            secTypes[stn] = dict(comp['type'])
                            sIdx += 1
                            
                        stn = sUnique.index(sKey)
                        sections.append({'type':stn,'length':float(comp['L'])})
                    else:
                        # add connector if it isn't empty
                        if not comp['m'] == 0 or not comp['CdA'] == 0 or not comp['v'] == 0:
                            # this is not an empty connector
                            if not 'type' in comp:
                                # make a new connector type
                                connTypes[str(int(len(connTypes)))] = dict(comp)
                                ctn = str(int(len(connTypes)-1)) # connector type name
                            else:
                                ctn = str(comp['type'])
                                connTypes[ctn] = dict(comp)
                                    
                            sections.append({'connectorType':ctn})
                                        
            
            # put mooring config dictionary together
            mooringConfigs[str(j)] = {'name':str(j),'span':float(conf['span']),'sections':sections}

                    
        # cables setup
        cables = []
        cableTypes = {}
        cableConfigs = {}
        cUnique = []
        bUnique = []
        appendageTypes = {}
        jUnique = []
        jIdx = 0
        for jj,cab in enumerate(self.cableList.values()):
            endA = cab.attached_to[0]
            endB = cab.attached_to[1]
            angA = np.pi/2 - np.arctan2(cab.subcomponents[0].rB[1]-cab.rA[1],cab.subcomponents[0].rB[0]-cab.rA[0])
            headA = float(np.degrees(angA - endA.phi))
            angB = np.pi/2 - np.arctan2(cab.subcomponents[-1].rA[1]-cab.rB[1],cab.subcomponents[-1].rA[0]-cab.rB[0])
            headB = float(np.degrees(angB - endB.phi))
            coords = []
            statcab = 'None'
            dynCabs = [None,None]
            burial = None
            jA = None
            jB = None
            jtubesA = [att['obj'] for att in endA.attachments.values() if isinstance(att['obj'], Jtube)]
            jtubesB = [att['obj'] for att in endB.attachments.values() if isinstance(att['obj'], Jtube)]
            
            for kk,sub in enumerate(cab.subcomponents):
                currentConfig = {}
                    
                if isinstance(sub,StaticCable):
                    # pull out cable config and compare it to existing cableConfigs
                    ctw = sub.dd['cable_type']['w']
                    ctA = sub.dd['cable_type']['A'] 
                    cKey = (ctw,ctA)
                    ctf = False
                    # # check if made with getCableProps (then we can skip writing out cable type info)
                    # if 'notes' in sub.dd['cable_type']:
                    #     if 'made with getCableProps' in sub.dd['cable_type']['notes']:
                    #         ctk = 'cableFamily'
                    #         ctn = 'static_cable_'+str(int(sub.voltage))
                    #         ctf = True
                    
                    # create current cable config dictionary
                    if not ctf:
                        if not cKey in cUnique:
                            cUnique.append(cKey)                        
                            ctn = 'stat_cab_'+str(len(cUnique)-1)
                            cableTypes[ctn] = sub.dd['cable_type']
                        else:
                            cIdx = cUnique.index(cKey)
                            ctn = 'stat_cab_'+str(cIdx)
                    

                    
                    # check for routing coordinates
                    if hasattr(sub,'x') and len(sub.x)>0:
                        if hasattr(sub,'r'):
                            coords.extend([[sub.x[aa],sub.y[aa],sub.r[aa]] for aa in range(len(sub.x))])
                        else:
                            coords.extend([[sub.x[aa],sub.y[aa]] for aa in range(len(sub.x))])
                        
                    if hasattr(sub,'burial'):
                        burial = sub.burial
                        
                    statcab = ctn

                        
                elif isinstance(sub,DynamicCable):
                    jtube = [att for att in sub.attached_to if isinstance(att,Jtube)]                      
                    # grab index of fairlead list from end B 
                    
                    for jj in jtube:
                        if jj.attached_to == endA:
                            jA = jtubesA.index(jj)+1

                        elif jj.attached_to == endB:
                            jB = jtubesB.index(jj)+1
                    # pull out cable config and compare it to existing cableConfigs
                    ct = sub.dd['type'] # static or dynamic
                    ctw = sub.dd['cable_type']['w']
                    ctA = sub.dd['A'] 
                    cKey = (ctw,ctA)
                    ctf = False; ctk = 'cable_type'
                    # # check if made with getCableProps (then we can skip writing out cable type info)
                    # if 'notes' in sub.dd['cable_type']:
                    #     if 'made with getCableProps' in sub.dd['cable_type']['notes']:
                    #         ctn = ct+'_cable_'+str(int(sub.voltage))
                    #         ctf = True
                    # check if cable type has already been written
                    if not ctf:
                        if not cKey in cUnique:
                            cUnique.append(cKey)                        
                            ctn = 'dyn_cab_'+str(len(cUnique)-1)
                            cableTypes[ctn] = sub.dd['cable_type']
                        else:
                            cIdx = cUnique.index(cKey)
                            ctn = 'dyn_cab_'+str(cIdx)
                    # collect buoyancy sections (and appendages) info if applicable
                    bs = []
                    ac = 0
                    if 'appendages' in sub.dd:
                        for app in sub.dd['appendages']:
                            if not app['type'] in appendageTypes:
                                appendageTypes[app['type']] = app
                            elif appendageTypes[app['type']] != app:
                                # adjust new name
                                app['type'] = app['type']+'_'+str(ac)
                                while app['type'] in appendageTypes and appendageTypes[app['type']] != app:
                                    ac += 1
                                    app['type'] = app['type'][:-1]+str(ac)
                                appendageTypes[app['type']] = app
                                
                            bs.append({'type':app['type']})
                    if 'buoyancy_sections' in sub.dd:
                        for b in sub.dd['buoyancy_sections']:
                            btw = b['module_props']['w']; btv = b['module_props']['volume']
                            if not (btw,btv) in bUnique:
                                bUnique.append((btw,btv))
                                btn = 'buoy_'+str(len(bUnique)-1)
                                appendageTypes[btn] = b['module_props']
                                appendageTypes[btn]['type'] = 'buoy'
                            else:
                                bid = bUnique.index((btw,btv))
                                btn = 'buoy_'+str(bid)
                            bs.append({'L_mid':b['L_mid'],'N_modules':b['N_modules'],
                                      'spacing':b['spacing'],'V':b['module_props']['volume'],
                                      'type':btn})
                            
                    # grab joint info
                    if kk == 0 and len(cab.subcomponents)>1:
                        sc = cab.subcomponents[kk+1] 
                    else:
                        sc = cab.subcomponents[kk-1]
                    jsub = isinstance(sc,Joint)
                    if jsub:
                        # grab joint info and add
                        if 'm' in sc or 'v' in sc and (sc['m']!=0 or sc['v']!=0):
                            jKey = (getFromDict(sc,'m',default=0),getFromDict(sc,'v',default=0))
                            if not jKey in jUnique:
                                jUnique.append(jKey)
                                jtn = 'joint_'+str(len(jUnique)-1)
                                appendageTypes[jtn] = dict(deepcopy(sc))
                                appendageTypes[jtn]['type'] = 'joint'
                                if 'r' in sc:
                                    appendageTypes[jtn].pop('r')
                            else:
                                jtd = deepcopy(sc)
                                if 'r' in sc:
                                    jtd.pop('r')
                                jIdx = jUnique.index(jKey)
                                jtn = 'joint_'+str(jIdx)
                            bs.append({'type':jtn})
                    # create current cable config dictionary
                    currentConfig = {ctk:ctn,'A':ctA,
                                     'span':sub.dd['span'],'length':sub.L,
                                     'voltage':sub.dd['cable_type']['voltage'],'sections':bs}
                    if 'rJTube' in sub.dd:
                        currentConfig['rJTube'] = sub.dd['rJTube']
                    # check if current cable config already exists in cable configs dictionary
                    if currentConfig in cableConfigs.values():
                        ccn = [key for key,val in cableConfigs.items() if val==currentConfig][0] # get cable config key
                    else:
                        # create new cable config entry in dictionary
                        ccn = 'dynamic_'+str(len(cableConfigs))
                        cableConfigs[ccn] = currentConfig
                        
                    if kk>0:
                        didx = 1
                    else:
                        didx = 0 
                    dynCabs[didx] = ccn
                    # currentCable.append({'type':ccn})

                
            cid = 'array_cable'+str(len(cables))
            endAdict = {'attachID':endA.id,
                        'heading':headA,
                        'dynamicID':dynCabs[0] if dynCabs[0] else 'None'}
            if jA:
                endAdict['JTube'] = jA
            endBdict = {'attachID':endB.id,
                        'heading':headB,
                        'dynamicID':dynCabs[1] if dynCabs[1] else 'None'}
            if jB:
                endBdict['JTube'] = jB
            
            cables.append({'name':cid,'endA':endAdict,'endB':endBdict,'type':statcab})
            
            if coords:
                cables[-1]['routing_x_y_r'] = coords
            if burial:
                cables[-1]['burial'] = burial
                
            
            
        
        
            
         
        # create master output dictionary for yaml
        if arrayMoor:
            arrayMooring = {'anchor_keys':anchKeys, 'anchor_data':arrayAnch,
                            'line_keys':lineKeys[:len(arrayMoor[0])], 'line_data':arrayMoor}
        else:
            arrayMooring = {}
        output = {'site':site, 'array':{'keys':arrayKeys,'data':arrayData}, 
                  pfkey:pfTypes, 
                  'topsides': topList, 
                  'array_mooring':arrayMooring,
                  'mooring_systems':moor_systems,
                  'mooring_line_configs':mooringConfigs,
                  'mooring_line_types':secTypes, 
                  'mooring_connector_types':connTypes,
                  'anchor_types':anchConfigs,
                  'cables':cables,'dynamic_cable_configs':cableConfigs,'cable_types':cableTypes, 
                  'cable_appendages':appendageTypes}

        output = cleanDataTypes(output, convert_lists=True)
        import ruamel.yaml
        yaml = ruamel.yaml.YAML()
        
        # write out to file
        with open(file,'w') as f:    
            yaml.dump(output,f)
        
    def extractFarmInfo(self, cmax=5, fmax=10/6, Cmeander=1.9, force=1.95e6, direction=0.0, retainForce=False):
        '''
        Function to extract farm-level information required to create FAST.Farm case simulations. [Under developement]:

        Parameters
        ----------
        cmax : float, optional
            Maximum rotor induction factor to be used in FAST.Farm simulations
        fmax: float, optional
            Maximum rotor frequency to be used in FAST.Farm simulations (Hz)
        Cmeander: float, optional
            Meandering coefficient to be used in FAST.Farm simulations
        force: float, optional
            Magnitude of the external force applied to each platform (N)
        direction: float, optional
            Direction of the applied force in degrees (0deg = +x +ve CCW)
        retainForce: bool, optional
            A flag to retain the applied force after extracting the platform offsets. If False, the force is removed and the equilibrium is re-solved.
        Returns
        -------
        wts : dict
            General farm-level information needed for FAST.Farm from project class
        yaw_init : list 
            initial yaw offset values (for not it's set as just the platform orientation adjusted for rotational convention variation between FAM and FF)
        '''      


        # ----------- Extract Wind Farm Data
        wts = {}
        i = 0
        yaw_init = np.zeros((1, len(self.platformList.items())))     
        for _, pf in self.platformList.items():
            x, y, z   = pf.r[0], pf.r[1], pf.r[2]
            phi_deg       = np.degrees(pf.phi)  # float((90 - np.degrees(pf.phi)) % 360)  # Converting FAD's rotational convention (0deg N, +ve CW) into FF's rotational convention (0deg E, +ve CCW)
            phi_deg       = (phi_deg + 180) % 360 - 180  # Shift range to -180 to 180
            for att in pf.attachments.values():
                if isinstance(att['obj'],Turbine):
                    if hasattr(att['obj'], 'D'):
                        D = int(att['obj'].D)
                    else:
                        D = 242
                    zhub = att['obj'].dd['hHub']
            
            wts[i] = {
                'x': x, 'y': y, 'z': z, 'phi_deg': phi_deg, 'D': D, 'zhub': zhub, 
                'cmax': cmax, 'fmax': fmax, 'Cmeander': Cmeander
                }
            yaw_init[0, i] = -phi_deg
            i += 1

        # Apply force and compute initial platform offsets
        fx = force*np.cos(np.radians(direction))
        fy = force*np.sin(np.radians(direction))   
        for _, pf in self.platformList.items():
            # if pf.entity=='FOWT':        #TODO: Rudy - maybe this should be replaced with an attribute: operatingPlatform: True or False (or a percentage of curtailment) [because we do want all platforms available in project class to be transformed to FFarm and if platform is not operating, we still want its information]
            pf.body.f6Ext = np.array([fx, fy, 0, 0, 0, 0])

        
        self.ms.solveEquilibrium3(DOFtype='both')
        
        i = 0
        for _, pf in self.platformList.items():
            x, y, z = wts[i]['x'], wts[i]['y'], wts[i]['z']
            xi = pf.body.r6[0] - x
            yi = pf.body.r6[1] - y
            zi = pf.body.r6[2] - z
            wts[i]['xi'] = xi
            wts[i]['yi'] = yi
            wts[i]['zi'] = zi
            i += 1
        
        # Return to original status if requested
        if not retainForce:
            for _, pf in self.platformList.items():   
                pf.body.f6Ext = np.array([0, 0, 0, 0, 0, 0])

            
            self.ms.solveEquilibrium3(DOFtype='both')

        # store farm-level wind turbine information
        self.wts = wts

        return wts, yaw_init  
    
    def FFarmCompatibleMDOutput(self, filename, MDoptionsDict=None, **kwargs):
        '''
        Function to create FFarm-compatible MoorDyn input file (assumes project.ms is already created and if subsystem converted to lines):

        Parameters
        ----------
        filename : str
            Name of the MoorDyn output file (.dat)
        MDoptionsDict: dict, optional
            MoorDyn Options. If not given, default options are considered.
        **kwargs : optional
            dynamicStiffness : bool
            unrotateTurbines : bool
                A flag to unrotate turbine (body) objects when passing it to MoorPy unload function 
                [FFarm takes fairlead points in the local-unrotated reference frame]
            renameBody : bool
                A flag to rename `Body` objects in the output MD file into `Turbine` to be compatible with FFarm.
            removeBody : bool
                A flag to remove 'Body' objects in the Bodies list in the output MD file to be compatible with FFarm.
            outputList : dict
                Output options for MoorDyn.
            bathymetryFile : str
                Path to bathymetry file.
            flag : str
                Extra flag to append to MD entries.
        '''

        # --- Default values ---
        defaults = {
            "unrotateTurbines": True,
            "dynamicStiffness": False,
            "renameBody": True,
            "removeBody": True,
            "outputList": [],
            "bathymetryFile": None,
            "flag": "-",
            "factor": [1,1]
        }

        # Merge defaults with kwargs
        opts = {**defaults, **kwargs}  # Basically, it srything in the defaults dictionary, then overwrite any entries with the values provided in kwargs.

        # Assign variables for convenience
        unrotateTurbines = opts["unrotateTurbines"]
        renameBody       = opts["renameBody"]
        removeBody       = opts["removeBody"]
        outputList       = opts["outputList"]
        bathymetryFile   = opts["bathymetryFile"]
        dynamicStiffness = opts["dynamicStiffness"]

        flag             = opts["flag"]
        factor           = opts["factor"]

        if MDoptionsDict is None:
            MDoptionsDict = {}       
        from moorpy.helpers import ss2lines    
        
        ms = self.ms
        
        # Unrotate turbines if needed
        if unrotateTurbines:
            if self.wts:
                phi = [wt['phi_deg'] for wt in self.wts.values()]  # to unrotate the platforms when unloading MoorDyn
            else:
                raise ValueError("wts is empty. Please run project.extractFarmInfo first before extracting MoorDyn")
        else:
            phi = None
        
        # Setup nNodes of lines manually based on the segment length desired.
        from moorpy.helpers import lengthAwareSegmentation
        
        lengthAwareSegmentation(ms.lineList, factor=factor)

        ms.unload(fileName=filename, phi=phi, dynamicStiffness=dynamicStiffness, MDoptionsDict=MDoptionsDict, outputList=outputList, flag=flag, cleanLineTypeName=True)
        
        # rename Body to Turbine if needed
        if renameBody:
            # Rename Body to Turbine:
            with open(filename, 'r') as f:
                filedata = f.read()

                filedata = filedata.replace('Body', 'Turbine')
                with open(filename, 'w') as f:
                    f.write(filedata)

                f.close()       
        
        if removeBody:
            with open(filename, 'r') as f:
                lines = f.readlines()

            newLines = []
            skipCount = 0

            for i, line in enumerate(lines):
                if '---' in line and ('BODIES' in line.upper() or 'BODY LIST' in line.upper() or 'BODY PROPERTIES' in line.upper()):
                    newLines.append(line)
                    newLines.append(next(iter(lines[i+1:])))  # Append 2 lines
                    newLines.append(next(iter(lines[i+2:]))) 

                    skipCount = 2 + len(self.platformList)  # Skip the number of platforms and the already appended lines above
                    continue

                if skipCount > 0:
                    skipCount -= 1
                else:
                    newLines.append(line)

            with open(filename, 'w') as f:
                f.writelines(newLines)
        
        if bathymetryFile:
            with open(filename, 'r') as f:
                lines = f.readlines()

            newLines = []

            for i, line in enumerate(lines):
                if 'DEPTH' not in line.upper():
                    newLines.append(line)
                if '---' in line and 'OPTIONS' in line.upper():
                    newLines.append(f"{bathymetryFile}             depth\n")
                
            with open(filename, 'w') as f:
                f.writelines(newLines)
            
            
            with open(filename, 'w') as f:
                f.writelines(newLines)
                
    def resetArrayCenter(self, FOWTOnly=True):
        '''
        Function to reset array center such that the farm origin is the mid-point 
        between all FOWT platforms in y and the minimum_x turbine location in x:

        Parameters
        ----------
        FOWTOnly : bool
            find the center between only FOWT-entity platforms if True.
        '''       
        x       =  np.min([p.r[0] for p in self.platformList.values() if p.entity=='FOWT' or not FOWTOnly])
        yCenter = np.mean([p.r[1] for p in self.platformList.values() if p.entity=='FOWT' or not FOWTOnly])          
        delta   = np.array([x, yCenter])
        
        # Change boundaries
        self.boundary -= delta

        # Change bathymetry
        self.grid_x -= delta[0]
        self.grid_y -= delta[1]
        
        # Change all platform locations and associated anchors/moorings/cables
        for pf in self.platformList.values():
            r2 = pf.r[:2] - delta
            r3 = np.append(r2, 0)
            pf.setPosition(r3)

        # Change RAFTDict if available.
        if self.RAFTDict:
            x_idx = self.RAFTDict['array']['keys'].index('x_location')
            y_idx = self.RAFTDict['array']['keys'].index('y_location')            
            for i in range(len(self.platformList.values())):
                self.RAFTDict['array']['data'][i][x_idx] -= delta[0]
                self.RAFTDict['array']['data'][i][y_idx] -= delta[1]

            if 'platforms' in self.RAFTDict or 'platform' in self.RAFTDict:
                    self.getRAFT(self.RAFTDict)
        self.getMoorPyArray()
        
    def reorientArray(self, windHeading=None, degrees=False):
        '''
        Reorients the array based on a given wind heading. The array will be reoriented such that wind faces East (the zero in FFarm). 
        Useful to allign the array with the wind direction.

        Parameters
        ----------
        windHeading, float (optional)
            The heading of the wind [deg or rad] depending on
            degrees parameter. The heading is based on compass convention (North=0deg and +ve CW).
        degrees : bool (optional)
            Determines whether to use degree or radian for heading.
        '''

        from scipy.interpolate import griddata

        # Check if windHeading is given
        if windHeading is None:
            raise ValueError("windHeading is not given. Please provide a valid wind heading.")
        
        if degrees:
            windHeading = np.radians(windHeading)
        
        # reference wind heading (aligned with x-axis)
        windHeadingRef = np.radians(270)
        # Calculate the phi angle with which we will rotate the array
        phi = ((np.pi/2 - windHeading) + np.pi) % (2*np.pi)
        
        # Compute rotation matrix for faster computation
        R = np.array([[np.cos(phi), np.sin(phi)],
                  [-np.sin(phi), np.cos(phi)]])
        
        # Rotate the boundary
        self.boundary = np.dot(R, self.boundary.T).T
        
        # Rotate the bathymetry
        X, Y = np.meshgrid(self.grid_x, self.grid_y)
        coords_flat = np.stack([X.flatten(), Y.flatten()], axis=-1)
        rotated_coords_flat = np.dot(R, coords_flat.T).T
        rotated_X_flat, rotated_Y_flat = rotated_coords_flat[:, 0], rotated_coords_flat[:, 1]
        X_rot = rotated_X_flat.reshape(X.shape)
        Y_rot = rotated_Y_flat.reshape(Y.shape)

        min_X = np.min(X_rot)
        max_X = np.max(X_rot)
        min_Y = np.min(Y_rot)
        max_Y = np.max(Y_rot)

        self.grid_x = np.arange(min_X, max_X + np.min(np.diff(self.grid_x)), np.min(np.diff(self.grid_x)))
        self.grid_y = np.arange(min_Y, max_Y + np.min(np.diff(self.grid_y)), np.min(np.diff(self.grid_y)))
        X_rot, Y_rot = np.meshgrid(self.grid_x, self.grid_y)
        
        # Interpolate self.grid_depth onto the rotated grid
        depth_flat = self.grid_depth.flatten()  # Flatten the depth values
        rotated_depth = griddata(
            points=rotated_coords_flat,  # Original grid points
            values=depth_flat,   # Original depth values
            xi=(X_rot, Y_rot),   # New rotated grid points
            method='linear'      # Interpolation method (can also use 'nearest' or 'cubic')
        )        

        if np.isnan(rotated_depth).any():
            nan_mask = np.isnan(rotated_depth)
            nearest_depth = griddata(
                points=rotated_coords_flat,
                values=depth_flat,
                xi=(X_rot, Y_rot),
                method='nearest'
            )
            rotated_depth[nan_mask] = nearest_depth[nan_mask]

        self.grid_depth = rotated_depth
        
        # Rotate the platforms
        for pf in self.platformList.values():
            r2 = np.dot(R, pf.r[:2].T).T
            phi = pf.phi + windHeadingRef - windHeading
            r3 = np.append(r2, 0)
            pf.setPosition(r3, heading=phi, degrees=False)

        # Change RAFTDict if available.
        if self.RAFTDict:
            x_idx = self.RAFTDict['array']['keys'].index('x_location')
            y_idx = self.RAFTDict['array']['keys'].index('y_location')            
            p_idx = self.RAFTDict['array']['keys'].index('heading_adjust')            
            for i, pf in enumerate(self.platformList.values()):
                self.RAFTDict['array']['data'][i][x_idx] = pf.r[0]
                self.RAFTDict['array']['data'][i][y_idx] = pf.r[1]
                self.RAFTDict['array']['data'][i][p_idx] = np.degrees(pf.phi)

            if 'platforms' in self.RAFTDict or 'platform' in self.RAFTDict:
                    self.getRAFT(self.RAFTDict)        
        
        self.getMoorPyArray()
        
    def repositionArray(self,platform_locs,platform_headings,anch_resize=False,
                        return_costs=False):
        '''
        Method to reposition all platforms in the array at once with input arrays of positions and headings

        Parameters
        ----------
        platform_locs : array
            2D or 3D (x,y or x,y,z) positions for each platform in an array
            Each row represents the position of the corresponding platform in the platformList.
            PlatformList is a dictionary, but it is ordered. It therefore follows this order
        platform_headings : array
            Each entry represents the heading of the corresponding platform in the platformList

        Returns
        -------
        None.

        '''
        # reset platform, moorings, and anchor positions. If adjuster function available for moorings, adjust mooring for new depth
        for i,pf in enumerate(self.platformList.values()):
            pf.setPosition(platform_locs[i],heading=platform_headings[i],project=self)

        if anch_resize:
            # get some preliminary max loads for the anchors
            self.arrayWatchCircle()
            # resize the anchors as needed
            for anch in self.anchorList.values():
                anch.getSize(anch.dd['design'].values(), anch.dd['design'].keys(), loads=None, minfs={'Ha':1.6,'Va':2},
                            LD_con=[4,8], fix_zlug=False, FSdiff_max={'Ha':.05,'Va':.05}, plot=False)

        if return_costs:
            total_cost = self.getArrayCost()
            return(total_cost['anchor cost'], total_cost['moor cost'])

    def updateFailureProbability(self):
        '''
        Function to populate (or update) failure probability dictionaries in each object 
        based on failure probability calculations developed by Emma Slack
        
        To be filled in...

        Returns
        -------
        None.

        '''
        pass    

    def mapRAFTResults(self, results=None, SFs=True):
        '''
        Function to map RAFT results to the project class. This
        maps the results from RAFT to the project class.

        Parameters
        ----------
        results : dict, optional
            Dictionary of results from RAFT. The default is None. 
            If `analyzeCases` is ran in raft, then the self.array
            results will be used. if results dictionary is empty, 
            we try to call analyze cases
        

        Returns
        -------
        None.

        '''
        if results:  # if results are given, overwrite self.array.results
            self.array.results = results
        
        if not self.array.results:
            if not self.array.design['cases']:
                raise ValueError("RAFT cases dictionary is empty. Please populate raft cases to generate RAFT results.")
            
            # If there is cases dictionary but has not been run, make sure to run RAFT to get it
            self.array.analyzeCases(display=1)        
        
        else:
            # Get results from RAFT
            # Map platform-related results:
            nCases = len(self.array.design['cases']['data'])
            for iCase in range(nCases):
                for i, pf in enumerate(self.platformList.values()):
                    pf.raftResults[iCase] = self.array.results['case_metrics'][iCase][i]

            # Map mooring-related results:
            if SFs:
                for moor in self.mooringList.values():
                    moor.safety_factors['tension'] = 1e10
            
            for iCase in range(nCases):
                for i, moor in enumerate(self.mooringList.values()):
                    # Find the corresponding line in the proj.ms.lineList
                    idx = moor.ss.number - 1
                    
                    moor.raftResults[iCase] = {
                        'Tmoor_avg': self.array.results['case_metrics'][iCase]['array_mooring']['Tmoor_avg'][[idx, idx+len(self.ms.lineList)]],
                        'Tmoor_std': self.array.results['case_metrics'][iCase]['array_mooring']['Tmoor_std'][[idx, idx+len(self.ms.lineList)]],
                        'Tmoor_min': self.array.results['case_metrics'][iCase]['array_mooring']['Tmoor_min'][[idx, idx+len(self.ms.lineList)]],
                        'Tmoor_max': self.array.results['case_metrics'][iCase]['array_mooring']['Tmoor_max'][[idx, idx+len(self.ms.lineList)]],
                        'Tmoor_PSD': self.array.results['case_metrics'][iCase]['array_mooring']['Tmoor_PSD'][[idx, idx+len(self.ms.lineList)], :]
                    }
                    if SFs:
                        SF = np.zeros((len(moor.ss.lineList)))
                        for j, line in enumerate(moor.ss.lineList):
                            line_MBL = line.type['MBL']
                            SF[j] = line_MBL/np.mean(moor.raftResults[iCase]['Tmoor_avg'])
                        
                        if min(SF) < moor.safety_factors['tension']:
                            moor.safety_factors['tension'] = min([moor.safety_factors['tension'], min(SF)])
                            moor.safety_factors['analysisType'] = f'(RAFT) MoorMod={self.array.moorMod}'
                        
            
    def generateSheets(self, filename):
        """
        Generates sheets in an Excel workbook with RAFT cases, platform, and mooring line information.

        Parameters
        ----------
        filename (str): The name of the Excel file to save the generated sheets.


        Returns
        -------
        None

        """
        
        def style_it(sheet, row, col_start, col_end, fill_color="FFFF00"):
            """
            Applies styling to a range of cells in the given sheet.

            Parameters
            ----------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to apply styling to.
            row : int
                The row number to apply styling to.
            col_start : int
                The starting column number for the range.
            col_end : int
                The ending column number for the range.

            Returns
            -------
            None
            """
            for col in range(col_start, col_end + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)

        import openpyxl
        from openpyxl.drawing.image import Image

        if not self.array.results:
            if not self.array.design['cases']:
                raise ValueError("RAFT cases dictionary is empty. Please populate raft cases to generate RAFT results.")

            self.array.analyzeCases(display=1)
        else:
            nCases = len(self.array.design['cases']['data'])

        # Create a new workbook
        workbook = openpyxl.Workbook()
        # Delete the default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        if self.array.design:
            # Create a sheet for cases information 
            cases_sheet = workbook.create_sheet(title="Cases")
            cases_sheet.append(self.array.design['cases']['keys'])
            for iCase in range(nCases):
                cases_sheet.append(self.array.design['cases']['data'][iCase])
            
        # Create a sheet for platforms
        platform_sheet = workbook.create_sheet(title="Platforms")
        platform_sheet.append(["ID", "X", "Y", "Depth", "Case", "Results (Avg)", " ", " ", " ", " ", " ", " ", " ", "Results (Std)"])
        platform_sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=13)
        platform_sheet.merge_cells(start_row=1, start_column=14, end_row=1, end_column=21)
        platform_sheet.append(["  ", " ", " ", "     ", "    ", 
                               "Surge (m)", "Sway (m)", "Heave (m)", "Roll (deg)", "Pitch (deg)", "Yaw (deg)", "NacAcc (m/s^2)", "TwrBend (kNm)",   #, "RtrSpd (RPM)", "RtrTrq (Nm)", "Power (MW)"
                               "Surge (m)", "Sway (m)", "Heave (m)", "Roll (deg)", "Pitch (deg)", "Yaw (deg)", "NacAcc (m/s^2)", "TwrBend (kNm)"])  #, "RtrSpd (RPM)", "RtrTrq (Nm)", "Power (MW)"
        platform_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        platform_sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        platform_sheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        platform_sheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
        platform_sheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)
        excMax = float('-inf')
        avg_surgeMax = float('-inf')
        avg_swayMax = float('-inf')
        avg_rollMax = float('-inf')
        avg_pitchMax = float('-inf')
        avg_nacAccMax = float('-inf')
        avg_twrBendMax = float('-inf')

        std_surgeMax = float('-inf')
        std_swayMax = float('-inf')
        std_rollMax = float('-inf')
        std_pitchMax = float('-inf')
        std_nacAccMax = float('-inf')
        std_twrBendMax = float('-inf')

        for pf in self.platformList.values():
            depth_at_pf = self.getDepthAtLocation(pf.r[0], pf.r[1])
            if hasattr(pf, 'raftResults'):
                for iCase in range(nCases):
                    if iCase==0:
                        platform_sheet.append([pf.id, round(pf.r[0], 3), round(pf.r[1], 3), round(depth_at_pf, 3), iCase,
                                               round(pf.raftResults[iCase]['surge_avg'], 3), round(pf.raftResults[iCase]['sway_avg'], 3), round(pf.raftResults[iCase]['heave_avg'], 3), 
                                               round(pf.raftResults[iCase]['roll_avg'], 3), round(pf.raftResults[iCase]['pitch_avg'], 3), round(pf.raftResults[iCase]['yaw_avg'], 3), 
                                               round(pf.raftResults[iCase]['AxRNA_avg'][0], 3), round(pf.raftResults[iCase]['Mbase_avg'][0]/1e3, 3),  #, round(pf.raftResults[iCase]['omega_avg'][0], 3), round(pf.raftResults[iCase]['torque_avg'][0], 3), round(pf.raftResults[iCase]['power_avg'][0]*1e-6, 3)])
                                               round(pf.raftResults[iCase]['surge_std'], 3), round(pf.raftResults[iCase]['sway_std'], 3), round(pf.raftResults[iCase]['heave_std'], 3), 
                                               round(pf.raftResults[iCase]['roll_std'], 3), round(pf.raftResults[iCase]['pitch_std'], 3), round(pf.raftResults[iCase]['yaw_std'], 3), 
                                               round(pf.raftResults[iCase]['AxRNA_std'][0], 3), round(pf.raftResults[iCase]['Mbase_std'][0]/1e3, 3)])  #, round(pf.raftResults[iCase]['omega_avg'][0], 3), round(pf.raftResults[iCase]['torque_avg'][0], 3), round(pf.raftResults[iCase]['power_avg'][0]*1e-6, 3)])
                    else:
                        platform_sheet.append([" ", " ", " ", " ", iCase,
                                               round(pf.raftResults[iCase]['surge_avg'], 3), round(pf.raftResults[iCase]['sway_avg'], 3), round(pf.raftResults[iCase]['heave_avg'], 3), 
                                               round(pf.raftResults[iCase]['roll_avg'], 3), round(pf.raftResults[iCase]['pitch_avg'], 3), round(pf.raftResults[iCase]['yaw_avg'], 3), 
                                               round(pf.raftResults[iCase]['AxRNA_avg'][0], 3), round(pf.raftResults[iCase]['Mbase_avg'][0]/1e3, 3),  #, round(pf.raftResults[iCase]['omega_avg'][0], 3), round(pf.raftResults[iCase]['torque_avg'][0], 3), round(pf.raftResults[iCase]['power_avg'][0]*1e-6, 3)])
                                               round(pf.raftResults[iCase]['surge_std'], 3), round(pf.raftResults[iCase]['sway_std'], 3), round(pf.raftResults[iCase]['heave_std'], 3), 
                                               round(pf.raftResults[iCase]['roll_std'], 3), round(pf.raftResults[iCase]['pitch_std'], 3), round(pf.raftResults[iCase]['yaw_std'], 3), 
                                               round(pf.raftResults[iCase]['AxRNA_std'][0], 3), round(pf.raftResults[iCase]['Mbase_std'][0]/1e3, 3)])  #, round(pf.raftResults[iCase]['omega_avg'][0], 3), round(pf.raftResults[iCase]['torque_avg'][0], 3), round(pf.raftResults[iCase]['power_avg'][0]*1e-6, 3)])
                    # Update min and max values
                    excMax = max(excMax, np.sqrt(pf.raftResults[iCase]['surge_avg']**2+pf.raftResults[iCase]['sway_avg']**2))
                    avg_surgeMax = max(avg_surgeMax, abs(pf.raftResults[iCase]['surge_avg']));      avg_swayMax = max(avg_swayMax, abs(pf.raftResults[iCase]['sway_avg']))
                    avg_rollMax = max(avg_rollMax, abs(pf.raftResults[iCase]['roll_avg']));         avg_pitchMax = max(avg_pitchMax, abs(pf.raftResults[iCase]['pitch_avg']))
                    avg_nacAccMax = max(avg_nacAccMax, abs(pf.raftResults[iCase]['AxRNA_avg'][0])); avg_twrBendMax = max(avg_twrBendMax, abs(pf.raftResults[iCase]['Mbase_avg'][0]/1e3))

                    std_surgeMax = max(std_surgeMax, abs(pf.raftResults[iCase]['surge_std']));      std_swayMax = max(std_swayMax, abs(pf.raftResults[iCase]['sway_std']))
                    std_rollMax = max(std_rollMax, abs(pf.raftResults[iCase]['roll_std']));         std_pitchMax = max(std_pitchMax, abs(pf.raftResults[iCase]['pitch_std']))
                    std_nacAccMax = max(std_nacAccMax, abs(pf.raftResults[iCase]['AxRNA_std'][0])); std_twrBendMax = max(std_twrBendMax, abs(pf.raftResults[iCase]['Mbase_std'][0]/1e3))
                    
                platform_sheet.merge_cells(start_row=platform_sheet.max_row-nCases+1, start_column=1, end_row=platform_sheet.max_row, end_column=1)
                platform_sheet.merge_cells(start_row=platform_sheet.max_row-nCases+1, start_column=2, end_row=platform_sheet.max_row, end_column=2)
                platform_sheet.merge_cells(start_row=platform_sheet.max_row-nCases+1, start_column=3, end_row=platform_sheet.max_row, end_column=3)
                platform_sheet.merge_cells(start_row=platform_sheet.max_row-nCases+1, start_column=4, end_row=platform_sheet.max_row, end_column=4)
            else:
                platform_sheet.append([pf.id, round(pf.r[0], 3), round(pf.r[1], 3), round(depth_at_pf, 3)])

        if hasattr(pf, 'raftResults'):
            platform_sheet.append(["----------------------"])
            platform_sheet.append(["Highest average values"])
            for cell in platform_sheet[platform_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True)

            platform_sheet.append(["Surge (m)", "Sway (m)", "Roll (deg)", "Pitch (deg)", "NacAcc (m/s^2)", "TwrBend (Nm)",  "Excursion (m)"])
            platform_sheet.append([round(avg_surgeMax, 3), round(avg_swayMax, 3), round(avg_rollMax, 3), round(avg_pitchMax, 3), round(avg_nacAccMax, 3), round(avg_twrBendMax, 3), round(excMax, 3)])
            # style maximum values (bold and italic)
            for cell in platform_sheet[platform_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True, italic=True)

            platform_sheet.append(["----------------------"])
            platform_sheet.append(["Highest std values"])
            for cell in platform_sheet[platform_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True)

            platform_sheet.append(["Surge (m)", "Sway (m)", "Roll (deg)", "Pitch (deg)", "NacAcc (m/s^2)", "TwrBend (Nm)"])
            platform_sheet.append([round(std_surgeMax, 3), round(std_swayMax, 3), round(std_rollMax, 3), round(std_pitchMax, 3), round(std_nacAccMax, 3), round(std_twrBendMax, 3)])
            # style maximum values (bold and italic)
            for cell in platform_sheet[platform_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True, italic=True)

        # Create a sheet for mooring lines
        mooring_sheet = workbook.create_sheet(title="Mooring Lines")
        mooring_sheet.append(["ID", "endA", "endB", "Shrd", "chain dnom [mm]", "rope dnom [mm]", "Safety Factors", "Fid Level", "Case", "Avg EndA Tension (kN)", "Std EndA Tension (kN)", "Avg EndB Tension (kN)", "Std EndB Tension (kN)"])
        
        avg_Tmoor_endA = float('-inf')
        avg_Tmoor_endB = float('-inf')
        std_Tmoor_endA = float('-inf')
        std_Tmoor_endB = float('-inf')

        for moor in self.mooringList.values():
            # Find nominal diameters in moor
            ch_dnom = '-'
            rp_dnom = '-'

            for line in moor.ss.lineList:
                if 'rope' in line.type['material'] or 'polyester' in line.type['material']:  # change the polyester/rope length
                    rp_dnom = line.type['d_nom'] * 1e3
                elif 'chain' in line.type['material']:
                    ch_dnom = line.type['d_nom'] * 1e3

            if hasattr(moor, 'raftResults'):
                for iCase in range(nCases):
                    if iCase==0:
                        mooring_sheet.append([moor.id, moor.attached_to[0].id, moor.attached_to[1].id, moor.shared, ch_dnom, rp_dnom, round(moor.safety_factors['tension'], 3), moor.safety_factors['analysisType'], iCase,
                                            round(moor.raftResults[iCase]['Tmoor_avg'][0], 3)/1e3, round(moor.raftResults[iCase]['Tmoor_std'][0], 3)/1e3,
                                            round(moor.raftResults[iCase]['Tmoor_avg'][1], 3)/1e3, round(moor.raftResults[iCase]['Tmoor_std'][1], 3)/1e3])
                        if moor.safety_factors['tension']<2.0:
                            style_it(mooring_sheet, mooring_sheet.max_row, 1, 8, fill_color="FF0000")
                        
                    else:
                        mooring_sheet.append([" ", " ", " ", " ", " ", " ", " ", " ", iCase,
                                            round(moor.raftResults[iCase]['Tmoor_avg'][0], 3)/1e3, round(moor.raftResults[iCase]['Tmoor_std'][0], 3)/1e3,
                                            round(moor.raftResults[iCase]['Tmoor_avg'][1], 3)/1e3, round(moor.raftResults[iCase]['Tmoor_std'][1], 3)/1e3])
                    if np.any(moor.raftResults[iCase]['Tmoor_avg']/1e3 < 100):
                        style_it(mooring_sheet, mooring_sheet.max_row, 9, mooring_sheet.max_column, fill_color="FFFF00")

                    # Update min and max values
                    avg_Tmoor_endA = max(avg_Tmoor_endA, moor.raftResults[iCase]['Tmoor_avg'][0]);      avg_Tmoor_endB = max(avg_Tmoor_endB, moor.raftResults[iCase]['Tmoor_avg'][1])
                    std_Tmoor_endA = max(std_Tmoor_endA, moor.raftResults[iCase]['Tmoor_std'][0]);      std_Tmoor_endB = max(std_Tmoor_endB, moor.raftResults[iCase]['Tmoor_std'][1])

                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=1, end_row=mooring_sheet.max_row, end_column=1)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=2, end_row=mooring_sheet.max_row, end_column=2)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=3, end_row=mooring_sheet.max_row, end_column=3)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=4, end_row=mooring_sheet.max_row, end_column=4)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=5, end_row=mooring_sheet.max_row, end_column=5)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=6, end_row=mooring_sheet.max_row, end_column=6)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=7, end_row=mooring_sheet.max_row, end_column=7)
                mooring_sheet.merge_cells(start_row=mooring_sheet.max_row-nCases+1, start_column=8, end_row=mooring_sheet.max_row, end_column=8)
            else:
                mooring_sheet.append([moor.id, moor.attached_to[0].id, moor.attached_to[1].id, moor.shared, ch_dnom, rp_dnom, round(moor.safety_factors['tension'], 3), moor.safety_factors['analysisType']])
        
        if hasattr(moor, 'raftResults'):
            mooring_sheet.append(["----------------------"])
            mooring_sheet.append(["Highest average values"])
            for cell in mooring_sheet[mooring_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True)

            mooring_sheet.append(["Tmoor_endA (kN)", "Tmoor_endB (kN)"])
            mooring_sheet.append([round(avg_Tmoor_endA, 3)/1e3, round(avg_Tmoor_endB, 3)/1e3])
            # style maximum values (bold and italic)
            for cell in mooring_sheet[mooring_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True, italic=True)        

            mooring_sheet.append(["----------------------"])
            mooring_sheet.append(["Highest std values"])
            for cell in mooring_sheet[mooring_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True)

            mooring_sheet.append(["Tmoor_endA (kN)", "Tmoor_endB (kN)"])
            mooring_sheet.append([round(std_Tmoor_endA, 3)/1e3, round(std_Tmoor_endB, 3)/1e3])
            # style maximum values (bold and italic)
            for cell in mooring_sheet[mooring_sheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True, italic=True)        

        # Create a sheet for a 2D Plot
        plot_sheet_2D = workbook.create_sheet(title="2D Plot")
        fig, ax = plt.subplots()
        ax = self.plot2d(ax=ax, plot_boundary=False, plot_bathymetry=True)
        filenameFolder = os.path.dirname(filename)
        imagePath = os.path.join(filenameFolder, 'temp_plot_2D.png')
        fig.savefig(imagePath, dpi=300)
        img = Image(imagePath)
        plot_sheet_2D.add_image(img, 'A1')

        # Create a sheet for a 3D Plot
        plot_sheet_3D = workbook.create_sheet(title="3D Plot")
        fig, ax = plt.subplots(subplot_kw={'projection': '3d'})
        ax = self.plot3d(ax=ax, fowt=True, draw_boundary=True, boundary_on_bath=True, draw_bathymetry=True)
        imagePath = os.path.join(filenameFolder, 'temp_plot_3D.png')
        fig.savefig(imagePath, dpi=300)
        img = Image(imagePath)
        plot_sheet_3D.add_image(img, 'A1')

        # Save the workbook
        workbook.save(filename)

    def _build_mooring_anchor_instance_shared(self, arrayMooring, arrayInfo, arrayAnchor, lineConfigs, connectorTypes, alph):
        from famodel.mooring.mooring import Mooring  # adjust import if needed
        from famodel.anchors.anchor import Anchor

        mooringList = {}
        anchorList = {}

        if arrayMooring == {}:
            return mooringList, anchorList
        
        # get mooring line info for all lines 
        for j in range(0, len(arrayMooring)): # run through each line            
            
            array_name = [] # platforms connected to the mooring line
            
            # Error check for putting an anchor (or something else) at end B
            if not any(ids['ID'] == arrayMooring[j]['endB'] for ids in arrayInfo):
                raise Exception("Input for end B must match an ID from the array table.")
            if any(ids['ID'] == arrayMooring[j]['endB'] for ids in arrayAnchor):
                raise Exception(f"input for end B of line_data table row '{j}' in array_mooring must be an ID for a FOWT from the array table. Any anchors should be listed as end A.")
            
            # Make sure no anchor IDs in arrayAnchor table are the same as IDs in array table
            for k in range(0,len(arrayInfo)):
                if any(ids['ID'] == arrayInfo[k] for ids in arrayAnchor):
                    raise Exception(f"ID for array table row {k} must be different from any ID in anchor_data table in array_mooring section")
            
            # determine if end A is an anchor or a platform
            if any(ids['ID'] == arrayMooring[j]['endA'] for ids in arrayInfo): # shared mooring line (no anchor)
                # get ID of platforms connected to line
                array_name.append(self.platformList[arrayMooring[j]['endB']])
                array_name.append(self.platformList[arrayMooring[j]['endA']])
                # find row in array table associated with these platform IDs and set locations
                for k in range(0, len(arrayInfo)):
                    if arrayInfo[k]['ID'] == array_name[0]:
                        rowB = arrayInfo[k]
                    elif arrayInfo[k]['ID'] == array_name[1]:
                        rowA = arrayInfo[k]

                # get configuration for the line 
                mooring_name = arrayMooring[j]['MooringConfigID']       
                
                # create mooring and connector dictionary for that line
                mdd = getMoorings(mooring_name, lineConfigs, 
                                    connectorTypes, array_name,
                                    lineProps=self.lineProps, 
                                    lineTypes=self.lineTypes, 
                                    rho=self.rho_water, g=self.g)
                
                # Add entries to mooring dictionary that depend on project class and platform info
                mdd['zAnchor']  = -self.depth
                mdd['rad_fair'] = self.platformList[array_name[0].id].rFair
                mdd['z_fair']   = self.platformList[array_name[0].id].zFair
                
                # create mooring class instance
                moor = Mooring.addMooring(id=str(array_name[1].id)+'-'+str(array_name[0].id), 
                                        dd=mdd,
                                        shared=1)
                
                # attach ends
                fairsB = attachFairleads(moor,
                                1,
                                array_name[0],
                                fair_ID_start=array_name[0].id+'_F',
                                fair_inds=arrayMooring[j]['fairleadB'])
                fairsA = attachFairleads(moor,
                                0,
                                array_name[1],
                                fair_ID_start=array_name[1].id+'_F',
                                fair_inds=arrayMooring[j]['fairleadA'])

                
                # determine heading
                points = [[f.r[:2] for f in fairsA], 
                            [f.r[:2] for f in fairsB]]
                headingB = calc_heading(points[0], points[1])

                # TODO: in the future, reposition should not depend on project class
                moor.reposition(r_center=[array_name[1].r,
                                            array_name[0].r],
                                heading=headingB, project=self)
                                

            elif any(ids['ID'] == arrayMooring[j]['endA'] for ids in arrayAnchor): # end A is an anchor
                # get ID of platform connected to line
                array_name.append(self.platformList[arrayMooring[j]['endB']])
                
                # get configuration for that line 
                mooring_name = arrayMooring[j]['MooringConfigID']                       
                # create mooring and connector dictionary for that line
                mdd = getMoorings(mooring_name, lineConfigs, 
                                    connectorTypes, array_name,
                                    lineProps=self.lineProps, 
                                    lineTypes=self.lineTypes, 
                                    rho=self.rho_water, g=self.g)
                
                # Add entries to mooring dictionary that depend on project class and platform info
                mdd['zAnchor']  = -self.depth
                mdd['rad_fair'] = self.platformList[array_name[0].id].rFair
                mdd['z_fair']   = self.platformList[array_name[0].id].zFair

                # get letter number for mooring line
                ind = len(array_name[0].getMoorings())
                
                # create mooring class instance
                moor = Mooring.addMooring(id=str(array_name[0].id)+alph[ind], 
                                        dd=mdd)
                
                # check if anchor instance already exists
                if any(tt == arrayMooring[j]['endA'] for tt in anchorList): # anchor name exists already in list
                    # find anchor class instance
                    for anch in anchorList:
                        if anch == arrayMooring[j]['endA']:
                            anchor = anchorList[anch]
                            anchor.shared=True
                else:
                    # find location of anchor in arrayAnchor table
                    for k in range(0,len(arrayAnchor)):
                        if arrayAnchor[k]['ID'] == arrayMooring[j]['endA']:
                            aloc = [arrayAnchor[k]['x'],arrayAnchor[k]['y']] 
                            aNum = k # get anchor row number
                            # set line anchor type and get dictionary of anchor information
                            lineAnch = arrayAnchor[k]['type']
                    ad, mass = getAnchors(lineAnch, self.anchorTypes) # call method to create dictionary
                    # create anchor object
                    anchor = Anchor.addAnchor(id=arrayAnchor[aNum]['ID'], dd=ad, mass=mass)
                    anchor.r[:2] = [aloc[0],aloc[1]]
                # attach anchor
                moor.attachTo(anchor,end='A')
                # attach platform
                fairsB = attachFairleads(moor,
                                            1,
                                            array_name[0],
                                            fair_ID_start=array_name[0].id+'_F',
                                            fair_inds=arrayMooring[j]['fairleadB'])

                # determine heading
                headingB = calc_heading(anchor.r[:2],[f.r[:2] for f in fairsB])
                
                # re-determine span as needed from anchor loc and end B midpoint
                # this is to ensure the anchor location does not change from that specified in the ontology
                moor.span = np.linalg.norm(anchor.r[:2]-
                                            np.array(calc_midpoint([f.r[:2] for f in fairsB])))

                # reposition mooring
                moor.reposition(r_center=array_name[0].r, heading=headingB, project=self)
                # update depths
                # NOTE: I am pretty sure we can move this method to anchor class 
                zAnew, nAngle = self.getDepthAtLocation(aloc[0],aloc[1], return_n=True)
                moor.dd['zAnchor'] = -zAnew
                moor.z_anch = -zAnew
                moor.setEndPosition([aloc[0],aloc[1],-zAnew], 0)
                
                # Position the subcomponents along the Mooring
                moor.positionSubcomponents()


            else: # error in input
                raise Exception(f"end A input in array_mooring line_data table line '{j}' must be either an ID from the anchor_data table (to specify an anchor) or an ID from the array table (to specify a FOWT).")
                                        
            # add heading to platform headings list
            moor.rel_heading = np.degrees(headingB-array_name[0].phi)#np.radians(arrayMooring[j]['headingB']))
            array_name[0].setPosition(r=array_name[0].r, project=self)
            if len(array_name)>1: # if shared line
                # headingA = headingB - np.pi
                # array_name[1].mooring_headings.append(headingA-array_name[1].phi) # add heading
                array_name[1].setPosition(r=array_name[1].r, project=self)

        return mooringList, anchorList
                
'''
Other future items:
Cost calc functions
System Reliability/failure analysis functions
Full scenario visualization functions
Load case setup and constraint eval?
'''

if __name__ == '__main__':
    '''
    project = Project()
    project.loadSoil(filename='tests/soil_sample.txt')
    # create project class instance from yaml file
    #project = Project(file='OntologySample600m.yaml')
    project = Project(file='../tests/simple_farm.yaml')
    
    project.getMoorPyArray(cables=1,plt=1,pristineLines=1)

    # make envelopes and watch circles
    # (Mooring.getEnvelope will call Platform.getWatchCircle when needed)
    for moor in project.mooringList.values():
        moor.getEnvelope()

    project.plot2d(plot_boundary=False)  # this should also plot the watch circles/envelopes!
    
    '''
    

    # point to location of yaml file with uniform array info
    filename = '../Examples/OntologySample600m_shared.yaml' # yaml file for project

    # load in yaml
    project = Project(file=filename,raft=False)


    project.getMoorPyArray()

    # plot in 2d and 3d
    #project.plot2d()
    #project.plot3d(fowt=True)

    #plt.show()
    
    
    # ----
    
    plt.show()