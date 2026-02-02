"""Core code for setting up a IO&M scenario"""

import os
import numpy as np
import matplotlib.pyplot as plt

import moorpy as mp
from moorpy.helpers import set_axes_equal
from moorpy import helpers
import yaml
from copy import deepcopy
import string
try: 
    import raft as RAFT
except:
    pass

#from shapely.geometry import Point, Polygon, LineString
from famodel.mooring.mooring import Mooring
from famodel.platform.platform import Platform
from famodel.anchors.anchor import Anchor
from famodel.mooring.connector import Connector
from famodel.substation.substation import Substation
from famodel.cables.cable import Cable
from famodel.cables.dynamic_cable import DynamicCable
from famodel.cables.static_cable import StaticCable
from famodel.cables.cable_properties import getCableProps, getBuoyProps, loadCableProps,loadBuoyProps
from famodel.cables.components import Joint
from famodel.turbine.turbine import Turbine
from famodel.famodel_base import Node

# Import select required helper functions
from famodel.helpers import (check_headings, head_adjust, getCableDD, getDynamicCables, 
                            getMoorings, getAnchors, getFromDict, cleanDataTypes, 
                            getStaticCables, getCableDesign, m2nm, loadYAML, 
                            configureAdjuster, route_around_anchors)

import networkx as nx
from action import Action, increment_name
from task import Task

from assets import Vessel, Port
from scheduler import Scheduler



def loadYAMLtoDict(info, already_dict=False):
    '''Reads a list or YAML file and prepares a dictionary'''
    
    if isinstance(info, str):
    
        with open(info) as file:
            data = yaml.load(file, Loader=yaml.FullLoader)
            if not data:
                raise Exception(f'File {info} does not exist or cannot be read. Please check filename.')
    elif isinstance(info, list):
        data = info
    else:
        raise Exception('loadYAMLtoDict must be passed a filename or list')
    
    # Go through contents and product the dictionary
    info_dict = {}
    
    if already_dict:
        # assuming it's already a dict 
        info_dict.update(data)
        
    else: # a list of dicts with name parameters
        # So we will convert into a dict based on those names
        for entry in data:
            if not 'name' in entry:
                print(entry)
                raise Exception('This entry does not have a required name field.')
                
            if entry['name'] in info_dict:
                print(entry)
                raise Exception('This entry has the same name as an existing entry.')
                
            info_dict[entry['name']] = entry  # could make this a copy operation if worried
    
    return info_dict
    

def printStruct(t, s=0):
    '''Prints a nested list/dictionary data structure with nice indenting.'''
    
    if not isinstance(t,dict) and not isinstance(t,list):
        print(" "*s+str(t))
    else:
        for key in t:
            if isinstance(t,dict) and not isinstance(t[key],dict) and not isinstance(t[key],list):
                print(" "*s+str(key)+"  :  "+str(t[key]))
            else:
                print(" "*s+str(key))
                if not isinstance(t,list):
                    printStruct(t[key], s=s+2)

#def storeState(project,...):


#def applyState():


def unifyUnits(d, display=0):
    '''Converts any capability specification/metric in supported non-SI units
    to be in SI units. Converts the key names as well.'''
    
    # load conversion data from YAML (eventually may want to store this in a class)
    with open('spec_conversions.yaml') as file:
        data = yaml.load(file, Loader=yaml.FullLoader)
    
    keys1 = []
    facts = []  # conversion factors
    keys2 = []
    
    for line in data:
        keys1.append(line[0])
        facts.append(float(line[1]))
        keys2.append(line[2])
    
    for name, asset in d.items():  # loop through each asset's dict
        
        capabilities = {}  # new dict of capabilities to built up
        
        for cap_key, cap_val in asset['capabilities'].items():
            
            # make the capability type sub-dictionary
            capabilities[cap_key] = {}
        
            for key, val in cap_val.items():  # look at each capability metric
                try:
                    i = keys1.index(key)  # find if key is on the list to convert
                    
                    
                    if keys2[i] in cap_val.keys():
                        raise Exception(f"Specification '{keys2[i]}' already exists")
                    
                    if display > 1: print(f"Converting from {key} to {keys2[i]}")
                    
                    capabilities[cap_key][keys2[i]] = val * facts[i]  # make converted entry
                    #capability[keys2[i]] = val * facts[i]  # create a new SI entry
                    #del capability[keys1[i]]  # remove the original?
                    
                except:
                    
                    capabilities[cap_key][key] = val  # copy over original form
                    
        # Copy over the standardized capability dict for this asset
        asset['capabilities'] = capabilities


class Scenario():

    def __init__(
        self,
        actions_yaml='actions.yaml',
        requirements_yaml='requirements.yaml',
        capabilities_yaml='capabilities.yaml',
        vessels_yaml='vessels.yaml',
        objects_yaml='objects.yaml',
        display=0
    ):
        '''Initialize a scenario object that can be used for IO&M modeling of
        of an offshore energy system. Eventually it will accept user-specified 
        settings files.
        '''
        
        # ----- Load database of supported things -----
        
        actionTypes = loadYAMLtoDict(actions_yaml, already_dict=True)  # Descriptions of actions that can be done
        requirements = loadYAMLtoDict(requirements_yaml, already_dict=True)  # Descriptions of requirements that can be done
        capabilities = loadYAMLtoDict(capabilities_yaml, already_dict=True)
        vessels = loadYAMLtoDict(vessels_yaml, already_dict=True)
        objects = loadYAMLtoDict(objects_yaml, already_dict=True)
        self.display = display
        
        unifyUnits(vessels, display=display)  # (function doesn't work yet!) <<<
        
        # ----- Validate internal cross references -----
        
        # Make sure vessels don't use nonexistent capabilities or actions
        for key, ves in vessels.items():
            
            #if key != ves['name']:
            #    raise Exception(f"Vessel key ({key}) contradicts its name ({ves['name']})")
            
            # Check capabilities
            if not 'capabilities' in ves:
                raise Exception(f"Vessel '{key}' is missing a capabilities list.")
                
            for capname, cap in ves['capabilities'].items():
                if not capname in capabilities and display > 0:
                    raise Exception(f"Vessel '{key}' capability '{capname}' is not in the global capability list.")
                
                # Could also check the sub-parameters of the capability
                for cap_param in cap:
                    if not cap_param in capabilities[capname] and display > 0:
                        #raise Exception(f"Vessel '{key}' capability '{capname}' parameter '{cap_param}' is not in the global capability's parameter list.")
                        print(f"Warning: Vessel '{key}' capability '{capname}' parameter '{cap_param}' is not in the global capability's parameter list.")
            
            # Check actions
            if not 'actions' in ves:
                raise Exception(f"Vessel '{key}' is missing an actions list.")
            
            for act in ves['actions']:
                if not act in actionTypes:
                    raise Exception(f"Vessel '{key}' action '{act}' is not in the global action list.")
        
        
        # Make sure actions refer to supported object types/properties and capabilities
        for key, act in actionTypes.items():
            
            act['type'] = key
            
            #if key != act['name']:
            #    raise Exception(f"Action key ({key}) contradicts its name ({act['name']})")
            
            # Check capabilities
            #if 'capabilities' in act:
            #    raise Exception(f"Action '{key}' is missing a capabilities list.")
            
            if 'capabilities' in act:
                
              for cap in act['capabilities']:
                if not cap in capabilities:
                    raise Exception(f"Action '{key}' capability '{cap}' is not in the global capability list.")
                
                # Could also check the sub-parameters of the capability
                #for cap_param in cap:
                #    if not cap_param in capabilities[cap['name']]:
                #        raise Exception(f"Action '{key}' capability '{cap['name']}' parameter '{cap_param}' is not in the global capability's parameter list.")
            
            if 'roles' in act:   # look through capabilities listed under each role
                for caps in act['roles'].values():
                    for cap in caps:
                        if not cap in capabilities:
                            raise Exception(f"Action '{key}' capability '{cap}' is not in the global capability list.")
                        
            
            # Check objects
            if not 'objects' in act:
                raise Exception(f"Action '{key}' is missing an objects list.")
            
            for obj in act['objects']:
                if not obj in objects:
                    raise Exception(f"Action '{key}' object '{obj}' is not in the global objects list.")
        
                # Could also check the sub-parameters of the object
                if isinstance(act['objects'], dict):  # if the object
                    for obj_param in act['objects'][obj]:
                        if not obj_param in objects[obj]:
                            raise Exception(f"Action '{key}' object '{obj}' parameter '{obj_param}' is not in the global object's parameter list.")
            
        
        # Store some things
        self.actionTypes  = actionTypes
        
        self.requirements  = requirements
        self.capabilities = capabilities  
        self.vessels      = vessels
        self.objects      = objects
        
        
        # Initialize some things
        self.actions = {}
        self.tasks = {}
        
        
    def registerAction(self, action):
        '''Registers an already created action'''
        
        # this also handles creation of unique dictionary keys
        
        if action.name in self.actions:  # check if there is already a key with the same name
            raise Warning(f"Action '{action.name}' is already registered.")
            print(f"Action name '{action.name}' is in the actions list so incrementing it...")
            action.name = increment_name(action.name)

        # What about handling of dependencies?? <<< done in the action object, 
        # but could check that each is in the list already...
        for dep in action.dependencies.values():
            if not dep in self.actions.values():
                raise Exception(f"New action '{action.name}' has a dependency '{dep.name}' this is not in the action list.")        
        
        # Check that all the requirements of all actions conform to the
        # options in requirements.yaml.
        for reqname, req in action.requirements.items():
            if req['base'] in self.requirements:  # ensure this requirement is listed
                for cap in req['capabilities']:
                    if not cap in self.capabilities:
                        raise Exception(f"Requirement '{reqname}' capability '{cap}' is not in the global capability list.")
            else:
                raise Exception(f"Action {action.name} requirement {req['base']} is not in requirements.yaml")

        # Add it to the actions dictionary
        self.actions[action.name] = action
        
        
    def addAction(self, action_type_name, action_name, **kwargs):
        '''Creates an action and adds it to the register'''
        
        if not action_type_name in self.actionTypes:
            raise Exception(f"Specified action type name {'action_type_name'} is not in the list of loaded action types.")
        
        # Get dictionary of action type information
        action_type = self.actionTypes[action_type_name]
        
        # Initialize full zero-valued dictionary of possible required capability specs
        reqs = {}  # Start a dictionary to hold the requirements -> capabilities -> specs
        for req in action_type['requirements']:
            
            # make sure it's a valid requirement
            if '-in' in req:  
                req_dir = 1  # this means the req is for storage and storage is being filled
                req_base = req[:-3]  # this is the name of the req as in requirements.yaml, no suffix
            elif '-out' in req:  
                req_dir = -1
                req_base = req[:-4]
            else:
                req_dir = 0
                req_base = req
            
            # Make sure the requirement and its direction are supported in the requirements yaml
            if not req_base in self.requirements:
                raise Exception(f"Requirement '{req_base}' is not in the requirements yaml.")
            if abs(req_dir) > 0 and ('directions' not in self.requirements[req_base] 
                                     or req_dir not in self.requirements[req_base]['directions']):
                raise Exception(f"Requirement '{req_base}' direction '{req_dir}' is not supported in the requirements yaml.")
            
            # Make the new requirements entry
            reqs[req] = {'base':req_base, 'direction':req_dir, 'capabilities':{}}
            
            # add the caps of the req
            for cap in self.requirements[req_base]['capabilities']: 
                reqs[req]['capabilities'][cap] = {}
                #print(f'   {cap}')
                # add the specs of the capability
                for spec in self.capabilities[cap]:
                    reqs[req]['capabilities'][cap][spec] = 0
                    #print(f'     {spec} = 0')
        # swap in the filled-out dict
        action_type['requirements'] = reqs
        
        # Create the action
        act = Action(action_type, action_name, display=self.display, **kwargs)        
        
        # Register the action
        self.registerAction(act)
        
        return act   # return the newly created action object, or its name?

    
    def addActionDependencies(self, action, dependencies):  
        '''Adds dependencies to an action, provided those dependencies have
        already been registered in the action list.
        '''
        
        if not isinstance(dependencies, list):
            dependencies = [dependencies]  # get into list form if singular
                
        for dep in dependencies:
            # Make sure the dependency is already registered
            if dep in self.actions.values():  
                action.addDependency(dep)
            else:
                raise Exception(f"New action '{action.name}' has a dependency '{dep.name}' this is not in the action list.")                  


    def visualizeActions(self):
        '''Generate a graph of the action dependencies.
        '''
        
        # Create the graph
        G = nx.DiGraph()
        for item, data in self.actions.items():
            for dep in data.dependencies:
                G.add_edge(dep, item, duration=data.duration)  # Store duration as edge attribute

        # Compute longest path & total duration
        longest_path = nx.dag_longest_path(G, weight='duration')
        longest_path_edges = list(zip(longest_path, longest_path[1:]))  # Convert path into edge pairs
        total_duration = sum(self.actions[node].duration for node in longest_path)
        if len(longest_path)>=1:
            last_node = longest_path[-1]  # Identify last node of the longest path
            # Define layout
            pos = nx.shell_layout(G)
            # Draw all nodes and edges (default gray)
            nx.draw(G, pos, with_labels=True, node_size=500, 
                    node_color='skyblue', font_size=10, font_weight='bold', 
                    font_color='black', edge_color='gray')

            # Highlight longest path in red
            nx.draw_networkx_edges(G, pos, edgelist=longest_path_edges, edge_color='red', width=2)

            # Annotate last node with total duration in red
            plt.text(pos[last_node][0], pos[last_node][1] - 0.1, f"{total_duration:.2f} hr", fontsize=12, color='red', fontweight='bold', ha='center')          
        else:
            pass
        plt.axis('equal')

        # Color first node (without dependencies) green
        i = 0
        for node in G.nodes():
            if G.in_degree(node) == 0:  # Check if the node has no incoming edges
                nx.draw_networkx_nodes(G, pos, nodelist=[node], node_color='green', node_size=500, label='Action starters' if i==0 else None)
                i += 1
        plt.legend()
        return G
    
    def visualizeActionsHierarchy(self):
        '''Generate a graph of the action dependencies with hierarchical layout.'''
        
        # Create the graph
        G = nx.DiGraph()
        for item, data in self.actions.items():
            for dep in data.dependencies:
                G.add_edge(dep, item, duration=data.duration)

        # Compute longest path & total duration
        longest_path = nx.dag_longest_path(G, weight='duration')
        longest_path_edges = list(zip(longest_path, longest_path[1:]))
        total_duration = sum(self.actions[node].duration for node in longest_path)
        
        if len(longest_path) >= 1:
            last_node = longest_path[-1]
            
            # Option A: Use multipartite layout (layers based on topological generations)
            for layer, nodes in enumerate(nx.topological_generations(G)):
                for node in nodes:
                    G.nodes[node]["layer"] = layer
            pos = nx.multipartite_layout(G, subset_key="layer", align='horizontal')
            
            # OR Option B: Use spring layout with specific parameters for hierarchy
            # pos = nx.spring_layout(G, k=2, iterations=50, seed=42)
            
            # OR Option C: Use Kamada-Kawai (good for hierarchical structure)
            # pos = nx.kamada_kawai_layout(G)
            
            plt.figure(figsize=(12, 8))
            
            # Draw all nodes and edges
            nx.draw(G, pos, with_labels=True, node_size=700,
                    node_color='skyblue', font_size=9, font_weight='bold',
                    font_color='black', edge_color='gray', arrows=True,
                    arrowsize=20, arrowstyle='->')

            # Highlight longest path
            nx.draw_networkx_edges(G, pos, edgelist=longest_path_edges, 
                                edge_color='red', width=3, arrows=True,
                                arrowsize=20)

            # Color start nodes green
            start_nodes = [node for node in G.nodes() if G.in_degree(node) == 0]
            nx.draw_networkx_nodes(G, pos, nodelist=start_nodes, 
                                node_color='green', node_size=700,
                                label='Start actions')
            
            # Color end nodes orange
            end_nodes = [node for node in G.nodes() if G.out_degree(node) == 0]
            nx.draw_networkx_nodes(G, pos, nodelist=end_nodes,
                                node_color='orange', node_size=700,
                                label='End actions')
            
            # Add edge labels with durations
            edge_labels = nx.get_edge_attributes(G, 'duration')
            edge_labels = {k: f"{v:.1f}h" for k, v in edge_labels.items()}
            nx.draw_networkx_edge_labels(G, pos, edge_labels, font_size=8)
            
            # Add title with critical path info
            plt.title(f"Action Dependency Graph\nCritical Path Duration: {total_duration:.2f} hr",
                    fontsize=14, fontweight='bold')
            
            plt.legend()
            plt.axis('off')
            plt.tight_layout()
            
        return G
    
    def registerTask(self, task):
        '''Registers an already created task'''
        
        # this also handles creation of unique dictionary keys
        
        if task.name in self.tasks:  # check if there is already a key with the same name
            raise Warning(f"Action '{task.name}' is already registered.")
            print(f"Task name '{task.name}' is in the tasks list so incrementing it...")
            task.name = increment_name(task.name)

        # Add it to the actions dictionary
        self.tasks[task.name] = task
        
        
    def addTask(self, task_name, actions, action_sequence, **kwargs):
        '''Creates a task and adds it to the register'''
        
        # Create the action
        task = Task(task_name, actions, action_sequence, **kwargs)        
        
        # Register the action
        self.registerTask(task)
        
        return task
    
    
    
    def findCompatibleVessels(self):
        '''Go through actions and identify which vessels have the required
        capabilities (could be based on capability presence, or quantitative.
        '''
        
        pass


    def figureOutTaskRelationships(self, time_interval=0.5):
        '''Calculate time constraints
        between tasks.
        '''
        
        # Figure out task durations (for a given set of asset assignments?)
        #for task in self.tasks.values():
            #task.calcTiming()
        
        # Figure out timing constraints between tasks based on action dependencies
        n = len(self.tasks)
        dt_min = np.zeros((n,n))  # matrix of required time offsets between tasks
        
        for i1, task1 in enumerate(self.tasks.values()):
            for i2, task2 in enumerate(self.tasks.values()):
                # look at all action dependencies from tasks 1 to 2 and
                # identify the limiting case (the largest time offset)...
                dt_min_1_2, dt_min_2_1 = findTaskDependencies(task1, task2, time_interval=time_interval)
                
                # for now, just look in one direction
                dt_min[i1, i2] = dt_min_1_2

        return dt_min
    

def findTaskDependencies(task1, task2, time_interval=0.5):
    '''Finds any time dependency between the actions of two tasks.
    Returns the minimum time separation required from task 1 to task 2,
    and from task 2 to task 1. I
    '''
    
    time_1_to_2 = []
    time_2_to_1 = []
    
    # Look for any dependencies where act2 depends on act1:
    #for i1, act1 in enumerate(task1.actions.values()):
    #    for i2, act2 in enumerate(task2.actions.values()):
    for a1, act1 in task1.actions.items():
        for a2, act2 in task2.actions.items():
        
            if a1 in act2.dependencies:  # if act2 depends on act1
                time_1_to_2.append(task1.actions_ti[a1] + act1.duration
                                   - task2.actions_ti[a2])
        
            if a2 in act1.dependencies:  # if act2 depends on act1
                time_2_to_1.append(task2.actions_ti[a2] + act2.duration
                                   - task1.actions_ti[a1])
    
    #print(time_1_to_2)
    #print(time_2_to_1)
    
    # TODO: provide cleaner handling of whether or not there is a time constraint in either direction <<<
    
    # Calculate minimum times (rounded to nearest interval)
    if time_1_to_2:
        raw_dt_min_1_2 = max(time_1_to_2)
        dt_min_1_2 = np.round(raw_dt_min_1_2 / time_interval) * time_interval
    else:
        dt_min_1_2 = -np.inf
    
    if time_2_to_1:
        raw_dt_min_2_1 = max(time_2_to_1)
        dt_min_2_1 = np.round(raw_dt_min_2_1 / time_interval) * time_interval
    else:
        dt_min_2_1 = -np.inf
    
    if dt_min_1_2 + dt_min_2_1 > 0:
        print(f"The timing between these two tasks seems to be impossible...")
    
    return dt_min_1_2, dt_min_2_1


def implementStrategy_staged(sc):
    '''This sets up Tasks in a way that implements a staged installation
    strategy where all of one thing is done before all of a next thing.
    '''
    
    # ----- Create a Task for all the anchor installs -----
    
    acts = []
    # first load each mooring
    for action in sc.actions.values():
        if action.type == 'load_anchor':
            acts.append(action)
    # then load each anchor
    for action in sc.actions.values():
        if action.type == 'install_anchor':
            acts.append(action)

    # create the task, passing in the sequence of actions
    sc.addTask('install_all_anchors', acts, action_sequence='series')
    
    
    # ----- Create a Task for all the mooring installs -----
    
    # gather the relevant actions
    acts = []
    # first load each mooring
    for action in sc.actions.values():
        if action.type == 'load_mooring':
            acts.append(action)
    # next lay each mooring (eventually route logic could be added)
    for action in sc.actions.values():
        if action.type == 'lay_mooring':
            acts.append(action)

    # create the task, passing in the sequence of actions
    sc.addTask('install_all_moorings', acts, action_sequence='series')
    
    
    # ----- Create a Task for the platform tow-out and hookup -----
    
    # gather the relevant actions
    acts = []
    # first tow out the platform
    acts.append(sc.actions['tow'])
    # next hook up each mooring
    for action in sc.actions.values():
        if action.type == 'mooring_hookup':
            acts.append(action)

    # create the task, passing in the sequence of actions
    sc.addTask('tow_and_hookup', acts, action_sequence='series')



def create_gantt_excel(tasks, scheduler_result, weather=None, filename='gantt_chart.xlsx'):
    """
    Creates a Gantt chart Excel file from scheduler results.
    
    Parameters:
    - tasks: Dictionary of tasks in the Scenario object
    - scheduler_result: Result object from the scheduler optimization
    - weather: Weather data list (optional, for coloring bad weather periods)
    - filename: Output Excel filename
    """

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"
    
    # Define colors
    task_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Task blue
    action_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")  # Action green
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Header light blue
    bad_weather_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Light pink for bad weather
    
    # Define fonts
    header_font = Font(color="000000", bold=True, size=10)  # Black text on light blue
    task_font = Font(bold=True, size=11)
    action_font = Font(size=10)
    cell_font = Font(color="000000", bold=True, size=9)  # Black text for cells
    
    # Only border the headers - no borders for data cells
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Extract scheduler results
    if scheduler_result.get('success', False):
        Xta = scheduler_result['Xta']  # Task-Asset assignments
        Xtp = scheduler_result['Xtp']  # Task-Period assignments  
        Xts = scheduler_result['Xts']  # Task-Start assignments
        period_duration = scheduler_result.get('period_duration', 0.25)
        max_periods = scheduler_result.get('num_periods', 50)
        
        # Create task schedule mapping
        task_schedule = {}
        for t, task_name in enumerate(scheduler_result['tasks']):
            # Find assigned asset group
            assigned_asset_group_idx = np.argmax(Xta[t, :]) if Xta[t, :].sum() > 0 else 0
            asset_group_info = scheduler_result['asset_groups'][assigned_asset_group_idx]
            
            # Extract asset group name and assets
            if isinstance(asset_group_info, dict):
                group_name = list(asset_group_info.keys())[0]
                asset_list = asset_group_info[group_name]
                if isinstance(asset_list, list):
                    asset_group_display = f"{', '.join(asset_list)}"
                else:
                    asset_group_display = str(group_name)
            else:
                asset_group_display = f"Group_{assigned_asset_group_idx}"
            
            # Find start time and duration
            start_time = np.argmax(Xts[t, :]) if Xts[t, :].sum() > 0 else 0
            active_periods = np.where(Xtp[t, :] > 0)[0]
            duration = len(active_periods) if len(active_periods) > 0 else 1
            
            task_schedule[task_name] = {
                'order': t,
                'start_time': int(start_time),
                'duration': int(duration),
                'asset_group': asset_group_display,
                'active_periods': active_periods
            }
    else:
        # Fallback if scheduler failed
        max_periods = 50
        period_duration = 0.25
        task_schedule = {}
        for task_name in tasks.keys():
            task_schedule[task_name] = {
                'order': t,
                'start_time': 0,
                'duration': 1, 
                'asset_group': 'Unknown',
                'active_periods': [0]
            }
    
    # Create headers - Name column + Period columns (no text for periods)
    headers = ['Name'] + ['' for i in range(max_periods)]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply weather-based coloring to period columns (skip first column which is 'Name')
        if col > 1 and weather is not None:
            period_index = col - 2  # Convert to 0-based period index
            if period_index < len(weather):
                if weather[period_index] != 1:  # Bad weather
                    cell.fill = bad_weather_fill
                else:  # Good weather
                    cell.fill = header_fill
            else:
                cell.fill = header_fill
        else:
            cell.fill = header_fill
    
    # Set row height for header
    ws.row_dimensions[1].height = 12
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Name column (adjusted for Excel display)
    for i in range(2, len(headers) + 1):  # Period columns
        ws.column_dimensions[get_column_letter(i)].width = 1.3
    
    current_row = 2
    
    # Sort tasks by start time if available
    sorted_tasks = list(tasks.items())
    if task_schedule:
        sorted_tasks.sort(key=lambda x: task_schedule.get(x[0], {}).get('order', 0))
    
    # Process each task
    for task_name, task in sorted_tasks:
        task_info = task_schedule.get(task_name, {})
        start_period = task_info.get('start_time', 0)
        duration = task_info.get('duration', 1)
        asset_group = task_info.get('asset_group', 'Unknown')
        active_periods = task_info.get('active_periods', range(start_period, start_period + duration))
        
        # Write task row
        task_cell = ws.cell(row=current_row, column=1, value=task_name)
        task_cell.font = task_font
        task_cell.alignment = Alignment(vertical='center')
        
        # Set row height for task row
        ws.row_dimensions[current_row].height = 12
        
        # Task Gantt bars (blue) - only for active periods
        if len(active_periods) > 0:
            # Create individual cells first
            task_cells = []
            for period in active_periods:
                if period < max_periods:
                    bar_cell = ws.cell(row=current_row, column=period + 2, value='')
                    bar_cell.fill = task_fill
                    task_cells.append((current_row, period + 2))
            
            # Merge cells if there are multiple periods for this task
            if len(task_cells) > 1:
                start_col = task_cells[0][1]
                end_col = task_cells[-1][1]
                try:
                    ws.merge_cells(start_row=current_row, start_column=start_col, 
                                    end_row=current_row, end_column=end_col)
                    # Add asset group name in the merged cell
                    merged_cell = ws.cell(row=current_row, column=start_col)
                    merged_cell.value = asset_group
                    merged_cell.fill = task_fill
                    merged_cell.font = Font(color="FFFFFF", bold=True, size=11)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                except:
                    # If merge fails, just put text in first cell
                    first_cell = ws.cell(row=current_row, column=start_col)
                    first_cell.value = asset_group
                    first_cell.font = Font(color="FFFFFF", bold=True, size=8)
                    first_cell.alignment = Alignment(horizontal='center', vertical='center')
            elif len(task_cells) == 1:
                # Single cell - just add the text
                single_cell = ws.cell(row=current_row, column=task_cells[0][1])
                single_cell.value = asset_group
                single_cell.font = Font(color="FFFFFF", bold=True, size=8)
                single_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        
        # Process actions within this task (with L-shaped indentation)
        if hasattr(task, 'actions') and task.actions:
            # Sort actions by their timing within the task
            sorted_actions = list(task.actions.items())
            if hasattr(task, 'actions_ti'):
                sorted_actions.sort(key=lambda x: task.actions_ti.get(x[0], 0))
            
            # Calculate action durations and adjust them to fit exactly within task duration
            task_duration_periods = len(active_periods)
            action_durations = []
            total_raw_duration = 0
            
            # First pass: calculate raw durations
            for action_name, action in sorted_actions:
                if hasattr(action, 'duration'):
                    raw_duration = max(1, int(action.duration / period_duration))
                else:
                    raw_duration = 1
                action_durations.append(raw_duration)
                total_raw_duration += raw_duration
            
            # Adjust durations to fit exactly within task duration
            if total_raw_duration != task_duration_periods and total_raw_duration > 0:
                # Proportionally scale each action duration
                adjusted_durations = []
                cumulative_adjusted = 0
                
                for i, raw_duration in enumerate(action_durations):
                    if i == len(action_durations) - 1:
                        # Last action gets remaining periods to ensure exact fit
                        adjusted_duration = task_duration_periods - cumulative_adjusted
                    else:
                        # Proportional scaling for other actions
                        adjusted_duration = max(1, int(raw_duration * task_duration_periods / total_raw_duration))
                    
                    adjusted_durations.append(max(1, adjusted_duration))  # Ensure minimum 1 period
                    cumulative_adjusted += adjusted_durations[-1]
                
                action_durations = adjusted_durations
            
            # Second pass: create action rows with adjusted durations
            action_start_period = start_period
            for i, (action_name, action) in enumerate(sorted_actions):
                # Create L-shaped symbol for hierarchy - all actions get L-shape
                indent_symbol = "└─ "  # All actions get the L-shape
                
                action_display_name = f"{indent_symbol}{action_name}"
                
                # Write action row
                action_cell = ws.cell(row=current_row, column=1, value=action_display_name)
                action_cell.font = action_font
                action_cell.alignment = Alignment(vertical='center')
                
                # Set row height for action row
                ws.row_dimensions[current_row].height = 12
                
                # Use adjusted duration
                action_duration = action_durations[i] if i < len(action_durations) else 1
                
                # Action Gantt bars (green) - sequential within task periods
                action_end = min(action_start_period + action_duration, start_period + task_duration_periods)
                for period in range(action_start_period, action_end):
                    if period < max_periods and period in active_periods:
                        bar_cell = ws.cell(row=current_row, column=period + 2, value='')
                        bar_cell.fill = action_fill
                
                # Next action starts where this one ended
                action_start_period = action_end
                current_row += 1
    
    # Freeze panes to keep headers and name column visible
    ws.freeze_panes = ws['B2']
    
    # Set zoom level to 150%
    ws.sheet_view.zoomScale = 150
    
    # Set print settings for fitting on page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically if needed
    
    # Set margins to minimum
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # Save the workbook
    wb.save(filename)
    print(f"Gantt chart saved to {filename}")
    
    return wb








if __name__ == '__main__':
    '''This is currently a script to explore how some of the workflow could
    work. Can move things into functions/methods as they solidify.
    '''
    
    # ----- Load up a Project -----    
    
    from famodel.project import Project


    print('Creating project without RAFT\n')
    print(os.getcwd())
    # create project object
    #project = Project(file='../../examples/OntologySample200m_1turb.yaml', raft=False)
    project = Project(file='../../examples/OntologySample_RM3.yaml', raft=False)
    # create moorpy system of the array, include cables in the system
    project.getMoorPyArray(cables=1)
    # plot in 3d, using moorpy system for the mooring and cable plots
    #project.plot2d()
    #project.plot3d()

    '''
    # project.arrayWatchCircle(ang_spacing=20)
    # save envelopes from watch circle information for each mooring line
    for moor in project.mooringList.values():
        moor.getEnvelope()

    # plot motion envelopes with 2d plot
    project.plot2d(save=True, plot_bathymetry=False)
    '''
    
    # Tally up some object properties (eventually make this built-in Project stuff)
    for mooring in project.mooringList.values():
        # sum up mooring quantities of interest
        L = 0  # length
        m = 0  # mass
        V = 0  # volume

        for sec in mooring.sections(): # add up the length of all sections in the mooring
            L += sec['L']
            m += sec['L'] * sec['type']['m']
            V += sec['L'] * np.pi/4 * sec['type']['d_vol']**2
        
        mooring.props = {}
        mooring.props['length'] = L
        mooring.props['pretension'] = 0   # <<< get this from MoorPy once this is moved into Mooring class?
        mooring.props['weight'] = 9.8*(m - 1025*V)
        mooring.props['mass'] = m
        mooring.props['volume'] = V
    
    for anchor in project.anchorList.values():

        anchor.props = {}

        if anchor.dd['design']['type'] == 'suction':
            L = anchor.dd['design']['L']    # [m]
            D = anchor.dd['design']['D']    # [m]
            t = (6.35 + 20*D)/1000          # [m]
            anchor.props['mass'] = 7850*np.pi*L * (D*t - t**2) + 7850*np.pi * ( (D**2*t/4) - (D*t**2) + t**3)


        

    

    display = 1

    sc = Scenario(display=display)  # class instance holding most of the info
                
    
    # ----- Create the interrelated actions (including their individual requirements) -----
    print("===== Create Actions =====")
    
    # add actions for each anchor object
    for akey, anchor in project.anchorList.items():
        
        # load each anchor
        a0 = sc.addAction('load_anchor', f'load_anchor-{akey}', objects=[anchor])

        # install each anchor
        a1 = sc.addAction('install_anchor', f'install_anchor-{akey}', objects=[anchor])
        sc.addActionDependencies(a1, [a0])
        
        # register the actions as necessary for the anchor <<< do this for all objects??
        anchor.install_dependencies = [a1]
    
    
    hookups = []  # list of hookup actions
    
    # add actions for each mooring object
    for mkey, mooring in project.mooringList.items():

        # create load mooring action
        a2 = sc.addAction('load_mooring', f'load_mooring-{mkey}', objects=[mooring])
        
        # create lay mooring action
        a3 = sc.addAction('lay_mooring', f'lay_mooring-{mkey}', objects=[mooring], dependencies=[a2])
        sc.addActionDependencies(a3, mooring.attached_to[0].install_dependencies)
        
        # create hookup action
        a4 = sc.addAction('mooring_hookup', f'mooring_hookup-{mkey}', 
                          objects=[mooring, mooring.attached_to[1]], dependencies=[a3])
        
        hookups.append(a4)
        
    
    # add the FOWT install action
    a5 = sc.addAction('tow', 'tow', objects=[list(project.platformList.values())[0]])
    for a in hookups:
        sc.addActionDependencies(a, [a5])  # make each hookup action dependent on the FOWT being towed out
    

    # ----- Do some graph analysis -----
    
    #G = sc.visualizeActions()
    #G = sc.visualizeActionsHierarchy()
    

    # ----- Generate tasks (sequences of Actions following specific strategies) -----
    print("===== Generate Tasks =====")

    # Call one of the task strategy implementers, which will create the tasks
    implementStrategy_staged(sc)
    
    time_interval = 0.25
    
    # ----- Try assigning assets to the tasks -----
    print('===== Assigning Asset Groups to Tasks =====')
    for task in sc.tasks.values():
        print(f"--- Looking at task {task.name} ---")
        if task.checkAssets([sc.vessels['AHTS_alpha']], display=1)[0]:
            print('Assigned AHTS')
            task.assignAssets([sc.vessels['AHTS_alpha']])
        elif task.checkAssets([sc.vessels['MPSV_01']], display=1)[0]:
            print('Assigned MPSV')
            task.assignAssets([sc.vessels['MPSV_01']])
        elif task.checkAssets([sc.vessels['AHTS_alpha'], sc.vessels['MPSV_01']], display=1):
            print('Assigned AHTS and MPSV')
            task.assignAssets([sc.vessels['AHTS_alpha'], sc.vessels['MPSV_01']])
        else:
            print('Something is wrong')
        
        
        # Calculation durations of the actions, and then of the task
        for a in task.actions.values():
            a.calcDurationAndCost()
        task.calcDuration(time_interval=time_interval)
    
   
    # Example task time adjustment and plot
    #sc.tasks['tow_and_hookup'].setStartTime(5)
    #sc.tasks['tow_and_hookup'].chart()


    # ----- Quantify the dependency offsets between tasks -----
    
    dt_min = sc.figureOutTaskRelationships(time_interval=time_interval)
    
    

    # ----- Call the scheduler -----

    # TASKS
    tasks_scheduler = list(sc.tasks.keys())
    
    # ASSETS
    for asset in sc.vessels.values():
        if 'name' in asset:
            if asset['name'] == 'AHTS_alpha':
                asset['max_weather'] = 2
            elif asset['name'] == 'MPSV_01':
                asset['max_weather'] = 1
        else:
            asset['max_weather'] = asset['transport']['Hs_m']
    assets_scheduler = list(sc.vessels.values())

    # ASSET GROUPS
    # >>>>> TODO: make this automated to find all possible combinations of "realistic" asset groups
    asset_groups_scheduler = [
        {'group1': ['AHTS_alpha']},
        {'group2': ['MPSV_01']},
        {'group3': ['AHTS_alpha', 'MPSV_01']}
    ]

    # TASK-ASSET-MATRIX
    task_asset_matrix_scheduler = np.zeros([len(tasks_scheduler), len(asset_groups_scheduler), 2])
    for i,task in enumerate(sc.tasks.values()):
        for j,asset_group in enumerate(asset_groups_scheduler):
            # Extract asset list from the dictionary - values() returns a list containing one list
            asset_names = list(asset_group.values())[0]
            asset_list = [sc.vessels[asset_name] for asset_name in asset_names]
            #task.checkAssets([sc.vessels['AHTS_alpha'], sc.vessels['HL_Giant'], sc.vessels['CSV_A']], display=1)
            if not task.checkAssets(asset_list, display=0)[0]:
                task_asset_matrix_scheduler[i,j] = (-1, -1)
            else:
                task.assignAssets(asset_list)
                for a in task.actions.values():
                    a.calcDurationAndCost()
                task.calcDuration(time_interval=time_interval)
                task.calcCost()
                duration_int = int(np.ceil(task.duration / time_interval))
                task_asset_matrix_scheduler[i,j] = (task.cost, duration_int)
                task.clearAssets()
    

    # DEPENDENCIES

    # Create dependency offset matrices for each pair of tasks (not the same task in pairs)
    # All rows represent the dependent task and all columns represent the prerequisite task
    # Each individual row or column represents a different asset group
    # Values in the matrices are the minimum time offset required between those tasks with those asset assignments
    # For example, if there are 3 tasks and 3 asset groups:
    #     if task2 depends on task1, if task2 uses the second asset group, and if task1 uses the third asset group,
    #     then there will be a non-inf value in the [1,2] (python-indexing) spot for the time offset
    
    dependency_offset_matrices = {}
    task_list = list(sc.tasks.values())
    task_names = list(sc.tasks.keys())
    
    # Initialize matrices for each dependency pair ("dependent_task->prerequisite_task")
    for i, task1 in enumerate(task_list):
        for k, task2 in enumerate(task_list):
            if i != k:  # Don't compare a task with itself
                dep_key = f"{task_names[k]}->{task_names[i]}"  # "task2->task1" = "dependent->prerequisite"
                dependency_offset_matrices[dep_key] = np.full((len(asset_groups_scheduler), len(asset_groups_scheduler)), -np.inf)
    
    # Calculate offsets for each valid task-asset-group combination
    for i, task1 in enumerate(task_list):
        for j, ag1 in enumerate(asset_groups_scheduler):
            # Get the list of vessels for this asset group
            asset_list1 = [sc.vessels[name] for name in list(ag1.values())[0]]
            
            # Check if this asset group can perform task1
            if task1.checkAssets(asset_list1, display=0)[0]:
                # Temporarily assign assets and calculate durations
                task1.assignAssets(asset_list1)
                for a in task1.actions.values():
                    a.calcDurationAndCost()
                task1.calcDuration(time_interval=time_interval)
                
                # Check dependencies with all other tasks
                for k, task2 in enumerate(task_list):
                    if i == k:  # Skip self-comparison
                        continue
                        
                    for l, ag2 in enumerate(asset_groups_scheduler):
                        asset_list2 = [sc.vessels[name] for name in list(ag2.values())[0]]
                        
                        # Check if this asset group can perform task2
                        if task2.checkAssets(asset_list2, display=0)[0]:
                            # Temporarily assign assets and calculate durations
                            task2.assignAssets(asset_list2)
                            for a in task2.actions.values():
                                a.calcDurationAndCost()
                            task2.calcDuration(time_interval=time_interval)
                            
                            # Calculate the minimum time offset from task1 to task2
                            dt_min_1_to_2, dt_min_2_to_1 = findTaskDependencies(task1, task2, time_interval=time_interval)
                            
                            # Store in the appropriate matrix (and convert from hours to periods)
                            dep_key_k_to_i = f"{task_names[k]}->{task_names[i]}"
                            if dt_min_1_to_2 != -np.inf:
                                dependency_offset_matrices[dep_key_k_to_i][l, j] = int(np.ceil(dt_min_1_to_2 / time_interval))
                            else:
                                dependency_offset_matrices[dep_key_k_to_i][l, j] = -np.inf
                            
                            dep_key_i_to_k = f"{task_names[i]}->{task_names[k]}"
                            if dt_min_2_to_1 != -np.inf:
                                dependency_offset_matrices[dep_key_i_to_k][j, l] = int(np.ceil(dt_min_2_to_1 / time_interval))
                            else:
                                dependency_offset_matrices[dep_key_i_to_k][j, l] = -np.inf
                            
                            task2.clearAssets()
                
                task1.clearAssets()
    
    

    # Calculate task_dependencies and dependency_types from the dependency_offset_matrices
    
    task_dependencies = {}
    dependency_types = {}
    
    for dep_key, matrix in dependency_offset_matrices.items():
        # Check if there are any valid offsets (non -inf) in this matrix
        if np.any(matrix != -np.inf):
            # Parse the dependency key: "dependent_task->prerequisite_task"
            dependent_task, prerequisite_task = dep_key.split('->')
            
            # Add to task_dependencies
            if dependent_task not in task_dependencies:
                task_dependencies[dependent_task] = []
            task_dependencies[dependent_task].append(prerequisite_task)
            
            # Set dependency type based on task characteristics (default is 'start_start') - can update this later with other dependency types
            dependency_types[dep_key] = 'start_start'




    # PERIODS
    # Calculate total number of periods needed for the scheduler
    # For each task, find the maximum duration across all valid asset group assignments
    total_task_duration = 0
    for i, task in enumerate(sc.tasks.values()):
        # Get durations for this task across all asset groups (index 1 is duration in periods)
        task_durations = task_asset_matrix_scheduler[i, :, 1]
        # Filter out invalid assignments (negative or zero duration)
        valid_durations = task_durations[task_durations > 0]
        if len(valid_durations) > 0:
            max_duration = np.max(valid_durations)
            total_task_duration += max_duration
    # Find maximum offset from all dependency matrices (ignoring -inf values and negative values)
    max_offsets_sum = 0
    for dep_key, matrix in dependency_offset_matrices.items():
        valid_offsets = matrix[matrix != -np.inf]
        if len(valid_offsets) > 0 and max(valid_offsets) >= 0:
            max_offsets_sum += np.max(valid_offsets)

    total_duration = total_task_duration + max_offsets_sum
    # total_duration is already in periods, so just convert to integer
    num_periods = int(np.ceil(total_duration))
    
    

    

    weather = [int(x) for x in np.ones(num_periods, dtype=int)]
    for i in range(6):
        weather[i] = 2
    for i in range(35, 50):
        weather[i] = 2


    scheduler = Scheduler(
        tasks=tasks_scheduler,
        assets=assets_scheduler,
        asset_groups=asset_groups_scheduler,
        task_asset_matrix=task_asset_matrix_scheduler,
        task_dependencies=task_dependencies,  # Derived from dependency_offset_matrices
        dependency_types=dependency_types,     # Derived from dependency_offset_matrices
        offsets=dependency_offset_matrices,    # Use asset-group-specific matrices
        weather=weather,
        period_duration=time_interval,
        wordy=1
    )

    scheduler.set_up_optimizer()

    result = scheduler.optimize()
    
    wb = create_gantt_excel(sc.tasks, result, weather, 'installation_gantt_chart.xlsx')

    a = 2